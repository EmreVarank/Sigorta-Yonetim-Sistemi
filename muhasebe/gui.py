#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AXA Muhasebe ve Komisyon Takip Sistemi - GUI
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
import sys
import sqlite3
import pandas as pd
from datetime import datetime
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), 'parsers'))

from commission import (
    get_application_path,
    PARSER_AVAILABLE,
    COLORS,
    AYLAR,
    KISILER,
    TURLER,
    ODEME_TURLERI,
    format_turkish_currency,
    is_iptal_police,
    process_cari_pdf,
    get_cari_data_for_month,
    process_pdf,
    hesapla_komisyon,
    parse_turkish_amount
)
from database import MuhasebeDB


class MuhasebeGUI:
    """Ana GUI Sınıfı"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("AXA Muhasebe ve Komisyon Takip Sistemi")
        self.root.geometry("1500x850")
        
        # Veritabanı
        self.db = MuhasebeDB()
        
        # Mevcut seçimler
        self.current_kisi = "YAŞAR"
        self.current_yil = datetime.now().year
        self.current_ay = AYLAR[datetime.now().month - 1]  # Şu anki ay
        
        # UI oluştur
        self.create_ui()
        
        # Başlangıç
        self.log("✅ Uygulama başlatıldı", "success")
        self.load_data()
        self.update_summary()
    
    def create_ui(self):
        """Kullanıcı arayüzünü oluştur"""
        self.create_menu()
        self.create_header()
        self.create_tabs()
        self.create_content()
        self.create_status_bar()
    
    def create_menu(self):
        """Menü çubuğu"""
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        
        # Dosya
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Dosya", menu=file_menu)
        file_menu.add_command(label="PDF Klasörü İşle", command=self.process_pdf_folder)
        file_menu.add_command(label="Excel'den İçe Aktar", command=self.import_excel_data)
        file_menu.add_command(label="Excel'e Aktar", command=self.export_to_excel)
        file_menu.add_separator()
        file_menu.add_command(label="Çıkış", command=self.root.quit)
        
        # Düzenle
        edit_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Düzenle", menu=edit_menu)
        edit_menu.add_command(label="Yeni Kayıt Ekle", command=self.add_manual_entry)
        edit_menu.add_command(label="Seçili Kaydı Sil", command=self.delete_selected)
        
        # Görünüm
        view_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Görünüm", menu=view_menu)
        view_menu.add_command(label="Yenile", command=self.load_data)
        view_menu.add_command(label="Cari Görüntüle", command=self.show_cari_window)
        view_menu.add_command(label="Logları Temizle", command=self.clear_logs)
        
        # Yardım
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Yardım", menu=help_menu)
        help_menu.add_command(label="Hakkında", command=self.show_about)
    
    def create_header(self):
        """Başlık bölümü"""
        header = tk.Frame(self.root, bg=COLORS['primary'], height=40)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        
        title = tk.Label(
            header,
            text="AXA MUHASEBE VE KOMİSYON TAKİP SİSTEMİ",
            font=('Segoe UI', 13, 'bold'),
            bg=COLORS['primary'],
            fg='#e2e8f0'
        )
        title.pack(pady=8)
    
    def create_tabs(self):
        """Kişi tab'ları"""
        tab_frame = tk.Frame(self.root, bg='#1e293b')
        tab_frame.pack(fill=tk.X)
        
        self.tab_buttons = {}
        for kisi in KISILER:
            btn = tk.Button(
                tab_frame,
                text=kisi,
                font=('Segoe UI', 9, 'bold'),
                bg='#0d9488' if kisi == self.current_kisi else '#1e293b',
                fg='white' if kisi == self.current_kisi else '#94a3b8',
                relief=tk.FLAT,
                padx=20,
                pady=4,
                cursor='hand2',
                command=lambda k=kisi: self.switch_tab(k)
            )
            btn.pack(side=tk.LEFT, padx=1, pady=2)
            self.tab_buttons[kisi] = btn
        
        # ============ AY SEKMELERİ ============
        self.create_month_tabs()
    
    def create_month_tabs(self):
        """Ay sekmeleri - kişi sekmelerinin altında"""
        month_frame = tk.Frame(self.root, bg='#0f172a')
        month_frame.pack(fill=tk.X)
        
        # Sol ok butonu
        left_arrow = tk.Button(
            month_frame,
            text="<",
            font=('Segoe UI', 10, 'bold'),
            bg='#0f172a',
            fg='#64748b',
            relief=tk.FLAT,
            padx=5,
            cursor='hand2',
            command=self.scroll_months_left
        )
        left_arrow.pack(side=tk.LEFT, padx=2)
        
        # Sağ ok butonu
        right_arrow = tk.Button(
            month_frame,
            text=">",
            font=('Segoe UI', 10, 'bold'),
            bg='#0f172a',
            fg='#64748b',
            relief=tk.FLAT,
            padx=5,
            cursor='hand2',
            command=self.scroll_months_right
        )
        right_arrow.pack(side=tk.LEFT, padx=2)
        
        # Ay butonları için frame - expand ile tam genişlik
        self.months_container = tk.Frame(month_frame, bg='#0f172a')
        self.months_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        self.month_buttons = {}
        for ay in AYLAR:
            btn = tk.Button(
                self.months_container,
                text=ay,
                font=('Segoe UI', 9, 'bold'),
                bg='#0f172a' if ay != self.current_ay else '#0d9488',
                fg='#64748b' if ay != self.current_ay else 'white',
                relief=tk.FLAT,
                pady=4,
                cursor='hand2',
                command=lambda a=ay: self.switch_month(a)
            )
            btn.pack(side=tk.LEFT, padx=1, pady=3, fill=tk.X, expand=True)
            self.month_buttons[ay] = btn
    
    def switch_month(self, ay):
        """Ay sekmesi değiştir"""
        self.current_ay = ay
        
        # Buton görünümlerini güncelle
        for a, btn in self.month_buttons.items():
            if a == ay:
                btn.config(bg='#0d9488', fg='white')
            else:
                btn.config(bg='#0f172a', fg='#64748b')
        
        self.load_data()
        self.update_summary()
        self.log(f"📅 {ay} ayına geçildi", "info")
    
    def scroll_months_left(self):
        """Ay listesinde sola kay"""
        current_idx = AYLAR.index(self.current_ay)
        if current_idx > 0:
            self.switch_month(AYLAR[current_idx - 1])
    
    def scroll_months_right(self):
        """Ay listesinde sağa kay"""
        current_idx = AYLAR.index(self.current_ay)
        if current_idx < len(AYLAR) - 1:
            self.switch_month(AYLAR[current_idx + 1])
    
    def switch_tab(self, kisi):
        """Tab değiştir"""
        self.current_kisi = kisi
        
        # Buton görünümlerini güncelle
        for k, btn in self.tab_buttons.items():
            if k == kisi:
                btn.config(bg='#0d9488', fg='white')
            else:
                btn.config(bg='#1e293b', fg='#94a3b8')
        
        self.load_data()
        self.update_summary()
        self.log(f"📂 {kisi} sekmesine geçildi", "info")
    
    def create_content(self):
        """Ana içerik alanı"""
        main_frame = tk.Frame(self.root, bg=COLORS['bg_light'])
        main_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        
        # Sol panel
        left_panel = tk.Frame(main_frame, bg=COLORS['card'], width=170)
        left_panel.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 8))
        left_panel.pack_propagate(False)
        
        self.create_filter_panel(left_panel)
        self.create_buttons(left_panel)
        self.create_log_panel(left_panel)  # Log paneli sol panele taşındı
        
        # Sağ panel
        right_panel = tk.Frame(main_frame, bg=COLORS['bg_light'])
        right_panel.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        self.create_table(right_panel)
    
    def create_filter_panel(self, parent):
        """Filtre paneli"""
        filter_frame = tk.LabelFrame(
            parent,
            text=" FİLTRELER",
            font=('Segoe UI', 9, 'bold'),
            bg=COLORS['card'],
            fg=COLORS['primary_accent']
        )
        filter_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # Sigortalı
        tk.Label(filter_frame, text="Sigortalı:", font=('Segoe UI', 8), bg=COLORS['card'], fg='#475569').pack(anchor='w', padx=5, pady=(5, 0))
        self.filter_sigortali = tk.Entry(filter_frame, font=('Arial', 8))
        self.filter_sigortali.pack(fill=tk.X, padx=5, pady=(1, 3))
        
        # Poliçe No
        tk.Label(filter_frame, text="Poliçe No:", font=('Segoe UI', 8), bg=COLORS['card'], fg='#475569').pack(anchor='w', padx=5, pady=(2, 0))
        self.filter_police_no = tk.Entry(filter_frame, font=('Arial', 8))
        self.filter_police_no.pack(fill=tk.X, padx=5, pady=(1, 3))
        
        # Plaka
        tk.Label(filter_frame, text="Plaka:", font=('Segoe UI', 8), bg=COLORS['card'], fg='#475569').pack(anchor='w', padx=5, pady=(2, 0))
        self.filter_plaka = tk.Entry(filter_frame, font=('Arial', 8))
        self.filter_plaka.pack(fill=tk.X, padx=5, pady=(1, 3))
        
        # Tür
        tk.Label(filter_frame, text="Tür:", font=('Segoe UI', 8), bg=COLORS['card'], fg='#475569').pack(anchor='w', padx=5, pady=(2, 0))
        self.filter_tur = ttk.Combobox(filter_frame, values=['Tümü'] + TURLER, font=('Arial', 8), state='readonly')
        self.filter_tur.set('Tümü')
        self.filter_tur.pack(fill=tk.X, padx=5, pady=(1, 3))
        
        # İptal Checkbox
        self.filter_iptal_var = tk.BooleanVar(value=False)
        self.filter_iptal_cb = tk.Checkbutton(
            filter_frame,
            text="Sadece İptaller",
            variable=self.filter_iptal_var,
            font=('Segoe UI', 8, 'bold'),
            bg=COLORS['card'],
            fg='#dc2626',
            activebackground=COLORS['card'],
            cursor='hand2'
        )
        self.filter_iptal_cb.pack(anchor='w', padx=5, pady=(5, 3))
        
        # Buton frame
        btn_frame = tk.Frame(filter_frame, bg=COLORS['card'])
        btn_frame.pack(fill=tk.X, padx=5, pady=(5, 5))
        
        # Ara butonu
        search_btn = tk.Button(
            btn_frame,
            text="ARA",
            command=self.apply_filters,
            font=('Segoe UI', 8, 'bold'),
            bg=COLORS['primary_accent'],
            fg='white',
            relief=tk.FLAT,
            cursor='hand2',
            pady=4
        )
        search_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 2))
        
        # Temizle butonu
        clear_btn = tk.Button(
            btn_frame,
            text="↻",
            command=self.clear_filters,
            font=('Segoe UI', 10, 'bold'),
            bg='#475569',
            fg='white',
            relief=tk.FLAT,
            cursor='hand2',
            pady=2
        )
        clear_btn.pack(side=tk.LEFT, padx=(2, 0))
    
    def create_buttons(self, parent):
        """Butonlar"""
        btn_frame = tk.LabelFrame(
            parent,
            text=" İŞLEMLER",
            font=('Segoe UI', 9, 'bold'),
            bg=COLORS['card']
        )
        btn_frame.pack(fill=tk.X, padx=5, pady=5)
        
        buttons = [
            ("PDF İşle", self.process_pdf_folder, '#0d9488'),
            ("Exceli Aç", self.import_excel_data, '#6366f1'),
            ("Excel'e Aktar", self.export_to_excel, '#d97706'),
            ("Ayı Sil", self.delete_current_month, COLORS['danger']),
            ("Seneyi Sil", self.delete_current_year, '#7f1d1d')
        ]
        
        for text, command, color in buttons:
            btn = tk.Button(
                btn_frame,
                text=text,
                command=command,
                font=('Segoe UI', 8, 'bold'),
                bg=color,
                fg='white',
                relief=tk.FLAT,
                cursor='hand2',
                padx=5,
                pady=4
            )
            btn.pack(fill=tk.X, padx=5, pady=2)
    
    def create_summary_panel(self, parent):
        """Özet paneli"""
        summary_frame = tk.LabelFrame(
            parent,
            text=" ÖZET",
            font=('Segoe UI', 9, 'bold'),
            bg=COLORS['card']
        )
        summary_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.summary_text = tk.Text(
            summary_frame,
            font=('Consolas', 8),
            bg='#f8fafc',
            fg='#1e293b',
            relief=tk.FLAT,
            wrap=tk.WORD
        )
        self.summary_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
    
    def create_table(self, parent):
        """İkili kayıt tablosu - Sol: AXA, Sağ: Diğer Şirketler"""
        self.table_frame = tk.Frame(parent, bg=COLORS['card'])
        self.table_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 8))
        
        # İki panel için sabit container - tam %50-%50 bölünmüş
        tables_container = tk.Frame(self.table_frame, bg=COLORS['border'])
        tables_container.pack(fill=tk.BOTH, expand=True)
        tables_container.grid_columnconfigure(0, weight=1, uniform="half")
        tables_container.grid_columnconfigure(1, weight=1, uniform="half")
        tables_container.grid_rowconfigure(0, weight=1)
        
        # ============ SOL PANEL: AXA POLİÇELERİ ============
        left_frame = tk.Frame(tables_container, bg='#f0fdf4')
        left_frame.grid(row=0, column=0, sticky='nsew', padx=(0, 1))
        
        # Sol panel başlık ve + butonu
        left_header = tk.Frame(left_frame, bg='#f0fdf4')
        left_header.pack(fill=tk.X)
        
        tk.Label(left_header, text="AXA POLİÇELERİ", font=('Segoe UI', 10, 'bold'),
                 bg='#f0fdf4', fg='#166534').pack(side=tk.LEFT, padx=5, pady=2)
        
        self.btn_add_axa = tk.Button(
            left_header, text="+", font=('Segoe UI', 12, 'bold'),
            bg='#059669', fg='white', width=2, cursor='hand2',
            command=lambda: self.add_manual_entry_to_table('axa')
        )
        self.btn_add_axa.pack(side=tk.RIGHT, padx=5, pady=2)
        
        # Sol tablo scrollbar
        left_scroll_y = ttk.Scrollbar(left_frame, orient=tk.VERTICAL)
        
        # Sol sütunlar (ŞİRKET sütunu VAR, PLAKA yok)
        columns_axa = ('ID', 'ŞİRKET', 'SİGORTALI', 'TARİH', 'POLİÇE_NO', 'TÜR', 
                       'ÖDEME', 'BRÜT_PRİM', 'NET_PRİM', 'KOMİSYON', 'ÖDENEN')
        
        self.tree_axa = ttk.Treeview(
            left_frame,
            columns=columns_axa,
            show='headings',
            yscrollcommand=left_scroll_y.set,
            selectmode='extended'
        )
        
        headers_axa = {
            'ID': 'ID', 'ŞİRKET': 'ŞİRKET', 'SİGORTALI': 'SİGORTALI', 'TARİH': 'TARİH',
            'POLİÇE_NO': 'POLİÇE NO', 'TÜR': 'TÜR',
            'ÖDEME': 'ÖDEME', 'BRÜT_PRİM': 'BRÜT PRİM', 'NET_PRİM': 'NET PRİM',
            'KOMİSYON': 'KOMİSYON', 'ÖDENEN': 'ÖDENEN'
        }
        
        widths_axa = {
            'ID': 20, 'ŞİRKET': 55, 'SİGORTALI': 120, 'TARİH': 75, 'POLİÇE_NO': 110,
            'TÜR': 55, 'ÖDEME': 50, 'BRÜT_PRİM': 75,
            'NET_PRİM': 70, 'KOMİSYON': 65, 'ÖDENEN': 65
        }
        
        for col in columns_axa:
            self.tree_axa.heading(col, text=headers_axa[col])
            self.tree_axa.column(col, width=widths_axa[col], minwidth=30, stretch=True, anchor='center' if col not in ['SİGORTALI', 'ŞİRKET'] else 'w')
        
        left_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree_axa.pack(fill=tk.BOTH, expand=True)
        
        left_scroll_y.config(command=self.tree_axa.yview)
        
        # ============ SAĞ PANEL: DİĞER ŞİRKETLER ============
        right_frame = tk.Frame(tables_container, bg='#eff6ff')
        right_frame.grid(row=0, column=1, sticky='nsew', padx=(1, 0))
        
        # Sağ panel başlık ve + butonu
        right_header = tk.Frame(right_frame, bg='#eff6ff')
        right_header.pack(fill=tk.X)
        
        tk.Label(right_header, text="DİĞER ŞİRKETLER", font=('Segoe UI', 10, 'bold'),
                 bg='#eff6ff', fg='#1e3a5f').pack(side=tk.LEFT, padx=5, pady=2)
        
        self.btn_add_other = tk.Button(
            right_header, text="+", font=('Segoe UI', 12, 'bold'),
            bg='#059669', fg='white', width=2, cursor='hand2',
            command=lambda: self.add_manual_entry_to_table('other')
        )
        self.btn_add_other.pack(side=tk.RIGHT, padx=5, pady=2)
        
        # Sağ tablo scrollbar
        right_scroll_y = ttk.Scrollbar(right_frame, orient=tk.VERTICAL)
        
        # Sağ sütunlar (şirket VAR, PLAKA yok)
        columns_other = ('ID', 'ŞİRKET', 'SİGORTALI', 'TARİH', 'POLİÇE_NO', 'TÜR', 
                         'ÖDEME', 'BRÜT_PRİM', 'NET_PRİM', 'KOMİSYON', 'ÖDENEN')
        
        self.tree_other = ttk.Treeview(
            right_frame,
            columns=columns_other,
            show='headings',
            yscrollcommand=right_scroll_y.set,
            selectmode='extended'
        )
        
        headers_other = {
            'ID': 'ID', 'ŞİRKET': 'ŞİRKET', 'SİGORTALI': 'SİGORTALI', 'TARİH': 'TARİH',
            'POLİÇE_NO': 'POLİÇE NO', 'TÜR': 'TÜR',
            'ÖDEME': 'ÖDEME', 'BRÜT_PRİM': 'BRÜT PRİM', 'NET_PRİM': 'NET PRİM',
            'KOMİSYON': 'KOMİSYON', 'ÖDENEN': 'ÖDENEN'
        }
        
        widths_other = {
            'ID': 20, 'ŞİRKET': 55, 'SİGORTALI': 120, 'TARİH': 75, 'POLİÇE_NO': 110,
            'TÜR': 55, 'ÖDEME': 50, 'BRÜT_PRİM': 75,
            'NET_PRİM': 70, 'KOMİSYON': 65, 'ÖDENEN': 65
        }
        
        for col in columns_other:
            self.tree_other.heading(col, text=headers_other[col])
            self.tree_other.column(col, width=widths_other[col], minwidth=30, stretch=True, anchor='center' if col not in ['SİGORTALI', 'ŞİRKET'] else 'w')
        
        right_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree_other.pack(fill=tk.BOTH, expand=True)
        
        right_scroll_y.config(command=self.tree_other.yview)
        
        # Ana tree referansı (uyumluluk için - AXA tablosu)
        self.tree = self.tree_axa
        
        # Çift tık düzenleme - her iki tablo için
        self.tree_axa.bind('<Double-Button-1>', self.on_double_click)
        self.tree_other.bind('<Double-Button-1>', self.on_double_click_other)
        
        # Sağ tık menü
        self.context_menu = tk.Menu(self.root, tearoff=0)
        self.context_menu.add_command(label="🗑️ Sil", command=self.delete_selected)
        self.tree_axa.bind('<Button-3>', self.show_context_menu)
        self.tree_other.bind('<Button-3>', self.show_context_menu_other)
        
        # Inline editing için değişkenler
        self.editing_new_row = False
        self.new_row_item = None
        self.current_edit_entry = None
        self.active_tree = self.tree_axa  # Aktif tablo
    
    def create_log_panel(self, parent):
        """Log paneli"""
        log_frame = tk.LabelFrame(
            parent,
            text="📋 LOG",
            font=('Arial', 8, 'bold'),
            bg='white'
        )
        log_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.log_text = tk.Text(
            log_frame,
            height=10,
            font=('Consolas', 7),
            bg='#fafafa',
            fg='#374151',
            wrap=tk.WORD
        )
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=3, pady=3)
        
        self.log_text.tag_config('success', foreground='#16a34a')
        self.log_text.tag_config('error', foreground='#dc2626')
        self.log_text.tag_config('warning', foreground='#ea580c')
        self.log_text.tag_config('info', foreground='#2563eb')
    
    def create_status_bar(self):
        """Durum çubuğu ve alt özet paneli - AXA ve Diğer Şirketler ayrı tablolar halinde"""
        # ============ ALT ÖZET PANELİ ============
        self.bottom_summary_frame = tk.Frame(self.root, bg='#0f172a', height=110)
        self.bottom_summary_frame.pack(side=tk.BOTTOM, fill=tk.X, before=None)
        self.bottom_summary_frame.pack_propagate(False)
        
        # Ana container
        main_container = tk.Frame(self.bottom_summary_frame, bg='#0f172a')
        main_container.pack(expand=True, fill=tk.BOTH, padx=8, pady=4)
        
        # ==================== AXA BÖLÜMÜ ====================
        axa_frame = tk.Frame(main_container, bg='#0f172a')
        axa_frame.pack(side=tk.LEFT, expand=True, fill=tk.BOTH)
        
        # AXA Başlık
        tk.Label(axa_frame, text="AXA", font=('Segoe UI', 10, 'bold'), bg='#0f172a', fg='#3b82f6').pack(anchor='w', padx=2)
        
        # AXA Tabloları yan yana
        axa_tables = tk.Frame(axa_frame, bg='#0f172a')
        axa_tables.pack(fill=tk.BOTH, expand=True)
        
        # --- AXA KREDİ KARTI TABLOSU ---
        axa_kk = tk.LabelFrame(axa_tables, text="KREDİ KARTI", font=('Segoe UI', 8, 'bold'), bg='#1e293b', fg='#f59e0b', bd=1, relief='solid')
        axa_kk.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=2)
        
        # Başlık satırı
        hdr = tk.Frame(axa_kk, bg='#1e293b')
        hdr.pack(fill=tk.X)
        for txt in ['Brüt', 'Net', 'Kom.', 'Ödn.']:
            tk.Label(hdr, text=txt, font=('Segoe UI', 7, 'bold'), bg='#1e293b', fg='#94a3b8', width=9, anchor='e').pack(side=tk.LEFT)
        
        # Toplam satırı
        row1 = tk.Frame(axa_kk, bg='#1e293b')
        row1.pack(fill=tk.X)
        self.lbl_axa_kk_brut = tk.Label(row1, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#10b981', width=9, anchor='e')
        self.lbl_axa_kk_brut.pack(side=tk.LEFT)
        self.lbl_axa_kk_net = tk.Label(row1, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#10b981', width=9, anchor='e')
        self.lbl_axa_kk_net.pack(side=tk.LEFT)
        self.lbl_axa_kk_kom = tk.Label(row1, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#10b981', width=9, anchor='e')
        self.lbl_axa_kk_kom.pack(side=tk.LEFT)
        self.lbl_axa_kk_ode = tk.Label(row1, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#10b981', width=9, anchor='e')
        self.lbl_axa_kk_ode.pack(side=tk.LEFT)
        
        # İptal satırı
        row2 = tk.Frame(axa_kk, bg='#1e293b')
        row2.pack(fill=tk.X)
        self.lbl_axa_kk_iptal_brut = tk.Label(row2, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#f43f5e', width=9, anchor='e')
        self.lbl_axa_kk_iptal_brut.pack(side=tk.LEFT)
        self.lbl_axa_kk_iptal_net = tk.Label(row2, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#f43f5e', width=9, anchor='e')
        self.lbl_axa_kk_iptal_net.pack(side=tk.LEFT)
        self.lbl_axa_kk_iptal_kom = tk.Label(row2, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#f43f5e', width=9, anchor='e')
        self.lbl_axa_kk_iptal_kom.pack(side=tk.LEFT)
        self.lbl_axa_kk_iptal_ode = tk.Label(row2, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#f43f5e', width=9, anchor='e')
        self.lbl_axa_kk_iptal_ode.pack(side=tk.LEFT)
        
        # --- AXA AÇIK TABLOSU ---
        axa_acik = tk.LabelFrame(axa_tables, text="AÇIK", font=('Segoe UI', 8, 'bold'), bg='#1e293b', fg='#f59e0b', bd=1, relief='solid')
        axa_acik.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=2)
        
        hdr = tk.Frame(axa_acik, bg='#1e293b')
        hdr.pack(fill=tk.X)
        for txt in ['Brüt', 'Net', 'Kom.', 'Ödn.']:
            tk.Label(hdr, text=txt, font=('Segoe UI', 7, 'bold'), bg='#1e293b', fg='#94a3b8', width=9, anchor='e').pack(side=tk.LEFT)
        
        row1 = tk.Frame(axa_acik, bg='#1e293b')
        row1.pack(fill=tk.X)
        self.lbl_axa_acik_brut = tk.Label(row1, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#10b981', width=9, anchor='e')
        self.lbl_axa_acik_brut.pack(side=tk.LEFT)
        self.lbl_axa_acik_net = tk.Label(row1, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#10b981', width=9, anchor='e')
        self.lbl_axa_acik_net.pack(side=tk.LEFT)
        self.lbl_axa_acik_kom = tk.Label(row1, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#10b981', width=9, anchor='e')
        self.lbl_axa_acik_kom.pack(side=tk.LEFT)
        self.lbl_axa_acik_ode = tk.Label(row1, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#10b981', width=9, anchor='e')
        self.lbl_axa_acik_ode.pack(side=tk.LEFT)
        
        row2 = tk.Frame(axa_acik, bg='#1e293b')
        row2.pack(fill=tk.X)
        self.lbl_axa_acik_iptal_brut = tk.Label(row2, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#f43f5e', width=9, anchor='e')
        self.lbl_axa_acik_iptal_brut.pack(side=tk.LEFT)
        self.lbl_axa_acik_iptal_net = tk.Label(row2, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#f43f5e', width=9, anchor='e')
        self.lbl_axa_acik_iptal_net.pack(side=tk.LEFT)
        self.lbl_axa_acik_iptal_kom = tk.Label(row2, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#f43f5e', width=9, anchor='e')
        self.lbl_axa_acik_iptal_kom.pack(side=tk.LEFT)
        self.lbl_axa_acik_iptal_ode = tk.Label(row2, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#f43f5e', width=9, anchor='e')
        self.lbl_axa_acik_iptal_ode.pack(side=tk.LEFT)
        
        # --- AXA GENEL TABLOSU ---
        axa_genel = tk.LabelFrame(axa_tables, text="GENEL", font=('Segoe UI', 8, 'bold'), bg='#1e293b', fg='#22d3ee', bd=1, relief='solid')
        axa_genel.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=2)
        
        hdr = tk.Frame(axa_genel, bg='#1e293b')
        hdr.pack(fill=tk.X)
        for txt in ['Brüt', 'Net', 'Kom.', 'Ödn.']:
            tk.Label(hdr, text=txt, font=('Segoe UI', 7, 'bold'), bg='#1e293b', fg='#94a3b8', width=9, anchor='e').pack(side=tk.LEFT)
        
        row1 = tk.Frame(axa_genel, bg='#1e293b')
        row1.pack(fill=tk.X)
        self.lbl_axa_genel_brut = tk.Label(row1, text="0", font=('Segoe UI', 8, 'bold'), bg='#1e293b', fg='#10b981', width=9, anchor='e')
        self.lbl_axa_genel_brut.pack(side=tk.LEFT)
        self.lbl_axa_genel_net = tk.Label(row1, text="0", font=('Segoe UI', 8, 'bold'), bg='#1e293b', fg='#10b981', width=9, anchor='e')
        self.lbl_axa_genel_net.pack(side=tk.LEFT)
        self.lbl_axa_genel_kom = tk.Label(row1, text="0", font=('Segoe UI', 8, 'bold'), bg='#1e293b', fg='#10b981', width=9, anchor='e')
        self.lbl_axa_genel_kom.pack(side=tk.LEFT)
        self.lbl_axa_genel_ode = tk.Label(row1, text="0", font=('Segoe UI', 8, 'bold'), bg='#1e293b', fg='#10b981', width=9, anchor='e')
        self.lbl_axa_genel_ode.pack(side=tk.LEFT)
        
        row2 = tk.Frame(axa_genel, bg='#1e293b')
        row2.pack(fill=tk.X)
        self.lbl_axa_genel_iptal_brut = tk.Label(row2, text="0", font=('Segoe UI', 8, 'bold'), bg='#1e293b', fg='#f43f5e', width=9, anchor='e')
        self.lbl_axa_genel_iptal_brut.pack(side=tk.LEFT)
        self.lbl_axa_genel_iptal_net = tk.Label(row2, text="0", font=('Segoe UI', 8, 'bold'), bg='#1e293b', fg='#f43f5e', width=9, anchor='e')
        self.lbl_axa_genel_iptal_net.pack(side=tk.LEFT)
        self.lbl_axa_genel_iptal_kom = tk.Label(row2, text="0", font=('Segoe UI', 8, 'bold'), bg='#1e293b', fg='#f43f5e', width=9, anchor='e')
        self.lbl_axa_genel_iptal_kom.pack(side=tk.LEFT)
        self.lbl_axa_genel_iptal_ode = tk.Label(row2, text="0", font=('Segoe UI', 8, 'bold'), bg='#1e293b', fg='#f43f5e', width=9, anchor='e')
        self.lbl_axa_genel_iptal_ode.pack(side=tk.LEFT)
        
        # ============ AYIRICI ÇİZGİ ============
        separator = tk.Frame(main_container, bg='#475569', width=3)
        separator.pack(side=tk.LEFT, fill=tk.Y, padx=10)
        
        # ==================== DİĞER BÖLÜMÜ ====================
        other_frame = tk.Frame(main_container, bg='#0f172a')
        other_frame.pack(side=tk.LEFT, expand=True, fill=tk.BOTH)
        
        # DİĞER Başlık
        tk.Label(other_frame, text="DİĞER", font=('Segoe UI', 10, 'bold'), bg='#0f172a', fg='#3b82f6').pack(anchor='w', padx=2)
        
        # DİĞER Tabloları yan yana
        other_tables = tk.Frame(other_frame, bg='#0f172a')
        other_tables.pack(fill=tk.BOTH, expand=True)
        
        # --- DİĞER KREDİ KARTI TABLOSU ---
        other_kk = tk.LabelFrame(other_tables, text="KREDİ KARTI", font=('Segoe UI', 8, 'bold'), bg='#1e293b', fg='#f59e0b', bd=1, relief='solid')
        other_kk.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=2)
        
        hdr = tk.Frame(other_kk, bg='#1e293b')
        hdr.pack(fill=tk.X)
        for txt in ['Brüt', 'Net', 'Kom.', 'Ödn.']:
            tk.Label(hdr, text=txt, font=('Segoe UI', 7, 'bold'), bg='#1e293b', fg='#94a3b8', width=9, anchor='e').pack(side=tk.LEFT)
        
        row1 = tk.Frame(other_kk, bg='#1e293b')
        row1.pack(fill=tk.X)
        self.lbl_other_kk_brut = tk.Label(row1, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#10b981', width=9, anchor='e')
        self.lbl_other_kk_brut.pack(side=tk.LEFT)
        self.lbl_other_kk_net = tk.Label(row1, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#10b981', width=9, anchor='e')
        self.lbl_other_kk_net.pack(side=tk.LEFT)
        self.lbl_other_kk_kom = tk.Label(row1, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#10b981', width=9, anchor='e')
        self.lbl_other_kk_kom.pack(side=tk.LEFT)
        self.lbl_other_kk_ode = tk.Label(row1, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#10b981', width=9, anchor='e')
        self.lbl_other_kk_ode.pack(side=tk.LEFT)
        
        row2 = tk.Frame(other_kk, bg='#1e293b')
        row2.pack(fill=tk.X)
        self.lbl_other_kk_iptal_brut = tk.Label(row2, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#f43f5e', width=9, anchor='e')
        self.lbl_other_kk_iptal_brut.pack(side=tk.LEFT)
        self.lbl_other_kk_iptal_net = tk.Label(row2, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#f43f5e', width=9, anchor='e')
        self.lbl_other_kk_iptal_net.pack(side=tk.LEFT)
        self.lbl_other_kk_iptal_kom = tk.Label(row2, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#f43f5e', width=9, anchor='e')
        self.lbl_other_kk_iptal_kom.pack(side=tk.LEFT)
        self.lbl_other_kk_iptal_ode = tk.Label(row2, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#f43f5e', width=9, anchor='e')
        self.lbl_other_kk_iptal_ode.pack(side=tk.LEFT)
        
        # --- DİĞER AÇIK TABLOSU ---
        other_acik = tk.LabelFrame(other_tables, text="AÇIK", font=('Segoe UI', 8, 'bold'), bg='#1e293b', fg='#f59e0b', bd=1, relief='solid')
        other_acik.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=2)
        
        hdr = tk.Frame(other_acik, bg='#1e293b')
        hdr.pack(fill=tk.X)
        for txt in ['Brüt', 'Net', 'Kom.', 'Ödn.']:
            tk.Label(hdr, text=txt, font=('Segoe UI', 7, 'bold'), bg='#1e293b', fg='#94a3b8', width=9, anchor='e').pack(side=tk.LEFT)
        
        row1 = tk.Frame(other_acik, bg='#1e293b')
        row1.pack(fill=tk.X)
        self.lbl_other_acik_brut = tk.Label(row1, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#10b981', width=9, anchor='e')
        self.lbl_other_acik_brut.pack(side=tk.LEFT)
        self.lbl_other_acik_net = tk.Label(row1, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#10b981', width=9, anchor='e')
        self.lbl_other_acik_net.pack(side=tk.LEFT)
        self.lbl_other_acik_kom = tk.Label(row1, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#10b981', width=9, anchor='e')
        self.lbl_other_acik_kom.pack(side=tk.LEFT)
        self.lbl_other_acik_ode = tk.Label(row1, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#10b981', width=9, anchor='e')
        self.lbl_other_acik_ode.pack(side=tk.LEFT)
        
        row2 = tk.Frame(other_acik, bg='#1e293b')
        row2.pack(fill=tk.X)
        self.lbl_other_acik_iptal_brut = tk.Label(row2, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#f43f5e', width=9, anchor='e')
        self.lbl_other_acik_iptal_brut.pack(side=tk.LEFT)
        self.lbl_other_acik_iptal_net = tk.Label(row2, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#f43f5e', width=9, anchor='e')
        self.lbl_other_acik_iptal_net.pack(side=tk.LEFT)
        self.lbl_other_acik_iptal_kom = tk.Label(row2, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#f43f5e', width=9, anchor='e')
        self.lbl_other_acik_iptal_kom.pack(side=tk.LEFT)
        self.lbl_other_acik_iptal_ode = tk.Label(row2, text="0", font=('Segoe UI', 8), bg='#1e293b', fg='#f43f5e', width=9, anchor='e')
        self.lbl_other_acik_iptal_ode.pack(side=tk.LEFT)
        
        # --- DİĞER GENEL TABLOSU ---
        other_genel = tk.LabelFrame(other_tables, text="GENEL", font=('Segoe UI', 8, 'bold'), bg='#1e293b', fg='#22d3ee', bd=1, relief='solid')
        other_genel.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=2)
        
        hdr = tk.Frame(other_genel, bg='#1e293b')
        hdr.pack(fill=tk.X)
        for txt in ['Brüt', 'Net', 'Kom.', 'Ödn.']:
            tk.Label(hdr, text=txt, font=('Segoe UI', 7, 'bold'), bg='#1e293b', fg='#94a3b8', width=9, anchor='e').pack(side=tk.LEFT)
        
        row1 = tk.Frame(other_genel, bg='#1e293b')
        row1.pack(fill=tk.X)
        self.lbl_other_genel_brut = tk.Label(row1, text="0", font=('Segoe UI', 8, 'bold'), bg='#1e293b', fg='#10b981', width=9, anchor='e')
        self.lbl_other_genel_brut.pack(side=tk.LEFT)
        self.lbl_other_genel_net = tk.Label(row1, text="0", font=('Segoe UI', 8, 'bold'), bg='#1e293b', fg='#10b981', width=9, anchor='e')
        self.lbl_other_genel_net.pack(side=tk.LEFT)
        self.lbl_other_genel_kom = tk.Label(row1, text="0", font=('Segoe UI', 8, 'bold'), bg='#1e293b', fg='#10b981', width=9, anchor='e')
        self.lbl_other_genel_kom.pack(side=tk.LEFT)
        self.lbl_other_genel_ode = tk.Label(row1, text="0", font=('Segoe UI', 8, 'bold'), bg='#1e293b', fg='#10b981', width=9, anchor='e')
        self.lbl_other_genel_ode.pack(side=tk.LEFT)
        
        row2 = tk.Frame(other_genel, bg='#1e293b')
        row2.pack(fill=tk.X)
        self.lbl_other_genel_iptal_brut = tk.Label(row2, text="0", font=('Segoe UI', 8, 'bold'), bg='#1e293b', fg='#f43f5e', width=9, anchor='e')
        self.lbl_other_genel_iptal_brut.pack(side=tk.LEFT)
        self.lbl_other_genel_iptal_net = tk.Label(row2, text="0", font=('Segoe UI', 8, 'bold'), bg='#1e293b', fg='#f43f5e', width=9, anchor='e')
        self.lbl_other_genel_iptal_net.pack(side=tk.LEFT)
        self.lbl_other_genel_iptal_kom = tk.Label(row2, text="0", font=('Segoe UI', 8, 'bold'), bg='#1e293b', fg='#f43f5e', width=9, anchor='e')
        self.lbl_other_genel_iptal_kom.pack(side=tk.LEFT)
        self.lbl_other_genel_iptal_ode = tk.Label(row2, text="0", font=('Segoe UI', 8, 'bold'), bg='#1e293b', fg='#f43f5e', width=9, anchor='e')
        self.lbl_other_genel_iptal_ode.pack(side=tk.LEFT)
        
        # ============ DURUM ÇUBUĞU ============
        self.status_bar = tk.Label(
            self.root,
            text="Hazır",
            font=('Arial', 9),
            bg='#e2e8f0',
            fg='#475569',
            anchor='w',
            padx=10
        )
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)
    
    # =========================================================================
    # VERİ İŞLEMLERİ
    # =========================================================================
    
    def on_filter_change(self):
        """Filtre değiştiğinde"""
        self.load_data()
        self.update_summary()
    
    def apply_filters(self):
        """Filtreleri uygula"""
        self.load_data()
        self.update_summary()
    
    def clear_filters(self):
        """Tüm filtreleri temizle"""
        self.filter_sigortali.delete(0, tk.END)
        self.filter_police_no.delete(0, tk.END)
        self.filter_plaka.delete(0, tk.END)
        self.filter_tur.set('Tümü')
        self.filter_iptal_var.set(False)
        self.load_data()
        self.update_summary()
        self.log("🔄 Filtreler temizlendi", "info")
    
    def load_data(self):
        """Verileri yükle - AXA ve diğer şirketleri ayrı tablolara"""
        # Her iki tabloyu temizle
        for item in self.tree_axa.get_children():
            self.tree_axa.delete(item)
        for item in self.tree_other.get_children():
            self.tree_other.delete(item)
        
        # İptal poliçeleri için kırmızı renk tag'i
        self.tree_axa.tag_configure('iptal', foreground='#dc2626')  # Kırmızı
        self.tree_other.tag_configure('iptal', foreground='#dc2626')  # Kırmızı
        
        # Ay filtresini de ekle
        df = self.db.get_records(kisi=self.current_kisi, yil=None, ay=self.current_ay)
        
        # Sigortalı filtresi
        sigortali_filter = self.filter_sigortali.get().strip().upper()
        if sigortali_filter and not df.empty:
            df = df[df['sigortali'].str.upper().str.contains(sigortali_filter, na=False)]
        
        # Poliçe No filtresi
        police_filter = self.filter_police_no.get().strip().upper()
        if police_filter and not df.empty:
            df = df[df['police_no'].astype(str).str.upper().str.contains(police_filter, na=False)]
        
        # Plaka filtresi
        plaka_filter = self.filter_plaka.get().strip().upper()
        if plaka_filter and not df.empty:
            df = df[df['plaka'].astype(str).str.upper().str.contains(plaka_filter, na=False)]
        
        # Tür filtresi
        tur_filter = self.filter_tur.get()
        if tur_filter != 'Tümü' and not df.empty:
            df = df[df['tur'].str.upper() == tur_filter.upper()]
        
        # İptal filtresi - checkbox işaretliyse sadece iptalleri göster
        if self.filter_iptal_var.get() and not df.empty:
            df = df[df['iptal'] == 1]
        
        if df.empty:
            self.log(f"ℹ️ {self.current_ay} ayında kayıt yok", "info")
            return
        
        axa_count = 0
        other_count = 0
        iptal_count = 0
        
        for _, row in df.iterrows():
            # Ödeme türü - isim göster
            odeme = str(row['odeme_turu']).upper().strip()
            if 'K.KART' in odeme or 'KART' in odeme:
                odeme_display = 'K.KART'
            else:
                odeme_display = 'AÇIK'
            
            # İptal kontrolü
            is_iptal = row.get('iptal', 0) == 1
            if is_iptal:
                iptal_count += 1
            
            # Şirket bilgisi - varsayılan AXA
            sirket = str(row.get('sirket', 'AXA')).upper().strip()
            if not sirket or sirket == 'NONE' or sirket == 'NAN':
                sirket = 'AXA'
            
            if sirket == 'AXA':
                axa_count += 1
                # AXA tablosuna ekle (ŞİRKET sütunu VAR, PLAKA yok)
                # ID yerine satır numarası göster, gerçek ID'yi iid olarak sakla
                values_axa = (
                    axa_count,  # Satır numarası (1, 2, 3...)
                    'AXA',
                    row['sigortali'],
                    row['tarih'],
                    row['police_no'],
                    row['tur'],
                    odeme_display,
                    format_turkish_currency(row['brut_prim']),
                    format_turkish_currency(row['net_prim']),
                    format_turkish_currency(row['toplam_komisyon']),
                    format_turkish_currency(row['odenen_komisyon'])
                )
                # İptal poliçeleri kırmızı renkte
                tags = ('iptal',) if is_iptal else ()
                # Gerçek veritabanı ID'sini iid parametresinde sakla
                self.tree_axa.insert('', tk.END, iid=f"axa_{row['id']}", values=values_axa, tags=tags)
            else:
                other_count += 1
                # Diğer şirketler tablosuna ekle (şirket VAR, PLAKA yok)
                # ID yerine satır numarası göster, gerçek ID'yi iid olarak sakla
                values_other = (
                    other_count,  # Satır numarası (1, 2, 3...)
                    sirket,
                    row['sigortali'],
                    row['tarih'],
                    row['police_no'],
                    row['tur'],
                    odeme_display,
                    format_turkish_currency(row['brut_prim']),
                    format_turkish_currency(row['net_prim']),
                    format_turkish_currency(row['toplam_komisyon']),
                    format_turkish_currency(row['odenen_komisyon'])
                )
                # İptal poliçeleri kırmızı renkte
                tags = ('iptal',) if is_iptal else ()
                # Gerçek veritabanı ID'sini iid parametresinde sakla
                self.tree_other.insert('', tk.END, iid=f"other_{row['id']}", values=values_other, tags=tags)
        
        # Log mesajı - iptal sayısını da göster
        if iptal_count > 0:
            self.log(f"✅ AXA: {axa_count}, Diğer: {other_count} kayıt (🔴 {iptal_count} iptal)", "success")
        else:
            self.log(f"✅ AXA: {axa_count}, Diğer: {other_count} kayıt yüklendi", "success")
    
    def update_summary(self):
        """Özeti güncelle - AXA ve Diğer Şirketler için ayrı ayrı, ödeme türüne göre"""
        # Detaylı özet al
        detailed = self.db.get_detailed_summary(kisi=self.current_kisi, ay=self.current_ay)
        
        # ============ AXA BÖLÜMÜ ============
        # Kredi Kart
        self.lbl_axa_kk_brut.config(text=format_turkish_currency(detailed['axa']['kk']['toplam']['brut']))
        self.lbl_axa_kk_net.config(text=format_turkish_currency(detailed['axa']['kk']['toplam']['net']))
        self.lbl_axa_kk_kom.config(text=format_turkish_currency(detailed['axa']['kk']['toplam']['komisyon']))
        self.lbl_axa_kk_ode.config(text=format_turkish_currency(detailed['axa']['kk']['toplam']['odenen']))
        self.lbl_axa_kk_iptal_brut.config(text=format_turkish_currency(detailed['axa']['kk']['iptal']['brut']))
        self.lbl_axa_kk_iptal_net.config(text=format_turkish_currency(detailed['axa']['kk']['iptal']['net']))
        self.lbl_axa_kk_iptal_kom.config(text=format_turkish_currency(detailed['axa']['kk']['iptal']['komisyon']))
        self.lbl_axa_kk_iptal_ode.config(text=format_turkish_currency(detailed['axa']['kk']['iptal']['odenen']))
        
        # Açık
        self.lbl_axa_acik_brut.config(text=format_turkish_currency(detailed['axa']['acik']['toplam']['brut']))
        self.lbl_axa_acik_net.config(text=format_turkish_currency(detailed['axa']['acik']['toplam']['net']))
        self.lbl_axa_acik_kom.config(text=format_turkish_currency(detailed['axa']['acik']['toplam']['komisyon']))
        self.lbl_axa_acik_ode.config(text=format_turkish_currency(detailed['axa']['acik']['toplam']['odenen']))
        self.lbl_axa_acik_iptal_brut.config(text=format_turkish_currency(detailed['axa']['acik']['iptal']['brut']))
        self.lbl_axa_acik_iptal_net.config(text=format_turkish_currency(detailed['axa']['acik']['iptal']['net']))
        self.lbl_axa_acik_iptal_kom.config(text=format_turkish_currency(detailed['axa']['acik']['iptal']['komisyon']))
        self.lbl_axa_acik_iptal_ode.config(text=format_turkish_currency(detailed['axa']['acik']['iptal']['odenen']))
        
        # Genel
        self.lbl_axa_genel_brut.config(text=format_turkish_currency(detailed['axa']['genel']['toplam']['brut']))
        self.lbl_axa_genel_net.config(text=format_turkish_currency(detailed['axa']['genel']['toplam']['net']))
        self.lbl_axa_genel_kom.config(text=format_turkish_currency(detailed['axa']['genel']['toplam']['komisyon']))
        self.lbl_axa_genel_ode.config(text=format_turkish_currency(detailed['axa']['genel']['toplam']['odenen']))
        self.lbl_axa_genel_iptal_brut.config(text=format_turkish_currency(detailed['axa']['genel']['iptal']['brut']))
        self.lbl_axa_genel_iptal_net.config(text=format_turkish_currency(detailed['axa']['genel']['iptal']['net']))
        self.lbl_axa_genel_iptal_kom.config(text=format_turkish_currency(detailed['axa']['genel']['iptal']['komisyon']))
        self.lbl_axa_genel_iptal_ode.config(text=format_turkish_currency(detailed['axa']['genel']['iptal']['odenen']))
        
        # ============ DİĞER ŞİRKETLER BÖLÜMÜ ============
        # Kredi Kart
        self.lbl_other_kk_brut.config(text=format_turkish_currency(detailed['other']['kk']['toplam']['brut']))
        self.lbl_other_kk_net.config(text=format_turkish_currency(detailed['other']['kk']['toplam']['net']))
        self.lbl_other_kk_kom.config(text=format_turkish_currency(detailed['other']['kk']['toplam']['komisyon']))
        self.lbl_other_kk_ode.config(text=format_turkish_currency(detailed['other']['kk']['toplam']['odenen']))
        self.lbl_other_kk_iptal_brut.config(text=format_turkish_currency(detailed['other']['kk']['iptal']['brut']))
        self.lbl_other_kk_iptal_net.config(text=format_turkish_currency(detailed['other']['kk']['iptal']['net']))
        self.lbl_other_kk_iptal_kom.config(text=format_turkish_currency(detailed['other']['kk']['iptal']['komisyon']))
        self.lbl_other_kk_iptal_ode.config(text=format_turkish_currency(detailed['other']['kk']['iptal']['odenen']))
        
        # Açık
        self.lbl_other_acik_brut.config(text=format_turkish_currency(detailed['other']['acik']['toplam']['brut']))
        self.lbl_other_acik_net.config(text=format_turkish_currency(detailed['other']['acik']['toplam']['net']))
        self.lbl_other_acik_kom.config(text=format_turkish_currency(detailed['other']['acik']['toplam']['komisyon']))
        self.lbl_other_acik_ode.config(text=format_turkish_currency(detailed['other']['acik']['toplam']['odenen']))
        self.lbl_other_acik_iptal_brut.config(text=format_turkish_currency(detailed['other']['acik']['iptal']['brut']))
        self.lbl_other_acik_iptal_net.config(text=format_turkish_currency(detailed['other']['acik']['iptal']['net']))
        self.lbl_other_acik_iptal_kom.config(text=format_turkish_currency(detailed['other']['acik']['iptal']['komisyon']))
        self.lbl_other_acik_iptal_ode.config(text=format_turkish_currency(detailed['other']['acik']['iptal']['odenen']))
        
        # Genel
        self.lbl_other_genel_brut.config(text=format_turkish_currency(detailed['other']['genel']['toplam']['brut']))
        self.lbl_other_genel_net.config(text=format_turkish_currency(detailed['other']['genel']['toplam']['net']))
        self.lbl_other_genel_kom.config(text=format_turkish_currency(detailed['other']['genel']['toplam']['komisyon']))
        self.lbl_other_genel_ode.config(text=format_turkish_currency(detailed['other']['genel']['toplam']['odenen']))
        self.lbl_other_genel_iptal_brut.config(text=format_turkish_currency(detailed['other']['genel']['iptal']['brut']))
        self.lbl_other_genel_iptal_net.config(text=format_turkish_currency(detailed['other']['genel']['iptal']['net']))
        self.lbl_other_genel_iptal_kom.config(text=format_turkish_currency(detailed['other']['genel']['iptal']['komisyon']))
        self.lbl_other_genel_iptal_ode.config(text=format_turkish_currency(detailed['other']['genel']['iptal']['odenen']))
    
    # =========================================================================
    # PDF İŞLEME
    # =========================================================================
    
    def process_pdf_folder(self):
        """PDF klasörü işle"""
        if not PARSER_AVAILABLE:
            messagebox.showerror("Hata", "axa_parser.py modülü bulunamadı!")
            return
        
        folder = filedialog.askdirectory(title="PDF Klasörünü Seçin")
        if not folder:
            return
        
        # Cari klasörü seçildiyse cari parser ile işle
        if os.path.basename(folder).lower() == 'cari':
            pdf_files = [f for f in os.listdir(folder) if f.lower().endswith('.pdf')]
            
            if not pdf_files:
                messagebox.showwarning("Uyarı", "Klasörde PDF dosyası bulunamadı!")
                return
            
            self.log(f"📂 {len(pdf_files)} Cari PDF dosyası bulundu", "info")
            
            basarili = 0
            basarisiz = 0
            
            for pdf_file in pdf_files:
                pdf_path = os.path.join(folder, pdf_file)
                try:
                    data = process_cari_pdf(pdf_path)
                    if data and data['isim']:
                        basarili += 1
                        self.log(f"✅ Cari: {pdf_file} - {data['isim']} - {data['tutar']:.2f} TL", "success")
                    else:
                        basarisiz += 1
                        self.log(f"⚠️ İşlenemedi: {pdf_file}", "warning")
                except Exception as e:
                    basarisiz += 1
                    self.log(f"❌ Hata ({pdf_file}): {str(e)}", "error")
            
            messagebox.showinfo("Cari İşleme Tamamlandı", 
                              f"✅ Başarılı: {basarili}\n❌ Başarısız: {basarisiz}\n\nCari penceresi açılıyor...")
            self.show_cari_window()
            return
        
        pdf_files = [f for f in os.listdir(folder) if f.lower().endswith('.pdf')]
        
        if not pdf_files:
            messagebox.showwarning("Uyarı", "Klasörde PDF dosyası bulunamadı!")
            return
        
        self.log(f"📂 {len(pdf_files)} PDF dosyası bulundu", "info")
        
        # Progress
        progress_win = tk.Frame(self.root, bg='white', relief=tk.RIDGE, borderwidth=2)
        progress_win.place(relx=0.5, rely=0.5, anchor='center', width=400, height=120)
        
        tk.Label(progress_win, text="📄 PDF İşleniyor...", font=('Arial', 11, 'bold'), bg='white').pack(pady=10)
        progress_label = tk.Label(progress_win, text="", font=('Arial', 9), bg='white')
        progress_label.pack(pady=5)
        progress_bar = ttk.Progressbar(progress_win, length=350, mode='determinate')
        progress_bar.pack(pady=10)
        
        basarili = 0
        basarisiz = 0
        duplicate = 0
        
        for idx, pdf_file in enumerate(pdf_files):
            pdf_path = os.path.join(folder, pdf_file)
            progress_label.config(text=pdf_file[:40])
            progress_bar['value'] = ((idx + 1) / len(pdf_files)) * 100
            progress_win.update()
            
            try:
                data = process_pdf(pdf_path, pdf_file)
                
                if data:
                    # İptal poliçesi kontrolü
                    iptal = 1 if is_iptal_police(pdf_path) else 0
                    
                    # Brüt prim parse
                    brut_prim = parse_turkish_amount(data.get('TUTAR', '0'))
                    
                    # Net prim - PDF'den al
                    net_prim_str = data.get('NET_PRIM', '0')
                    net_prim = parse_turkish_amount(net_prim_str) if net_prim_str != '0' else brut_prim
                    
                    # Komisyon hesapla (net prim ile)
                    komisyon = hesapla_komisyon(self.current_kisi, data.get('TÜR', ''), net_prim)
                    
                    # Tarihten yıl ve ay çıkar
                    tarih = data.get('TARİH', '')
                    yil = datetime.now().year
                    ay_from_tarih = self.current_ay  # Varsayılan: seçili ay
                    
                    if tarih:
                        # DD/MM/YYYY veya DD.MM.YYYY formatları
                        tarih_parts = None
                        if '/' in tarih:
                            tarih_parts = tarih.split('/')
                        elif '.' in tarih:
                            tarih_parts = tarih.split('.')
                        
                        if tarih_parts and len(tarih_parts) >= 3:
                            try:
                                ay_num = int(tarih_parts[1])
                                yil = int(tarih_parts[2])
                                # Ay numarasını ay ismine çevir
                                ay_isimleri = ['OCAK', 'ŞUBAT', 'MART', 'NİSAN', 'MAYIS', 'HAZİRAN', 
                                               'TEMMUZ', 'AĞUSTOS', 'EYLÜL', 'EKİM', 'KASIM', 'ARALIK']
                                if 1 <= ay_num <= 12:
                                    ay_from_tarih = ay_isimleri[ay_num - 1]
                            except:
                                pass
                    
                    record = {
                        'kisi': self.current_kisi,
                        'ay': ay_from_tarih,  # Tarihten çıkarılan ay
                        'yil': yil,
                        'sigortali': data.get('SİGORTALI', '-'),
                        'police_no': data.get('POLİÇE NO', '-'),
                        'tarih': tarih,
                        'plaka': data.get('PLAKA', '-'),
                        'tur': data.get('TÜR', '-'),
                        'odeme_turu': 'K.KART',
                        'brut_prim': brut_prim,
                        'net_prim': net_prim,
                        'komisyon_orani': komisyon['komisyon_orani'],
                        'toplam_komisyon': komisyon['komisyon'],
                        'odeme_orani': komisyon['odeme_orani'],
                        'odenen_komisyon': komisyon['odenen'],
                        'iptal': iptal,
                        'sirket': data.get('ŞİRKET', 'AXA')
                    }
                    
                    if self.db.insert_record(record):
                        basarili += 1
                        sirket = data.get('ŞİRKET', 'AXA')
                        iptal_text = " 🔴 İPTAL" if iptal else ""
                        self.log(f"✅ {pdf_file} - {sirket} - {data.get('TÜR', '?')}{iptal_text}", "success")
                    else:
                        duplicate += 1
                        self.log(f"⚠️ Duplicate: {pdf_file}", "warning")
                else:
                    basarisiz += 1
                    self.log(f"⚠️ İşlenemedi: {pdf_file}", "warning")
            
            except Exception as e:
                basarisiz += 1
                self.log(f"❌ Hata ({pdf_file}): {str(e)}", "error")
        
        progress_win.destroy()
        
        msg = f"✅ Başarılı: {basarili}\n"
        if duplicate > 0:
            msg += f"⚠️ Duplicate: {duplicate}\n"
        if basarisiz > 0:
            msg += f"❌ Hatalı: {basarisiz}"
        
        messagebox.showinfo("Tamamlandı", msg)
        self.load_data()
        self.update_summary()
    
    # =========================================================================
    # MANUEL KAYIT
    # =========================================================================
    
    def add_manual_entry(self):
        """Eski fonksiyon - geriye dönük uyumluluk için"""
        self.add_manual_entry_to_table('axa')
    
    def add_manual_entry_to_table(self, table_type='axa'):
        """Manuel kayıt ekle - popup dialog ile veri girişi"""
        self.current_table_type = table_type
        self.show_add_policy_dialog(table_type)
    
    def on_double_click_new_row(self, event):
        """Yeni satır için çift tıklama - inline editing"""
        region = self.active_tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        
        column = self.active_tree.identify_column(event.x)
        item = self.active_tree.identify_row(event.y)
        
        if not item:
            return
        
        # Eğer yeni satır değilse normal çift tık davranışı
        if item != self.new_row_item:
            if self.active_tree == self.tree_other:
                self.on_double_click_other(event)
            else:
                self.on_double_click(event)
            return
        
        col_index = int(column.replace('#', '')) - 1
        # Sütun listesi tabloya göre farklı
        if self.current_table_type == 'other':
            columns_list = ['ID', 'ŞİRKET', 'SİGORTALI', 'TARİH', 'POLİÇE_NO', 'TÜR', 
                            'ÖDEME', 'BRÜT_PRİM', 'NET_PRİM', 'KOMİSYON', 'ÖDENEN']
        else:
            columns_list = ['ID', 'ŞİRKET', 'SİGORTALI', 'TARİH', 'POLİÇE_NO', 'TÜR', 
                            'ÖDEME', 'BRÜT_PRİM', 'NET_PRİM', 'KOMİSYON', 'ÖDENEN']
        
        if col_index >= len(columns_list):
            return
        
        col_name = columns_list[col_index]
        
        # Düzenlenebilir sütunlar
        editable_cols = {
            'ŞİRKET': 'sirket',
            'SİGORTALI': 'sigortali',
            'TARİH': 'tarih', 
            'POLİÇE_NO': 'police_no',
            'TÜR': 'tur',
            'ÖDEME': 'odeme',
            'BRÜT_PRİM': 'brut_prim'
        }
        
        if col_name not in editable_cols:
            return
        
        # Önceki entry'yi kapat
        if self.current_edit_entry:
            self.current_edit_entry.destroy()
        
        # Hücre koordinatlarını al
        bbox = self.active_tree.bbox(item, column)
        if not bbox:
            return
        
        x, y, width, height = bbox
        
        # Koordinatları root pencereye göre hesapla (Combobox için önemli)
        tree_x = self.active_tree.winfo_rootx() - self.root.winfo_rootx()
        tree_y = self.active_tree.winfo_rooty() - self.root.winfo_rooty()
        abs_x = tree_x + x
        abs_y = tree_y + y
        
        field_key = editable_cols[col_name]
        
        # ŞİRKET, TÜR ve ÖDEME için Combobox (root'a yerleştir), diğerleri için Entry
        if col_name == 'ŞİRKET':
            sirketler = ['AXA', 'ALLIANZ', 'ANADOLU', 'GÜNEŞ', 'MAPFRE', 'QUICK', 'HDI', 'DOĞA', 'SOMPİ', 'RAY', 'DİĞER']
            self.current_edit_entry = ttk.Combobox(self.root, values=sirketler, 
                                                    font=('Arial', 9), state='readonly', width=10)
            current_sirket = self.new_row_data.get('sirket', sirketler[0]) or sirketler[0]
            if current_sirket in sirketler:
                self.current_edit_entry.set(current_sirket)
            else:
                self.current_edit_entry.set(sirketler[0])
            self.current_edit_entry.place(x=abs_x, y=abs_y, width=width, height=height)
            
            # DİĞER seçildiğinde elle giriş aç
            def on_sirket_selected_new(e=None):
                if self.current_edit_entry and self.current_edit_entry.get() == 'DİĞER':
                    self.current_edit_entry.destroy()
                    self.current_edit_entry = tk.Entry(self.root, font=('Arial', 9))
                    self.current_edit_entry.place(x=abs_x, y=abs_y, width=width, height=height)
                    self.current_edit_entry.focus_set()
                    self.current_edit_entry.bind('<Return>', lambda e: self.save_cell_edit(field_key, col_index))
                    self.current_edit_entry.bind('<FocusOut>', lambda e: self.save_cell_edit(field_key, col_index))
                else:
                    self.save_cell_edit(field_key, col_index)
            self.current_edit_entry.bind('<<ComboboxSelected>>', on_sirket_selected_new)
        elif col_name == 'TÜR':
            self.current_edit_entry = ttk.Combobox(self.root, values=TURLER, 
                                                    font=('Arial', 9), state='readonly', width=8)
            self.current_edit_entry.set(self.new_row_data.get('tur', TURLER[0]) or TURLER[0])
            self.current_edit_entry.place(x=abs_x, y=abs_y, width=width, height=height)
        elif col_name == 'ÖDEME':
            self.current_edit_entry = ttk.Combobox(self.root, values=ODEME_TURLERI, 
                                                    font=('Arial', 9), state='readonly', width=6)
            self.current_edit_entry.set(self.new_row_data.get('odeme', 'K.KART'))
            self.current_edit_entry.place(x=abs_x, y=abs_y, width=width, height=height)
        else:
            self.current_edit_entry = tk.Entry(self.active_tree, font=('Arial', 9))
            current_val = self.new_row_data.get(field_key, '')
            if current_val and not current_val.startswith('('):
                self.current_edit_entry.insert(0, current_val)
            self.current_edit_entry.place(x=x, y=y, width=width, height=height)
        
        self.current_edit_entry.lift()  # Widget'ı öne çıkar
        self.current_edit_entry.focus_set()
        
        # Enter ile kaydet
        self.current_edit_entry.bind('<Return>', lambda e: self.save_cell_edit(field_key, col_index))
        
        # Combobox için sadece seçim event'i kullan (FocusOut sorun yaratıyor)
        if isinstance(self.current_edit_entry, ttk.Combobox):
            self.current_edit_entry.bind('<<ComboboxSelected>>', lambda e: self.save_cell_edit(field_key, col_index))
        else:
            # Entry için FocusOut kullan
            self.current_edit_entry.bind('<FocusOut>', lambda e: self.save_cell_edit(field_key, col_index))
    
    def save_cell_edit(self, field_key, col_index, move_next=False):
        """Hücre düzenlemesini kaydet"""
        if not self.current_edit_entry:
            return
        
        value = self.current_edit_entry.get().strip()
        self.new_row_data[field_key] = value
        
        # Treeview'daki değeri güncelle
        current_values = list(self.active_tree.item(self.new_row_item, 'values'))
        
        # Değeri güncelle - tablo tipine göre farklı indeksler
        if self.current_table_type == 'other':
            # Sağ tablo: ID, ŞİRKET, SİGORTALI, TARİH, POLİÇE_NO, TÜR, ÖDEME, BRÜT_PRİM, NET, KOM, ÖDENEN
            if field_key == 'sirket':
                current_values[1] = value if value else 'AXA'
            elif field_key == 'sigortali':
                current_values[2] = value if value else '(Sigortalı)'
            elif field_key == 'tarih':
                current_values[3] = value if value else '(Tarih)'
            elif field_key == 'police_no':
                current_values[4] = value if value else '(Poliçe No)'
            elif field_key == 'tur':
                current_values[5] = value if value else '(Tür)'
            elif field_key == 'odeme':
                current_values[6] = value if value else 'K.KART'
            elif field_key == 'brut_prim':
                current_values[7] = value if value else '(Brüt Prim)'
        else:
            # Sol tablo: ID, ŞİRKET, SİGORTALI, TARİH, POLİÇE_NO, TÜR, ÖDEME, BRÜT_PRİM, NET, KOM, ÖDENEN
            if field_key == 'sirket':
                current_values[1] = value if value else 'AXA'
            elif field_key == 'sigortali':
                current_values[2] = value if value else '(Sigortalı)'
            elif field_key == 'tarih':
                current_values[3] = value if value else '(Tarih)'
            elif field_key == 'police_no':
                current_values[4] = value if value else '(Poliçe No)'
            elif field_key == 'tur':
                current_values[5] = value if value else '(Tür)'
            elif field_key == 'odeme':
                current_values[6] = value if value else 'K.KART'
            elif field_key == 'brut_prim':
                current_values[7] = value if value else '(Brüt Prim)'
        
        self.active_tree.item(self.new_row_item, values=current_values)
        
        # Entry'yi kapat
        if self.current_edit_entry:
            self.current_edit_entry.destroy()
            self.current_edit_entry = None
    
    def cancel_cell_edit(self):
        """Hücre düzenlemesini iptal et"""
        if self.current_edit_entry:
            self.current_edit_entry.destroy()
            self.current_edit_entry = None
    
    def cancel_edit_and_remove(self):
        """Düzenlemeyi iptal et ve yeni satırı sil"""
        if self.current_edit_entry:
            self.current_edit_entry.destroy()
            self.current_edit_entry = None
        
        if self.new_row_item and self.editing_new_row:
            self.active_tree.delete(self.new_row_item)
            self.editing_new_row = False
            self.new_row_item = None
            # Her iki tablonun event'ini de geri yükle
            self.tree_axa.bind('<Double-Button-1>', self.on_double_click)
            self.tree_other.bind('<Double-Button-1>', self.on_double_click_other)
            self.log("❌ Yeni kayıt ekleme iptal edildi", "warning")
    
    def check_and_save_new_row(self):
        """Girilen değerlerle kayıt ekle (zorunlu alan kontrolü yok)"""
        # Brüt prim kontrolü
        brut_prim_str = self.new_row_data.get('brut_prim', '0')
        if brut_prim_str and not brut_prim_str.startswith('('):
            try:
                brut_prim = float(brut_prim_str.replace('.', '').replace(',', '.'))
            except:
                brut_prim = 0
        else:
            brut_prim = 0
        
        tarih = self.new_row_data.get('tarih', '')
        if tarih and tarih.startswith('('):
            tarih = datetime.now().strftime('%d/%m/%Y')
        
        yil = datetime.now().year
        if tarih and '/' in tarih:
            try:
                yil = int(tarih.split('/')[2])
            except:
                pass
        
        tur = self.new_row_data.get('tur', '-')
        if tur.startswith('('):
            tur = '-'
        
        # Komisyon hesapla (sadece geçerli tür varsa)
        if tur and tur != '-' and brut_prim > 0:
            komisyon = hesapla_komisyon(self.current_kisi, tur, brut_prim)
        else:
            komisyon = {
                'komisyon_orani': 0,
                'komisyon': 0,
                'odeme_orani': 0,
                'odenen': 0
            }
        
        # Poliçe no kontrolü
        police_no = self.new_row_data.get('police_no', '')
        if not police_no or police_no.startswith('('):
            police_no = f"MNL-{datetime.now().strftime('%Y%m%d%H%M%S')}"
        
        sigortali = self.new_row_data.get('sigortali', '-')
        if sigortali.startswith('('):
            sigortali = '-'
        
        record = {
            'kisi': self.current_kisi,
            'ay': self.current_ay,  # Seçili ayı kullan
            'yil': yil,
            'sigortali': sigortali,
            'police_no': police_no,
            'tarih': tarih,
            'plaka': '-',
            'tur': tur,
            'odeme_turu': self.new_row_data.get('odeme', 'K.KART'),
            'brut_prim': brut_prim,
            'net_prim': brut_prim,
            'komisyon_orani': komisyon['komisyon_orani'],
            'toplam_komisyon': komisyon['komisyon'],
            'odeme_orani': komisyon['odeme_orani'],
            'odenen_komisyon': komisyon['odenen'],
            'notlar': self.new_row_data.get('notlar', ''),
            'sirket': self.new_row_data.get('sirket', 'AXA')
        }
        
        if self.db.insert_record(record):
            self.log("✅ Yeni kayıt başarıyla eklendi!", "success")
            self.editing_new_row = False
            self.new_row_item = None
            # Her iki tablonun event'ini de geri yükle
            self.tree_axa.bind('<Double-Button-1>', self.on_double_click)
            self.tree_other.bind('<Double-Button-1>', self.on_double_click_other)
            self.load_data()
            self.update_summary()
        else:
            self.log("❌ Kayıt eklenemedi! (Poliçe no zaten mevcut olabilir)", "error")
    
    # =========================================================================
    # DİĞER İŞLEMLER
    # =========================================================================
    
    def on_double_click(self, event, table_type='axa'):
        """Çift tıklama - hücreyi düzenle"""
        # Hangi tablo kullanılacak
        if table_type == 'other':
            tree = self.tree_other
        else:
            tree = self.tree_axa
        
        region = tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        
        column = tree.identify_column(event.x)
        item = tree.identify_row(event.y)
        
        if not item:
            return
        
        col_index = int(column.replace('#', '')) - 1
        
        # Sütun listesi tabloya göre değişir
        if table_type == 'other':
            # Diğer şirketler: ŞİRKET sütunu VAR
            columns = ('ID', 'ŞİRKET', 'SİGORTALI', 'TARİH', 'POLİÇE_NO', 'TÜR', 
                       'ÖDEME', 'BRÜT_PRİM', 'NET_PRİM', 'KOMİSYON', 'ÖDENEN')
            tur_index = 5  # TÜR sütun indexi
        else:
            # AXA: ŞİRKET sütunu VAR
            columns = ('ID', 'ŞİRKET', 'SİGORTALI', 'TARİH', 'POLİÇE_NO', 'TÜR', 
                       'ÖDEME', 'BRÜT_PRİM', 'NET_PRİM', 'KOMİSYON', 'ÖDENEN')
            tur_index = 5  # TÜR sütun indexi
        
        if col_index >= len(columns):
            return
        
        col_name = columns[col_index]
        values = tree.item(item, 'values')
        # iid formatından gerçek veritabanı ID'sini çıkar
        # iid formatı: "axa_<id>" veya "other_<id>"
        record_id = item.replace('axa_', '').replace('other_', '')
        
        # Düzenlenebilir sütunlar ve veritabanı alanları
        editable_cols = {
            'ŞİRKET': ('sirket', 'combo_sirket'),
            'SİGORTALI': ('sigortali', 'entry'),
            'TARİH': ('tarih', 'entry'),
            'POLİÇE_NO': ('police_no', 'entry'),
            'TÜR': ('tur', 'combo_tur'),
            'ÖDEME': ('odeme_turu', 'combo_odeme'),
            'BRÜT_PRİM': ('brut_prim', 'entry_number'),
            'NET_PRİM': ('net_prim', 'entry_number_simple')  # Bağımsız düzenleme
        }
        
        if col_name not in editable_cols:
            return
        
        db_field, widget_type = editable_cols[col_name]
        current_value = values[col_index]
        
        # Önceki entry'yi kapat
        if self.current_edit_entry:
            self.current_edit_entry.destroy()
            self.current_edit_entry = None
        
        # Hücre koordinatlarını al
        bbox = tree.bbox(item, column)
        if not bbox:
            return
        
        x, y, width, height = bbox
        
        # Koordinatları root pencereye göre hesapla (Combobox için önemli)
        tree_x = tree.winfo_rootx() - self.root.winfo_rootx()
        tree_y = tree.winfo_rooty() - self.root.winfo_rooty()
        abs_x = tree_x + x
        abs_y = tree_y + y
        
        # Widget oluştur - Combobox'lar için root kullan (dropdown seçimi için)
        if widget_type == 'combo_sirket':
            sirketler = ['AXA', 'ALLIANZ', 'ANADOLU', 'GÜNEŞ', 'MAPFRE', 'QUICK', 'HDI', 'DOĞA', 'SOMPİ', 'RAY', 'DİĞER']
            self.current_edit_entry = ttk.Combobox(self.root, values=sirketler, 
                                                    font=('Arial', 9), state='readonly')
            # Mevcut değeri ayarla
            current_sirket = str(current_value).strip()
            if current_sirket in sirketler:
                self.current_edit_entry.set(current_sirket)
            else:
                self.current_edit_entry.set(sirketler[0])
            self.current_edit_entry.place(x=abs_x, y=abs_y, width=width, height=height)
        elif widget_type == 'combo_tur':
            self.current_edit_entry = ttk.Combobox(self.root, values=TURLER, 
                                                    font=('Arial', 9), state='readonly')
            # Mevcut değeri ayarla
            current_tur = current_value
            if current_tur in TURLER:
                self.current_edit_entry.set(current_tur)
            else:
                self.current_edit_entry.set(TURLER[0])
            self.current_edit_entry.place(x=abs_x, y=abs_y, width=width, height=height)
        elif widget_type == 'combo_odeme':
            self.current_edit_entry = ttk.Combobox(self.root, values=ODEME_TURLERI, 
                                                    font=('Arial', 9), state='readonly')
            # Mevcut değeri temizle (● işaretini kaldır)
            clean_value = str(current_value).replace('●', '').replace('■', '').strip()
            if 'AÇIK' in clean_value or 'ACIK' in clean_value:
                self.current_edit_entry.set('AÇIK')
            else:
                self.current_edit_entry.set('K.KART')
            self.current_edit_entry.place(x=abs_x, y=abs_y, width=width, height=height)
        else:
            self.current_edit_entry = tk.Entry(tree, font=('Arial', 9))
            # Mevcut değeri ekle - sayısal değerler için özel işlem
            if widget_type in ('entry_number', 'entry_number_simple'):
                # Tabloda gösterilen format: 122.60 veya 1.234,56
                # Düzenleme için olduğu gibi bırak (kullanıcı gösterilen değeri görsün)
                clean_value = str(current_value)
            else:
                clean_value = str(current_value)
            self.current_edit_entry.insert(0, clean_value)
            self.current_edit_entry.select_range(0, tk.END)
            self.current_edit_entry.place(x=x, y=y, width=width, height=height)
        
        self.current_edit_entry.lift()  # Widget'ı öne çıkar
        self.current_edit_entry.focus_set()
        
        # Event bindings
        def save_edit(e=None):
            if not self.current_edit_entry:
                return
            try:
                new_value = self.current_edit_entry.get().strip()
            except:
                return
            
            # Brüt prim için özel işlem (sadece komisyon hesapla, net_prim değiştirme)
            if widget_type == 'entry_number':
                try:
                    # Akıllı format tespiti - son ayırıcıya bak
                    new_value_clean = new_value.replace(' ', '')
                    
                    # Hem nokta hem virgül varsa: hangisi son ayırıcı?
                    if '.' in new_value_clean and ',' in new_value_clean:
                        last_comma = new_value_clean.rfind(',')
                        last_dot = new_value_clean.rfind('.')
                        
                        # Son ayırıcıdan sonra kaç karakter var?
                        if last_comma > last_dot:
                            # Virgül en sonda → Türk formatı (1.234,56)
                            new_value_clean = new_value_clean.replace('.', '').replace(',', '.')
                        else:
                            # Nokta en sonda → İngiliz formatı (1,234.56)
                            new_value_clean = new_value_clean.replace(',', '')
                    # Sadece virgül varsa: Türk formatı (1234,56)
                    elif ',' in new_value_clean:
                        new_value_clean = new_value_clean.replace(',', '.')
                    # Sadece nokta varsa: Standart format (1234.56) - olduğu gibi bırak
                    
                    new_value = float(new_value_clean)
                    
                    # Komisyonu yeniden hesapla (net prim olarak brut prim kullanılır, kullanıcı brut primi değiştirdiğinde)
                    tur = values[tur_index]  # TÜR sütunu (tabloya göre index değişir)
                    komisyon = hesapla_komisyon(self.current_kisi, tur, new_value)
                    self.db.update_record(record_id, table_type,
                                         brut_prim=new_value,
                                         net_prim=new_value,
                                         toplam_komisyon=komisyon['komisyon'],
                                         odenen_komisyon=komisyon['odenen'])
                except:
                    self.log("⚠️ Geçersiz sayı formatı!", "warning")
                    if self.current_edit_entry:
                        self.current_edit_entry.destroy()
                        self.current_edit_entry = None
                    return
            # Net prim değiştiğinde komisyonu da yeniden hesapla
            elif widget_type == 'entry_number_simple':
                try:
                    # Akıllı format tespiti - son ayırıcıya bak
                    new_value_clean = new_value.replace(' ', '')
                    
                    if '.' in new_value_clean and ',' in new_value_clean:
                        last_comma = new_value_clean.rfind(',')
                        last_dot = new_value_clean.rfind('.')
                        
                        if last_comma > last_dot:
                            # Virgül en sonda → Türk formatı (1.234,56)
                            new_value_clean = new_value_clean.replace('.', '').replace(',', '.')
                        else:
                            # Nokta en sonda → İngiliz formatı (1,234.56)
                            new_value_clean = new_value_clean.replace(',', '')
                    elif ',' in new_value_clean:
                        new_value_clean = new_value_clean.replace(',', '.')
                    
                    new_value = float(new_value_clean)
                    
                    # TÜR'e göre komisyon yeniden hesapla
                    tur = values[tur_index]
                    komisyon = hesapla_komisyon(self.current_kisi, tur, new_value)
                    self.db.update_record(record_id, table_type,
                                         net_prim=new_value,
                                         toplam_komisyon=komisyon['komisyon'],
                                         odenen_komisyon=komisyon['odenen'])
                except:
                    self.log("⚠️ Geçersiz sayı formatı!", "warning")
                    if self.current_edit_entry:
                        self.current_edit_entry.destroy()
                        self.current_edit_entry = None
                    return
            else:
                # Şirket değişikliği - tablo arası taşıma gerekebilir
                if widget_type == 'combo_sirket':
                    new_sirket = new_value.upper().strip()
                    
                    # DİĞER seçildiyse henüz elle giriş yapılacak, kaydetme
                    if new_sirket == 'DİĞER' or not new_sirket:
                        return
                    
                    old_sirket = str(current_value).upper().strip()
                    
                    # AXA → Diğer şirket (axa tablosundan other tablosuna taşı)
                    if table_type == 'axa' and new_sirket != 'AXA':
                        try:
                            with sqlite3.connect(self.db.db_path) as conn:
                                cursor = conn.cursor()
                                # Mevcut kaydı oku
                                cursor.execute("SELECT * FROM komisyonlar_axa WHERE id = ?", (record_id,))
                                row = cursor.fetchone()
                                if row:
                                    col_names = [desc[0] for desc in cursor.description]
                                    row_dict = dict(zip(col_names, row))
                                    # Other tablosuna ekle
                                    cursor.execute('''
                                    INSERT INTO komisyonlar_other 
                                    (kisi, ay, yil, sigortali, police_no, tarih, plaka, tur, odeme_turu, 
                                     kart_sahibi, brut_prim, tramer, net_prim, komisyon_orani, toplam_komisyon, 
                                     odeme_orani, odenen_komisyon, ikinci_police, iptal, sirket, notlar)
                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                    ''', (
                                        row_dict.get('kisi', ''), row_dict.get('ay', ''), row_dict.get('yil', 0),
                                        row_dict.get('sigortali', ''), row_dict.get('police_no', ''),
                                        row_dict.get('tarih', ''), row_dict.get('plaka', '-'),
                                        row_dict.get('tur', ''), row_dict.get('odeme_turu', 'K.KART'),
                                        row_dict.get('kart_sahibi', ''), row_dict.get('brut_prim', 0),
                                        row_dict.get('tramer', 0), row_dict.get('net_prim', 0),
                                        row_dict.get('komisyon_orani', 0), row_dict.get('toplam_komisyon', 0),
                                        row_dict.get('odeme_orani', 0), row_dict.get('odenen_komisyon', 0),
                                        row_dict.get('ikinci_police', 0), row_dict.get('iptal', 0),
                                        new_sirket, row_dict.get('notlar', '')
                                    ))
                                    # AXA tablosundan sil
                                    cursor.execute("DELETE FROM komisyonlar_axa WHERE id = ?", (record_id,))
                                    conn.commit()
                                    self.log(f"✅ Kayıt AXA → {new_sirket} taşındı", "success")
                        except Exception as e:
                            self.log(f"❌ Taşıma hatası: {e}", "error")
                    
                    # Diğer şirket → AXA (other tablosundan axa tablosuna taşı)
                    elif table_type == 'other' and new_sirket == 'AXA':
                        try:
                            with sqlite3.connect(self.db.db_path) as conn:
                                cursor = conn.cursor()
                                # Mevcut kaydı oku
                                cursor.execute("SELECT * FROM komisyonlar_other WHERE id = ?", (record_id,))
                                row = cursor.fetchone()
                                if row:
                                    col_names = [desc[0] for desc in cursor.description]
                                    row_dict = dict(zip(col_names, row))
                                    # AXA tablosuna ekle
                                    cursor.execute('''
                                    INSERT INTO komisyonlar_axa 
                                    (kisi, ay, yil, sigortali, police_no, tarih, plaka, tur, odeme_turu, 
                                     kart_sahibi, brut_prim, tramer, net_prim, komisyon_orani, toplam_komisyon, 
                                     odeme_orani, odenen_komisyon, ikinci_police, iptal, notlar)
                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                    ''', (
                                        row_dict.get('kisi', ''), row_dict.get('ay', ''), row_dict.get('yil', 0),
                                        row_dict.get('sigortali', ''), row_dict.get('police_no', ''),
                                        row_dict.get('tarih', ''), row_dict.get('plaka', '-'),
                                        row_dict.get('tur', ''), row_dict.get('odeme_turu', 'K.KART'),
                                        row_dict.get('kart_sahibi', ''), row_dict.get('brut_prim', 0),
                                        row_dict.get('tramer', 0), row_dict.get('net_prim', 0),
                                        row_dict.get('komisyon_orani', 0), row_dict.get('toplam_komisyon', 0),
                                        row_dict.get('odeme_orani', 0), row_dict.get('odenen_komisyon', 0),
                                        row_dict.get('ikinci_police', 0), row_dict.get('iptal', 0),
                                        row_dict.get('notlar', '')
                                    ))
                                    # Other tablosundan sil
                                    cursor.execute("DELETE FROM komisyonlar_other WHERE id = ?", (record_id,))
                                    conn.commit()
                                    self.log(f"✅ Kayıt {old_sirket} → AXA taşındı", "success")
                        except Exception as e:
                            self.log(f"❌ Taşıma hatası: {e}", "error")
                    
                    # Aynı tablo içinde şirket güncelleme (other → other, farklı şirket)
                    elif table_type == 'other' and new_sirket != 'AXA':
                        self.db.update_record(record_id, table_type, sirket=new_sirket)
                    # AXA → AXA (değişiklik yok)
                    # else: pass
                elif widget_type == 'combo_tur':
                    # TÜR değiştiğinde komisyonu yeniden hesapla
                    new_tur = new_value.upper().strip()
                    # Mevcut net primi al (index 8 = NET_PRİM sütunu)
                    net_prim_str = str(values[8]).replace('.', '').replace(',', '.')
                    try:
                        net_prim = float(net_prim_str)
                    except:
                        net_prim = 0
                    komisyon = hesapla_komisyon(self.current_kisi, new_tur, net_prim)
                    self.db.update_record(record_id, table_type,
                                         tur=new_tur,
                                         komisyon_orani=komisyon['komisyon_orani'],
                                         toplam_komisyon=komisyon['komisyon'],
                                         odeme_orani=komisyon['odeme_orani'],
                                         odenen_komisyon=komisyon['odenen'])
                else:
                    # Normal alan güncelleme
                    self.db.update_record(record_id, table_type, **{db_field: new_value})
            
            self.log(f"✅ {col_name} güncellendi", "success")
            
            if self.current_edit_entry:
                self.current_edit_entry.destroy()
                self.current_edit_entry = None
            
            self.load_data()
            self.update_summary()
        
        def cancel_edit(e=None):
            if self.current_edit_entry:
                self.current_edit_entry.destroy()
                self.current_edit_entry = None
        
        self.current_edit_entry.bind('<Return>', save_edit)
        self.current_edit_entry.bind('<Escape>', cancel_edit)
        
        # Combobox için sadece seçim event'i kullan (FocusOut sorun yaratıyor)
        if isinstance(self.current_edit_entry, ttk.Combobox):
            if widget_type == 'combo_sirket':
                # DİĞER seçildiğinde elle giriş aç
                def on_sirket_selected(e=None):
                    if self.current_edit_entry and self.current_edit_entry.get() == 'DİĞER':
                        self.current_edit_entry.destroy()
                        self.current_edit_entry = tk.Entry(self.root, font=('Arial', 9))
                        self.current_edit_entry.place(x=abs_x, y=abs_y, width=width, height=height)
                        self.current_edit_entry.lift()
                        self.current_edit_entry.focus_set()
                        self.current_edit_entry.bind('<Return>', save_edit)
                        self.current_edit_entry.bind('<Escape>', cancel_edit)
                        self.current_edit_entry.bind('<FocusOut>', save_edit)
                    else:
                        save_edit()
                self.current_edit_entry.bind('<<ComboboxSelected>>', on_sirket_selected)
            else:
                self.current_edit_entry.bind('<<ComboboxSelected>>', save_edit)
        else:
            # Entry için FocusOut kullan
            self.current_edit_entry.bind('<FocusOut>', save_edit)
    
    def show_context_menu(self, event):
        """Sağ tık menüsü - AXA tablosu"""
        self.active_tree = self.tree_axa
        try:
            self.context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.context_menu.grab_release()
    
    def show_context_menu_other(self, event):
        """Sağ tık menüsü - Diğer şirketler tablosu"""
        self.active_tree = self.tree_other
        try:
            self.context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.context_menu.grab_release()
    
    def on_double_click_other(self, event):
        """Diğer şirketler tablosu için çift tıklama"""
        # on_double_click'i 'other' table_type ile çağır
        self.on_double_click(event, table_type='other')
    
    def delete_selected(self):
        """Seçili kayıtları sil (her iki tablodan)"""
        # Her iki tablodan seçimleri al
        selection_axa = self.tree_axa.selection()
        selection_other = self.tree_other.selection()
        
        all_selections = []
        for item in selection_axa:
            # iid formatı: "axa_<id>" - gerçek ID'yi çıkar
            record_id = item.replace('axa_', '')
            all_selections.append(('axa', record_id))
        for item in selection_other:
            # iid formatı: "other_<id>" - gerçek ID'yi çıkar
            record_id = item.replace('other_', '')
            all_selections.append(('other', record_id))
        
        if not all_selections:
            messagebox.showwarning("Uyarı", "Lütfen kayıt seçin!")
            return
        
        if messagebox.askyesno("Onay", f"{len(all_selections)} kayıt silinecek. Emin misiniz?"):
            for table_type, record_id in all_selections:
                self.db.delete_record(record_id, table_type)
            
            self.log(f"🗑️ {len(all_selections)} kayıt silindi", "warning")
            self.load_data()
            self.update_summary()
    
    def delete_current_month(self):
        """Seçili ayın kayıtlarını sil - tarih sütunundaki aya göre"""
        # Seçili aydaki kayıtları al (get_records tarih sütununa göre filtreler)
        df = self.db.get_records(kisi=self.current_kisi, yil=None, ay=self.current_ay)
        if df.empty:
            messagebox.showinfo("Bilgi", f"{self.current_ay} ayında silinecek kayıt yok!")
            return
        
        count = len(df)
        
        if messagebox.askyesno("⚠️ DİKKAT!", 
                              f"{self.current_kisi} - {self.current_ay} ayında\n"
                              f"toplam {count} kayıt silinecek!\n\n"
                              f"Bu işlem geri alınamaz. Emin misiniz?"):
            try:
                with sqlite3.connect(self.db.db_path) as conn:
                    cursor = conn.cursor()
                    
                    # get_records'dan dönen ID'leri kullanarak sil
                    # Bu sayede tarih sütunundaki aya göre filtrelenen kayıtlar silinir
                    axa_ids = df[df['sirket'] == 'AXA']['id'].tolist()
                    other_ids = df[df['sirket'] != 'AXA']['id'].tolist()
                    
                    # AXA kayıtlarını sil
                    if axa_ids:
                        placeholders = ','.join(['?' for _ in axa_ids])
                        cursor.execute(f"DELETE FROM komisyonlar_axa WHERE id IN ({placeholders})", axa_ids)
                    
                    # Diğer şirket kayıtlarını sil
                    if other_ids:
                        placeholders = ','.join(['?' for _ in other_ids])
                        cursor.execute(f"DELETE FROM komisyonlar_other WHERE id IN ({placeholders})", other_ids)
                    
                    conn.commit()
                
                self.log(f"🗑️ {self.current_ay} ayı temizlendi! {count} kayıt silindi.", "warning")
                self.load_data()
                self.update_summary()
            except Exception as e:
                self.log(f"❌ Silme hatası: {e}", "error")
    
    def delete_current_year(self):
        """Seçili kişinin TÜM yıl kayıtlarını sil - ÇİFT ONAY gerektirir"""
        try:
            # Doğrudan veritabanından kayıt sayısını al
            with sqlite3.connect(self.db.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM komisyonlar_axa WHERE kisi = ?", (self.current_kisi,))
                axa_count = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(*) FROM komisyonlar_other WHERE kisi = ?", (self.current_kisi,))
                other_count = cursor.fetchone()[0]
            
            count = axa_count + other_count
            
            if count == 0:
                messagebox.showinfo("Bilgi", f"{self.current_kisi} için silinecek kayıt yok!")
                return
            
            # BİRİNCİ ONAY
            if not messagebox.askyesno("⚠️ ÇOK DİKKAT!", 
                                       f"⚠️ {self.current_kisi} için TÜM YIL kayıtları silinecek!\n\n"
                                       f"AXA: {axa_count} kayıt\n"
                                       f"Diğer: {other_count} kayıt\n"
                                       f"Toplam: {count} kayıt silinecek!\n\n"
                                       f"Bu işlem GERİ ALINAMAZ!\n"
                                       f"Emin misiniz?"):
                return
            
            # İKİNCİ ONAY
            if not messagebox.askyesno("🔴 SON UYARI!", 
                                       f"🔴 BU SON UYARIDIR!\n\n"
                                       f"{self.current_kisi} için {count} kayıt\n"
                                       f"KALICI olarak SİLİNECEK!\n\n"
                                       f"Devam etmek istediğinize EMİN misiniz?"):
                messagebox.showinfo("İptal", "İşlem iptal edildi.")
                return
            
            with sqlite3.connect(self.db.db_path) as conn:
                cursor = conn.cursor()
                
                # Doğrudan SQL ile tüm kayıtları sil - kişiye göre
                cursor.execute("DELETE FROM komisyonlar_axa WHERE kisi = ?", (self.current_kisi,))
                deleted_axa = cursor.rowcount
                
                cursor.execute("DELETE FROM komisyonlar_other WHERE kisi = ?", (self.current_kisi,))
                deleted_other = cursor.rowcount
                
                conn.commit()
            
            total_deleted = deleted_axa + deleted_other
            self.log(f"🗑️ {self.current_kisi} için TÜM YIL temizlendi! {total_deleted} kayıt silindi (AXA: {deleted_axa}, Diğer: {deleted_other}).", "warning")
            self.load_data()
            self.update_summary()
            messagebox.showinfo("Tamamlandı", f"{self.current_kisi} için {total_deleted} kayıt silindi.\n(AXA: {deleted_axa}, Diğer: {deleted_other})")
        except Exception as e:
            self.log(f"❌ Silme hatası: {e}", "error")
            messagebox.showerror("Hata", f"Silme hatası: {e}")

    
    def export_to_excel(self):
        """Excel'e aktar - Her ay ayrı sheet olarak"""
        
        # Tüm kayıtları çek
        df_all = self.db.get_records(kisi=self.current_kisi, yil=None, ay=None)
        
        if df_all.empty:
            messagebox.showwarning("Uyarı", f"{self.current_kisi} için dışa aktarılacak veri yok!")
            return
        
        self.log(f"📊 Excel export başlatıldı - {self.current_kisi} için {len(df_all)} kayıt bulundu", "info")
        
        # Tarihten ay çıkarma fonksiyonu
        def extract_month(tarih):
            """Tarih string'inden ay numarasını çıkar (1-12)"""
            if pd.isna(tarih) or not tarih:
                return None
            tarih_str = str(tarih).strip()
            try:
                # DD.MM.YYYY formatı
                if '.' in tarih_str:
                    parts = tarih_str.split('.')
                    if len(parts) >= 2:
                        ay_str = parts[1].strip()
                        return int(ay_str) if ay_str.isdigit() else None
                # DD/MM/YYYY formatı
                elif '/' in tarih_str:
                    parts = tarih_str.split('/')
                    if len(parts) >= 2:
                        ay_str = parts[1].strip()
                        return int(ay_str) if ay_str.isdigit() else None
                # YYYY-MM-DD formatı
                elif '-' in tarih_str:
                    parts = tarih_str.split('-')
                    if len(parts) >= 2:
                        ay_str = parts[1].strip()
                        return int(ay_str) if ay_str.isdigit() else None
            except:
                pass
            return None
        
        # Ay bilgisini ekle
        df_all['_ay_num'] = df_all['tarih'].apply(extract_month)
        
        # Dosya adı
        default_name = f"{self.current_kisi}.xlsx"
        
        filename = filedialog.asksaveasfilename(
            title="Excel Dosyası Kaydet",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if not filename:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            # Varsayılan sheet'i sil (sonra aylar eklenecek)
            default_sheet = wb.active
            
            # Stiller
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Başlık stili - Koyu mavi
            header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=10)
            
            # İptal satır - Sadece kırmızı yazı (arka plan yok)
            iptal_font = Font(color="FF0000")  # Kırmızı yazı
            
            # Toplam satır - Koyu yeşil
            toplam_fill = PatternFill(start_color="00b050", end_color="00b050", fill_type="solid")
            toplam_font = Font(bold=True, color="FFFFFF", size=10)
            
            # Bakiye satır - Açık mavi
            bakiye_fill = PatternFill(start_color="bdd7ee", end_color="bdd7ee", fill_type="solid")
            bakiye_font = Font(bold=True, size=10)
            
            # Grup başlığı stili - Turuncu arka plan
            grup_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            grup_font = Font(bold=True, size=10)
            
            # ÖDEME TÜRÜ renkleri
            kkart_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Açık kırmızı
            kkart_font = Font(bold=True, color="8B0000")  # Koyu kırmızı yazı
            acik_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # Açık yeşil
            acik_font = Font(bold=True, color="006400")  # Koyu yeşil yazı
            
            # Ödeme türü renklendirme fonksiyonu
            def apply_odeme_turu_color(cell, odeme_turu_value):
                """ÖDEME TÜRÜ hücresine renk uygula"""
                odeme_upper = str(odeme_turu_value).upper().strip()
                if 'K.KART' in odeme_upper or 'KKART' in odeme_upper or 'KREDİ' in odeme_upper:
                    cell.fill = kkart_fill
                    cell.font = kkart_font
                elif 'AÇIK' in odeme_upper or 'ACIK' in odeme_upper:
                    cell.fill = acik_fill
                    cell.font = acik_font
            
            # Sütun başlıkları
            columns = ['SİGORTALI', 'TARİH', 'POLİÇE NO', 'TÜRÜ', 'ÖDEME ŞEKLİ', 
                      'TUTAR', 'NET', 'KOMİSYON', 'ALINAN KOMİSYON', 'ORAN']
            col_widths = [25, 12, 18, 10, 12, 14, 14, 12, 16, 8]
            
            # Ödeme oranını belirle
            if self.current_kisi == 'YAŞAR':
                odeme_orani = 0.60
                oran_text = "60%"
            else:
                odeme_orani = 0.50
                oran_text = "50%"
            
            sheet_count = 0
            
            # Her ay için ayrı sheet oluştur
            for ay_idx, ay in enumerate(AYLAR):
                ay_numarasi = ay_idx + 1  # 1=OCAK, 2=ŞUBAT, ...
                
                # Bu aya ait kayıtları filtrele (tarihten çıkarılan ay numarasına göre)
                df_ay = df_all[df_all['_ay_num'] == ay_numarasi].copy()
                
                if df_ay.empty:
                    continue  # Bu ayda kayıt yoksa atla
                
                # Yeni sheet oluştur
                ws = wb.create_sheet(title=ay)
                sheet_count += 1
                
                # iptal sütununu integer'a çevir (NaN değerler 0 olsun)
                df_ay['iptal'] = df_ay['iptal'].fillna(0).astype(int)
                
                # Sütun genişliklerini ayarla
                for i, width in enumerate(col_widths):
                    ws.column_dimensions[get_column_letter(i + 1)].width = width
                
                # ========== YAŞAR İÇİN ÖZEL FORMAT ==========
                if self.current_kisi == 'YAŞAR':
                    # YAŞAR için özel sütunlar (görüntüdeki sıraya göre)
                    yasar_columns = ['SİGORTALI', 'POL. NO.', 'TÜR', 'TARİH', 'PLAKA', 'ÖDEME TÜRÜ', 
                                    'KART SAHİBİ', 'BRÜT PRİM', 'NET PRİM', 'TOPLAM KOMİSYON', 
                                    'ÖDENEN KOMİSYON', 'ORAN']
                    yasar_col_widths = [22, 15, 10, 12, 12, 12, 18, 14, 14, 16, 16, 8]
                    
                    # Sütun genişliklerini ayarla
                    for i, width in enumerate(yasar_col_widths):
                        ws.column_dimensions[get_column_letter(i + 1)].width = width
                    
                    # Satır 1: Ay başlığı ve yıl
                    ws.cell(row=1, column=1).value = f"{ay} 2025 YILI HESAP ÖZETİ"
                    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
                    
                    # Satır 2: "ÖNCEKİ AYDAN DEVREDEN BAKİYE" + sağda boş hücre (elle doldurulacak)
                    ws.cell(row=2, column=1).value = "ÖNCEKİ AYDAN DEVREDEN BAKİYE"
                    ws.cell(row=2, column=1).font = Font(bold=True, size=10, color="FF0000")
                    # L sütununda (12) bakiye değeri için boş hücre
                    ws.cell(row=2, column=12).fill = bakiye_fill
                    ws.cell(row=2, column=12).number_format = '#,##0.00'
                    
                    # Satır 3: Sütun başlıkları
                    for i, col_name in enumerate(yasar_columns):
                        col_idx = i + 1
                        cell = ws.cell(row=3, column=col_idx)
                        cell.value = col_name
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border
                    
                    row_num = 4
                    
                    # Özet bölümü için değişkenleri başlat
                    kasko_acik_toplam = 0
                    kasko_toplam_alinan = 0
                    kasko_toplam_tutar = 0
                    kasko_toplam_net = 0
                    kasko_toplam_komisyon = 0
                    trafik_acik_toplam = 0
                    trafik_toplam_alinan = 0
                    trafik_toplam_tutar = 0
                    trafik_toplam_net = 0
                    trafik_toplam_komisyon = 0
                    dask_acik_toplam = 0
                    dask_toplam_alinan = 0
                    dask_toplam_tutar = 0
                    dask_toplam_net = 0
                    dask_toplam_komisyon = 0
                    diger_acik_toplam = 0
                    diger_toplam_alinan = 0
                    diger_toplam_tutar = 0
                    diger_toplam_net = 0
                    diger_toplam_komisyon = 0
                    # Genel iptal toplamları
                    yasar_iptal_toplam = 0
                    
                    # Satır numarası değişkenleri (varsayılan değerler - veri yoksa formüller düzgün çalışsın)
                    kasko_start_row = 0
                    kasko_end_row = 0
                    kasko_toplam_row = 0
                    trafik_start_row = 0
                    trafik_end_row = 0
                    trafik_toplam_row = 0
                    dask_start_row = 0
                    dask_end_row = 0
                    dask_toplam_row = 0
                    diger_start_row = 0
                    diger_end_row = 0
                    diger_toplam_row = 0
                    
                    # ============ KASKO BÖLÜMÜ ============
                    df_kasko = df_ay[df_ay['tur'].str.upper() == 'KASKO']
                    df_kasko_normal = df_kasko[df_kasko['iptal'] == 0]
                    df_kasko_iptal = df_kasko[df_kasko['iptal'] == 1]
                    
                    if not df_kasko.empty:
                        ws.cell(row=row_num, column=1).value = "KASKO"
                        ws.cell(row=row_num, column=1).font = grup_font
                        ws.cell(row=row_num, column=1).fill = grup_fill
                        row_num += 1
                        kasko_start_row = row_num  # KASKO veri satırlarının başlangıcı
                        
                        for _, row in df_kasko_normal.iterrows():
                            tutar = float(row['brut_prim']) if row['brut_prim'] else 0
                            net = float(row['net_prim']) if row['net_prim'] else 0
                            # Komisyon ve ödenen değerleri formüllerle hesaplanacak
                            plaka = str(row.get('plaka', '-')) if row.get('plaka') else '-'
                            if plaka == 'nan' or plaka == 'None': plaka = '-'
                            
                            # Sabit değerler
                            ws.cell(row=row_num, column=1).value = row['sigortali']
                            ws.cell(row=row_num, column=2).value = str(row['police_no'])
                            ws.cell(row=row_num, column=3).value = row['tur']  # TÜR
                            ws.cell(row=row_num, column=4).value = row['tarih']
                            ws.cell(row=row_num, column=5).value = plaka
                            ws.cell(row=row_num, column=6).value = row['odeme_turu']
                            ws.cell(row=row_num, column=7).value = '-'
                            ws.cell(row=row_num, column=8).value = tutar  # BRÜT PRİM
                            ws.cell(row=row_num, column=9).value = net    # NET PRİM
                            
                            # KOMİSYON formülü: =I{row}*0.15 (KASKO için %15)
                            ws.cell(row=row_num, column=10).value = f"=I{row_num}*0.15"
                            
                            # ÖDENEN KOMİSYON formülü: =J{row}*0.6 (YAŞAR için %60)
                            ws.cell(row=row_num, column=11).value = f"=J{row_num}*0.6"
                            
                            ws.cell(row=row_num, column=12).value = oran_text
                            
                            # Stil ve format
                            for i in range(12):
                                cell = ws.cell(row=row_num, column=i + 1)
                                cell.border = border
                                if i == 5:  # ÖDEME TÜRÜ sütunu
                                    apply_odeme_turu_color(cell, ws.cell(row=row_num, column=6).value)
                                if i >= 7 and i <= 10:
                                    cell.number_format = '#,##0.00'
                                    cell.alignment = Alignment(horizontal='right')
                                elif i == 11:
                                    cell.alignment = Alignment(horizontal='center')
                            
                            # Toplamlar için formül değerlerini hesapla (Python'da)
                            kasko_toplam_tutar += tutar
                            kasko_toplam_net += net
                            kasko_toplam_komisyon += net * 0.15
                            kasko_toplam_alinan += net * 0.15 * 0.6
                            
                            row_num += 1
                        kasko_end_row = row_num - 1  # KASKO veri satırlarının sonu (son veri satırı)
                        
                        for _, row in df_kasko_iptal.iterrows():
                            tutar = float(row['brut_prim']) if row['brut_prim'] else 0
                            net = float(row['net_prim']) if row['net_prim'] else 0
                            # Komisyon ve ödenen değerleri formüllerle hesaplanacak
                            alinan_hesaplanan = net * 0.15 * 0.6
                            yasar_iptal_toplam += alinan_hesaplanan  # İptal toplamına ekle
                            plaka = str(row.get('plaka', '-')) if row.get('plaka') else '-'
                            if plaka == 'nan' or plaka == 'None': plaka = '-'
                            
                            # Sabit değerler
                            ws.cell(row=row_num, column=1).value = row['sigortali']
                            ws.cell(row=row_num, column=2).value = str(row['police_no'])
                            ws.cell(row=row_num, column=3).value = row['tur']  # TÜR
                            ws.cell(row=row_num, column=4).value = row['tarih']
                            ws.cell(row=row_num, column=5).value = plaka
                            ws.cell(row=row_num, column=6).value = row['odeme_turu']
                            ws.cell(row=row_num, column=7).value = '-'
                            ws.cell(row=row_num, column=8).value = tutar  # BRÜT PRİM
                            ws.cell(row=row_num, column=9).value = net    # NET PRİM
                            
                            # KOMİSYON formülü: =I{row}*0.15 (KASKO için %15)
                            ws.cell(row=row_num, column=10).value = f"=I{row_num}*0.15"
                            
                            # ÖDENEN KOMİSYON formülü: =J{row}*0.6 (YAŞAR için %60)
                            ws.cell(row=row_num, column=11).value = f"=J{row_num}*0.6"
                            
                            ws.cell(row=row_num, column=12).value = oran_text
                            
                            # Stil ve format - iptal için kırmızı yazı
                            for i in range(12):
                                cell = ws.cell(row=row_num, column=i + 1)
                                cell.font = iptal_font
                                cell.border = border
                                if i == 5:  # ÖDEME TÜRÜ sütunu (iptal olsa bile renklendirme)
                                    apply_odeme_turu_color(cell, ws.cell(row=row_num, column=6).value)
                                if i >= 7 and i <= 10:
                                    cell.number_format = '#,##0.00'
                                    cell.alignment = Alignment(horizontal='right')
                                elif i == 11:
                                    cell.alignment = Alignment(horizontal='center')
                            row_num += 1
                        
                        row_num += 3
                        kasko_toplam_row = row_num  # KASKO TOPLAM satır numarasını kaydet
                        ws.cell(row=row_num, column=6).value = "KASKO TOPLAM"
                        ws.cell(row=row_num, column=6).font = toplam_font
                        ws.cell(row=row_num, column=6).fill = toplam_fill
                        
                        # SUM formülleri: H=BRÜT, I=NET, J=KOMİSYON, K=ÖDENEN
                        toplam_columns = ['H', 'I', 'J', 'K']  # 8, 9, 10, 11
                        for i, col_letter in enumerate(toplam_columns):
                            cell = ws.cell(row=row_num, column=8 + i)
                            # SUM formülü: kasko_start_row ile kasko_end_row arasındaki satırları topla
                            cell.value = f"=SUM({col_letter}{kasko_start_row}:{col_letter}{kasko_end_row})"
                            cell.number_format = '#,##0.00'
                            cell.font = toplam_font
                            cell.fill = toplam_fill
                            cell.alignment = Alignment(horizontal='right')
                            cell.border = border
                        row_num += 1
                        df_kasko_acik = df_kasko[(df_kasko['odeme_turu'].str.upper().str.contains('AÇIK|ACIK', na=False)) & (df_kasko['iptal'] == 0)]
                        kasko_acik_toplam = df_kasko_acik['brut_prim'].sum() if not df_kasko_acik.empty else 0
                        ws.cell(row=row_num, column=6).value = "KASKO AÇIK TOPLAMI"
                        ws.cell(row=row_num, column=6).font = Font(bold=True)
                        cell = ws.cell(row=row_num, column=8)
                        cell.value = kasko_acik_toplam
                        cell.number_format = '#,##0.00'
                        cell.font = Font(bold=True)
                        row_num += 3
                    
                    # ============ TRAFİK BÖLÜMÜ ============
                    df_trafik = df_ay[df_ay['tur'].str.upper() == 'TRAFİK']
                    df_trafik_normal = df_trafik[df_trafik['iptal'] == 0]
                    df_trafik_iptal = df_trafik[df_trafik['iptal'] == 1]
                    
                    if not df_trafik.empty:
                        ws.cell(row=row_num, column=1).value = "TRAFİK"
                        ws.cell(row=row_num, column=1).font = grup_font
                        ws.cell(row=row_num, column=1).fill = grup_fill
                        row_num += 1
                        trafik_start_row = row_num  # TRAFİK veri satırlarının başlangıcı
                        
                        trafik_iptal_alinan = 0  # TRAFİK iptal poliçeleri için ayrı toplam
                        
                        for _, row in df_trafik_normal.iterrows():
                            tutar = float(row['brut_prim']) if row['brut_prim'] else 0
                            net = float(row['net_prim']) if row['net_prim'] else 0
                            # Komisyon ve ödenen değerleri formüllerle hesaplanacak
                            plaka = str(row.get('plaka', '-')) if row.get('plaka') else '-'
                            if plaka == 'nan' or plaka == 'None': plaka = '-'
                            
                            # Sabit değerler
                            ws.cell(row=row_num, column=1).value = row['sigortali']
                            ws.cell(row=row_num, column=2).value = str(row['police_no'])
                            ws.cell(row=row_num, column=3).value = row['tur']  # TÜR
                            ws.cell(row=row_num, column=4).value = row['tarih']
                            ws.cell(row=row_num, column=5).value = plaka
                            ws.cell(row=row_num, column=6).value = row['odeme_turu']
                            ws.cell(row=row_num, column=7).value = '-'
                            ws.cell(row=row_num, column=8).value = tutar  # BRÜT PRİM
                            ws.cell(row=row_num, column=9).value = net    # NET PRİM
                            
                            # KOMİSYON formülü: =I{row}*0.10 (TRAFİK için %10)
                            ws.cell(row=row_num, column=10).value = f"=I{row_num}*0.1"
                            
                            # ÖDENEN KOMİSYON formülü: =J{row}*0.6 (YAŞAR için %60)
                            ws.cell(row=row_num, column=11).value = f"=J{row_num}*0.6"
                            
                            ws.cell(row=row_num, column=12).value = oran_text
                            
                            # Stil ve format
                            for i in range(12):
                                cell = ws.cell(row=row_num, column=i + 1)
                                cell.border = border
                                if i == 5:  # ÖDEME TÜRÜ sütunu
                                    apply_odeme_turu_color(cell, ws.cell(row=row_num, column=6).value)
                                if i >= 7 and i <= 10:
                                    cell.number_format = '#,##0.00'
                                    cell.alignment = Alignment(horizontal='right')
                                elif i == 11:
                                    cell.alignment = Alignment(horizontal='center')
                            
                            # Toplamlar için formül değerlerini hesapla (Python'da)
                            trafik_toplam_tutar += tutar
                            trafik_toplam_net += net
                            trafik_toplam_komisyon += net * 0.10
                            trafik_toplam_alinan += net * 0.10 * 0.6
                            
                            row_num += 1
                        trafik_end_row = row_num - 1  # TRAFİK veri satırlarının sonu
                        
                        for _, row in df_trafik_iptal.iterrows():
                            tutar = float(row['brut_prim']) if row['brut_prim'] else 0
                            net = float(row['net_prim']) if row['net_prim'] else 0
                            # Komisyon ve ödenen değerleri formüllerle hesaplanacak
                            alinan_hesaplanan = net * 0.10 * 0.6
                            yasar_iptal_toplam += alinan_hesaplanan  # İptal toplamına ekle
                            trafik_iptal_alinan += alinan_hesaplanan  # TRAFİK iptal toplamına ekle
                            plaka = str(row.get('plaka', '-')) if row.get('plaka') else '-'
                            if plaka == 'nan' or plaka == 'None': plaka = '-'
                            
                            # Sabit değerler
                            ws.cell(row=row_num, column=1).value = row['sigortali']
                            ws.cell(row=row_num, column=2).value = str(row['police_no'])
                            ws.cell(row=row_num, column=3).value = row['tur']  # TÜR
                            ws.cell(row=row_num, column=4).value = row['tarih']
                            ws.cell(row=row_num, column=5).value = plaka
                            ws.cell(row=row_num, column=6).value = row['odeme_turu']
                            ws.cell(row=row_num, column=7).value = '-'
                            ws.cell(row=row_num, column=8).value = tutar  # BRÜT PRİM
                            ws.cell(row=row_num, column=9).value = net    # NET PRİM
                            
                            # KOMİSYON formülü: =I{row}*0.10 (TRAFİK için %10)
                            ws.cell(row=row_num, column=10).value = f"=I{row_num}*0.1"
                            
                            # ÖDENEN KOMİSYON formülü: =J{row}*0.6 (YAŞAR için %60)
                            ws.cell(row=row_num, column=11).value = f"=J{row_num}*0.6"
                            
                            ws.cell(row=row_num, column=12).value = oran_text
                            
                            # Stil ve format - iptal için kırmızı yazı
                            for i in range(12):
                                cell = ws.cell(row=row_num, column=i + 1)
                                cell.font = iptal_font
                                cell.border = border
                                if i == 5:  # ÖDEME TÜRÜ sütunu (iptal olsa bile)
                                    apply_odeme_turu_color(cell, ws.cell(row=row_num, column=6).value)
                                if i >= 7 and i <= 10:
                                    cell.number_format = '#,##0.00'
                                    cell.alignment = Alignment(horizontal='right')
                                elif i == 11:
                                    cell.alignment = Alignment(horizontal='center')
                            row_num += 1
                        
                        row_num += 3
                        trafik_toplam_row = row_num  # TRAFİK TOPLAM satır numarasını kaydet
                        ws.cell(row=row_num, column=6).value = "TRAFİK TOPLAM"
                        ws.cell(row=row_num, column=6).font = toplam_font
                        ws.cell(row=row_num, column=6).fill = toplam_fill
                        
                        # SUM formülleri: H=BRÜT, I=NET, J=KOMİSYON, K=ÖDENEN
                        toplam_columns = ['H', 'I', 'J', 'K']
                        for i, col_letter in enumerate(toplam_columns):
                            cell = ws.cell(row=row_num, column=8 + i)
                            cell.value = f"=SUM({col_letter}{trafik_start_row}:{col_letter}{trafik_end_row})"
                            cell.number_format = '#,##0.00'
                            cell.font = toplam_font
                            cell.fill = toplam_fill
                            cell.alignment = Alignment(horizontal='right')
                            cell.border = border
                        row_num += 1
                        df_trafik_acik = df_trafik[(df_trafik['odeme_turu'].str.upper().str.contains('AÇIK|ACIK', na=False)) & (df_trafik['iptal'] == 0)]
                        trafik_acik_toplam = df_trafik_acik['brut_prim'].sum() if not df_trafik_acik.empty else 0
                        ws.cell(row=row_num, column=6).value = "TRAFİK AÇIK TOPLAMI"
                        ws.cell(row=row_num, column=6).font = Font(bold=True)
                        cell = ws.cell(row=row_num, column=8)
                        cell.value = trafik_acik_toplam
                        cell.number_format = '#,##0.00'
                        cell.font = Font(bold=True)
                        row_num += 3
                    
                    # ============ DİĞER BÖLÜMÜ (KASKO ve TRAFİK hariç - DASK dahil) ============
                    df_diger = df_ay[~df_ay['tur'].str.upper().isin(['KASKO', 'TRAFİK'])]
                    df_diger_normal = df_diger[df_diger['iptal'] == 0]
                    df_diger_iptal = df_diger[df_diger['iptal'] == 1]
                    
                    if not df_diger.empty:
                        ws.cell(row=row_num, column=1).value = "DİĞER"
                        ws.cell(row=row_num, column=1).font = grup_font
                        ws.cell(row=row_num, column=1).fill = grup_fill
                        row_num += 1
                        diger_start_row = row_num  # DİĞER veri satırlarının başlangıcı
                        
                        for _, row in df_diger_normal.iterrows():
                            tutar = float(row['brut_prim']) if row['brut_prim'] else 0
                            net = float(row['net_prim']) if row['net_prim'] else 0
                            plaka = str(row.get('plaka', '-')) if row.get('plaka') else '-'
                            if plaka == 'nan' or plaka == 'None': plaka = '-'
                            
                            # Tür kontrolü - DASK için farklı komisyon oranı
                            tur_upper = str(row['tur']).upper()
                            is_dask = 'DASK' in tur_upper
                            komisyon_oran = 0.0725 if is_dask else 0.15
                            
                            # Sabit değerler
                            ws.cell(row=row_num, column=1).value = row['sigortali']
                            ws.cell(row=row_num, column=2).value = str(row['police_no'])
                            ws.cell(row=row_num, column=3).value = row['tur']  # TÜR
                            ws.cell(row=row_num, column=4).value = row['tarih']
                            ws.cell(row=row_num, column=5).value = plaka
                            ws.cell(row=row_num, column=6).value = row['odeme_turu']
                            ws.cell(row=row_num, column=7).value = '-'
                            ws.cell(row=row_num, column=8).value = tutar  # BRÜT PRİM
                            ws.cell(row=row_num, column=9).value = net    # NET PRİM
                            
                            # KOMİSYON formülü: DASK için %7.25, diğerleri için %15
                            ws.cell(row=row_num, column=10).value = f'=I{row_num}*{komisyon_oran}'
                            
                            # ÖDENEN KOMİSYON formülü: =J{row}*0.6 (YAŞAR için %60)
                            ws.cell(row=row_num, column=11).value = f"=J{row_num}*0.6"
                            
                            ws.cell(row=row_num, column=12).value = oran_text
                            
                            # Stil ve format
                            for i in range(12):
                                cell = ws.cell(row=row_num, column=i + 1)
                                cell.border = border
                                if i == 5:  # ÖDEME TÜRÜ sütunu
                                    apply_odeme_turu_color(cell, ws.cell(row=row_num, column=6).value)
                                if i >= 7 and i <= 10:
                                    cell.number_format = '#,##0.00'
                                    cell.alignment = Alignment(horizontal='right')
                                elif i == 11:
                                    cell.alignment = Alignment(horizontal='center')
                            
                            # Toplamlar için formül değerlerini hesapla (Python'da)
                            diger_toplam_tutar += tutar
                            diger_toplam_net += net
                            diger_toplam_komisyon += net * komisyon_oran
                            diger_toplam_alinan += net * komisyon_oran * 0.6
                            
                            row_num += 1
                        diger_end_row = row_num - 1  # DİĞER veri satırlarının sonu
                        
                        for _, row in df_diger_iptal.iterrows():
                            tutar = float(row['brut_prim']) if row['brut_prim'] else 0
                            net = float(row['net_prim']) if row['net_prim'] else 0
                            plaka = str(row.get('plaka', '-')) if row.get('plaka') else '-'
                            if plaka == 'nan' or plaka == 'None': plaka = '-'
                            
                            # Tür kontrolü - DASK için farklı komisyon oranı
                            tur_upper = str(row['tur']).upper()
                            is_dask = 'DASK' in tur_upper
                            komisyon_oran = 0.0725 if is_dask else 0.15
                            
                            alinan_hesaplanan = net * komisyon_oran * 0.6
                            yasar_iptal_toplam += alinan_hesaplanan  # İptal toplamına ekle
                            
                            # Sabit değerler
                            ws.cell(row=row_num, column=1).value = row['sigortali']
                            ws.cell(row=row_num, column=2).value = str(row['police_no'])
                            ws.cell(row=row_num, column=3).value = row['tur']  # TÜR
                            ws.cell(row=row_num, column=4).value = row['tarih']
                            ws.cell(row=row_num, column=5).value = plaka
                            ws.cell(row=row_num, column=6).value = row['odeme_turu']
                            ws.cell(row=row_num, column=7).value = '-'
                            ws.cell(row=row_num, column=8).value = tutar  # BRÜT PRİM
                            ws.cell(row=row_num, column=9).value = net    # NET PRİM
                            
                            # KOMİSYON formülü: DASK için %7.25, diğerleri için %15
                            ws.cell(row=row_num, column=10).value = f'=I{row_num}*{komisyon_oran}'
                            
                            # ÖDENEN KOMİSYON formülü: =J{row}*0.6 (YAŞAR için %60)
                            ws.cell(row=row_num, column=11).value = f"=J{row_num}*0.6"
                            
                            ws.cell(row=row_num, column=12).value = oran_text
                            
                            # Stil ve format - iptal için kırmızı yazı
                            for i in range(12):
                                cell = ws.cell(row=row_num, column=i + 1)
                                cell.font = iptal_font
                                cell.border = border
                                if i == 5:  # ÖDEME TÜRÜ sütunu (iptal olsa bile)
                                    apply_odeme_turu_color(cell, ws.cell(row=row_num, column=6).value)
                                if i >= 7 and i <= 10:
                                    cell.number_format = '#,##0.00'
                                    cell.alignment = Alignment(horizontal='right')
                                elif i == 11:
                                    cell.alignment = Alignment(horizontal='center')
                            row_num += 1
                        
                        row_num += 3
                        diger_toplam_row = row_num  # DİĞER TOPLAM satır numarasını kaydet
                        ws.cell(row=row_num, column=6).value = "DİĞER TOPLAM"
                        ws.cell(row=row_num, column=6).font = toplam_font
                        ws.cell(row=row_num, column=6).fill = toplam_fill
                        
                        # SUM formülleri: H=BRÜT, I=NET, J=KOMİSYON, K=ÖDENEN
                        toplam_columns = ['H', 'I', 'J', 'K']
                        for i, col_letter in enumerate(toplam_columns):
                            cell = ws.cell(row=row_num, column=8 + i)
                            cell.value = f"=SUM({col_letter}{diger_start_row}:{col_letter}{diger_end_row})"
                            cell.number_format = '#,##0.00'
                            cell.font = toplam_font
                            cell.fill = toplam_fill
                            cell.alignment = Alignment(horizontal='right')
                            cell.border = border
                        row_num += 1
                        df_diger_acik = df_diger[(df_diger['odeme_turu'].str.upper().str.contains('AÇIK|ACIK', na=False)) & (df_diger['iptal'] == 0)]
                        diger_acik_toplam = df_diger_acik['brut_prim'].sum() if not df_diger_acik.empty else 0
                        ws.cell(row=row_num, column=6).value = "DİĞER AÇIK TOPLAMI"
                        ws.cell(row=row_num, column=6).font = Font(bold=True)
                        cell = ws.cell(row=row_num, column=8)
                        cell.value = diger_acik_toplam
                        cell.number_format = '#,##0.00'
                        cell.font = Font(bold=True)
                        row_num += 2
                    
                    # Bölümler arası boşluk
                    row_num += 1
                    
                    # ============ ÖZET BÖLÜMÜ ============
                    # Açık mavi arka plan
                    acik_mavi = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
                    
                    # Satır 1: KASKO AÇIK TOPLAMI | TRAFİK AÇIK TOPLAMI | İŞYERİ/KONUT/DİĞER AÇIK TOPLM
                    ws.cell(row=row_num, column=1).value = "KASKO AÇIK TOPLAMI"
                    ws.cell(row=row_num, column=1).font = Font(bold=True)
                    ws.cell(row=row_num, column=1).fill = acik_mavi
                    cell = ws.cell(row=row_num, column=2)
                    cell.value = kasko_acik_toplam
                    cell.number_format = '#,##0.00'
                    cell.fill = acik_mavi
                    
                    ws.cell(row=row_num, column=4).value = "TRAFİK AÇIK"
                    ws.cell(row=row_num, column=4).font = Font(bold=True)
                    ws.cell(row=row_num, column=4).fill = acik_mavi
                    cell = ws.cell(row=row_num, column=5)
                    cell.value = trafik_acik_toplam
                    cell.number_format = '#,##0.00'
                    cell.fill = acik_mavi
                    
                    ws.cell(row=row_num, column=7).value = "İŞYERİ/KONUT/-"
                    ws.cell(row=row_num, column=7).font = Font(bold=True)
                    ws.cell(row=row_num, column=7).fill = acik_mavi
                    cell = ws.cell(row=row_num, column=8)
                    cell.value = diger_acik_toplam
                    cell.number_format = '#,##0.00'
                    cell.fill = acik_mavi
                    row_num += 1
                    
                    # Satır 2: ALINAN TAKSİTLER TOPLAMI | - | ALINAN DİĞER TAKSİTLER TOPLAMI
                    ws.cell(row=row_num, column=1).value = "ALINAN TAKSİTLER TOPLAMI"
                    ws.cell(row=row_num, column=1).font = Font(bold=True)
                    ws.cell(row=row_num, column=2).value = "-"
                    
                    ws.cell(row=row_num, column=4).value = "-"
                    
                    ws.cell(row=row_num, column=7).value = "ALINAN DİĞER"
                    ws.cell(row=row_num, column=7).font = Font(bold=True)
                    ws.cell(row=row_num, column=8).value = "-"
                    row_num += 1
                    
                    # Satır 3: YAŞAR KASKO KOMİSYON | YAŞAR TRAFİK KOMİSYON | YAŞAR DİĞER KOMİSYON
                    ws.cell(row=row_num, column=1).value = "YAŞAR KASKO KOMİSYON"
                    ws.cell(row=row_num, column=1).font = bakiye_font
                    ws.cell(row=row_num, column=1).fill = bakiye_fill
                    cell = ws.cell(row=row_num, column=2)
                    cell.value = kasko_toplam_alinan
                    cell.number_format = '#,##0.00'
                    cell.fill = bakiye_fill
                    
                    ws.cell(row=row_num, column=4).value = "YAŞAR TRAFİK KOMİSYON"
                    ws.cell(row=row_num, column=4).font = bakiye_font
                    ws.cell(row=row_num, column=4).fill = bakiye_fill
                    cell = ws.cell(row=row_num, column=5)
                    cell.value = trafik_toplam_alinan
                    cell.number_format = '#,##0.00'
                    cell.fill = bakiye_fill
                    
                    ws.cell(row=row_num, column=7).value = "YAŞAR DİĞER KOMİSYON"
                    ws.cell(row=row_num, column=7).font = bakiye_font
                    ws.cell(row=row_num, column=7).fill = bakiye_fill
                    cell = ws.cell(row=row_num, column=8)
                    cell.value = diger_toplam_alinan
                    cell.number_format = '#,##0.00'
                    cell.fill = bakiye_fill
                    row_num += 3
                    
                    # ============ HESAP AYRINTISI BÖLÜMÜ ============
                    # Sarı arka plan stili (başlıklar için)
                    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    # Hafif gri arkaplan (veriler için)
                    light_gray_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                    
                    # Başlık satırı: YAŞAR TRAFİK HESAP AYRINTISI | NİHAT DİŞLİOĞLU HESAP AYRINTISI | CARİ
                    ws.cell(row=row_num, column=1).value = "YAŞAR TRAFİK HESAP AYRINTISI"
                    ws.cell(row=row_num, column=1).font = Font(bold=True)
                    ws.cell(row=row_num, column=1).fill = yellow_fill
                    
                    ws.cell(row=row_num, column=4).value = "NİHAT DİŞLİOĞLU HESAP AYRINTISI"
                    ws.cell(row=row_num, column=4).font = Font(bold=True)
                    ws.cell(row=row_num, column=4).fill = yellow_fill
                    
                    ws.cell(row=row_num, column=9).value = "CARİ"
                    ws.cell(row=row_num, column=9).font = Font(bold=True)
                    ws.cell(row=row_num, column=9).fill = yellow_fill
                    cari_header_row = row_num  # CARİ başlık satırını kaydet
                    row_num += 1
                    
                    # YAŞAR TRAFİK HESAP AYRINTISI satırları (Sütun A-B)
                    yasar_satirlar = [
                        "ÖNCEKİ AYDAN DEVREDEN ALACAK BAK.",
                        "Y.KÜÇÜK'E ÖDENEN TOPLAM KOMSYN",
                        "Y KÜÇÜK CARİ ALACAKLAR",
                        "SATIŞ İPT.İADE",
                        "TOPLAM"
                    ]
                    
                    # NİHAT DİŞLİOĞLU HESAP AYRINTISI satırları (Sütun D-E)
                    nihat_satirlar = [
                        "ÖNCEKİ AYDAN DEVREDEN ALACAK BAKİYESİ",
                        "AXA AÇIK KASKO ALACAK",
                        "AXA AÇIK TRAFİK ALACAK",
                        "TOPLAM"
                    ]
                    
                    # En uzun liste kadar satır oluştur
                    max_rows = max(len(yasar_satirlar), len(nihat_satirlar), 15)  # CARİ için 15 satır
                    
                    for i in range(max_rows):
                        # YAŞAR sütunu
                        if i < len(yasar_satirlar):
                            cell_a = ws.cell(row=row_num, column=1)
                            cell_a.value = yasar_satirlar[i]
                            cell_a.font = Font(bold=True) if yasar_satirlar[i] == "TOPLAM" else Font()
                            cell_a.fill = light_gray_fill
                            ws.cell(row=row_num, column=2).fill = light_gray_fill
                            
                            # Y.KÜÇÜK'E ÖDENEN TOPLAM KOMSYN satırına formül yaz
                            if yasar_satirlar[i] == "Y.KÜÇÜK'E ÖDENEN TOPLAM KOMSYN":
                                cell = ws.cell(row=row_num, column=2)
                                # Formül: KASKO + TRAFİK + DİĞER toplam ödenen komisyonları (K sütunu)
                                cell.value = f"=K{kasko_toplam_row}+K{trafik_toplam_row}+K{diger_toplam_row}"
                                cell.number_format = '#,##0.00'
                                cell.alignment = Alignment(horizontal='right')
                                odenen_toplam_row = row_num  # Bu satırı kaydet (TOPLAM formülünde kullanılacak)
                            
                            # SATIŞ İPT.İADE satırına tüm iptal poliçelerin komisyon toplamını yaz
                            elif yasar_satirlar[i] == "SATIŞ İPT.İADE":
                                cell = ws.cell(row=row_num, column=2)
                                cell.value = yasar_iptal_toplam
                                cell.number_format = '#,##0.00'
                                cell.alignment = Alignment(horizontal='right')
                                iptal_toplam_row = row_num  # Bu satırı kaydet
                            
                            # TOPLAM satırına net toplamı yaz (ödenen - iptal) - formül olarak
                            elif yasar_satirlar[i] == "TOPLAM":
                                cell = ws.cell(row=row_num, column=2)
                                # Formül: Ödenen Toplam - İptal Toplam
                                cell.value = f"=B{odenen_toplam_row}-B{iptal_toplam_row}"
                                cell.number_format = '#,##0.00'
                                cell.font = Font(bold=True)
                                cell.alignment = Alignment(horizontal='right')
                                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                                    top=Side(style='thin'), bottom=Side(style='thin'))
                                yasar_hesap_toplam_row = row_num  # Hesap ayrıntısı toplam satırını kaydet
                        
                        # NİHAT sütunu
                        if i < len(nihat_satirlar):
                            cell_d = ws.cell(row=row_num, column=4)
                            cell_d.value = nihat_satirlar[i]
                            cell_d.font = Font(bold=True) if nihat_satirlar[i] == "TOPLAM" else Font()
                            cell_d.fill = light_gray_fill
                            ws.cell(row=row_num, column=5).fill = light_gray_fill
                            
                            # AXA AÇIK KASKO ALACAK = Kasko Toplam Komisyon - Kasko Ödenen Komisyon
                            if nihat_satirlar[i] == "AXA AÇIK KASKO ALACAK":
                                cell = ws.cell(row=row_num, column=5)
                                # Formül: J{kasko_toplam_row} - K{kasko_toplam_row}
                                cell.value = f"=J{kasko_toplam_row}-K{kasko_toplam_row}"
                                cell.number_format = '#,##0.00'
                                cell.alignment = Alignment(horizontal='right')
                                kasko_alacak_row = row_num  # Bu satırı kaydet
                            
                            # AXA AÇIK TRAFİK ALACAK = Trafik Toplam Komisyon - Trafik Ödenen Komisyon
                            elif nihat_satirlar[i] == "AXA AÇIK TRAFİK ALACAK":
                                cell = ws.cell(row=row_num, column=5)
                                # Formül: J{trafik_toplam_row} - K{trafik_toplam_row}
                                cell.value = f"=J{trafik_toplam_row}-K{trafik_toplam_row}"
                                cell.number_format = '#,##0.00'
                                cell.alignment = Alignment(horizontal='right')
                                trafik_alacak_row = row_num  # Bu satırı kaydet
                            
                            # TOPLAM = Kasko Alacak + Trafik Alacak
                            elif nihat_satirlar[i] == "TOPLAM":
                                cell = ws.cell(row=row_num, column=5)
                                # Formül: E{kasko_alacak_row} + E{trafik_alacak_row}
                                cell.value = f"=E{kasko_alacak_row}+E{trafik_alacak_row}"
                                cell.number_format = '#,##0.00'
                                cell.font = Font(bold=True)
                                cell.alignment = Alignment(horizontal='right')
                        
                        # CARİ sütunları için cari verilerini yaz
                        # I: isim, J: tutar, K: tarih
                        
                        row_num += 1
                    
                    # ============ CARİ VERİLERİNİ YAZ ============
                    # Cari klasöründen bu aya ait verileri çek
                    script_dir = get_application_path()
                    cari_folder = os.path.join(script_dir, 'cari')
                    cari_data = get_cari_data_for_month(cari_folder, ay_numarasi)
                    
                    # Cari verilerini CARİ başlığının hemen altından başlat
                    cari_row = cari_header_row + 1
                    cari_toplam = 0
                    
                    for cari in cari_data:
                        cell_h = ws.cell(row=cari_row, column=9)
                        cell_h.value = cari['isim']
                        cell_h.alignment = Alignment(horizontal='left')
                        cell_h.fill = light_gray_fill
                        
                        cell_i = ws.cell(row=cari_row, column=10)
                        cell_i.value = cari['tutar']
                        cell_i.number_format = '#,##0.00'
                        cell_i.alignment = Alignment(horizontal='right')
                        cell_i.fill = light_gray_fill
                        cari_toplam += cari['tutar']
                        
                        cell_j = ws.cell(row=cari_row, column=11)
                        cell_j.value = cari['tarih']
                        cell_j.alignment = Alignment(horizontal='center')
                        cell_j.fill = light_gray_fill
                        
                        cari_row += 1
                    
                    # CARİ için TOPLAM satırı - cari verilerinin hemen altına
                    cell_toplam = ws.cell(row=cari_row, column=9)
                    cell_toplam.value = "TOPLAM"
                    cell_toplam.font = Font(bold=True)
                    cell_toplam.fill = light_gray_fill
                    
                    cell = ws.cell(row=cari_row, column=10)
                    cell.value = cari_toplam
                    cell.number_format = '#,##0.00'
                    cell.font = Font(bold=True)
                    cell.fill = light_gray_fill
                    cell.alignment = Alignment(horizontal='right')
                    row_num = max(row_num, cari_row + 1) + 1  # Bir sonraki bölüm için satır numarasını güncelle
                    
                    # Genel TOPLAM satırı - CMC formatı gibi tüm sütunlar için toplamları yaz
                    yasar_genel_brut = kasko_toplam_tutar + trafik_toplam_tutar + diger_toplam_tutar
                    yasar_genel_net = kasko_toplam_net + trafik_toplam_net + diger_toplam_net
                    yasar_genel_komisyon = kasko_toplam_komisyon + trafik_toplam_komisyon + diger_toplam_komisyon
                    yasar_genel_alinan = kasko_toplam_alinan + trafik_toplam_alinan + diger_toplam_alinan
                    
                    ws.cell(row=row_num, column=6).value = "TOPLAM"
                    ws.cell(row=row_num, column=6).font = toplam_font
                    ws.cell(row=row_num, column=6).fill = toplam_fill
                    
                    for i, val in enumerate([yasar_genel_brut, yasar_genel_net, yasar_genel_komisyon, yasar_genel_alinan]):
                        cell = ws.cell(row=row_num, column=8 + i)
                        cell.value = val
                        cell.number_format = '#,##0.00'
                        cell.font = toplam_font
                        cell.fill = toplam_fill
                        cell.alignment = Alignment(horizontal='right')
                        cell.border = border
                    
                    genel_toplam_row = row_num  # Bu satırı kaydet
                    row_num += 1
                    
                    # İPTAL TOPLAMI satırı
                    ws.cell(row=row_num, column=6).value = "İPTAL TOPLAMI"
                    ws.cell(row=row_num, column=6).font = Font(bold=True)
                    
                    cell = ws.cell(row=row_num, column=11)
                    cell.value = yasar_iptal_toplam  # İptal değeri Python'da kalıyor
                    cell.number_format = '#,##0.00'
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='right')
                    
                    alt_iptal_row = row_num  # Bu satırı kaydet
                    row_num += 2
                    
                    # BAKİYE satırı - CMC formatı gibi
                    yasar_bakiye = yasar_genel_alinan - yasar_iptal_toplam
                    
                    ws.cell(row=row_num, column=6).value = "BAKİYE"
                    ws.cell(row=row_num, column=6).font = bakiye_font
                    ws.cell(row=row_num, column=6).fill = bakiye_fill
                    
                    cell = ws.cell(row=row_num, column=8)
                    cell.value = yasar_bakiye
                    cell.number_format = '#,##0.00'
                    cell.font = bakiye_font
                    cell.fill = bakiye_fill
                    cell.alignment = Alignment(horizontal='right')
                    
                    cell = ws.cell(row=row_num, column=11)
                    cell.value = yasar_bakiye
                    cell.number_format = '#,##0.00'
                    cell.font = bakiye_font
                    cell.fill = bakiye_fill
                    cell.alignment = Alignment(horizontal='right')
                
                # ========== TEZER VE KAMİL İÇİN ÖZEL FORMAT (İKİ PANELLİ) ==========
                elif self.current_kisi in ['TEZER', 'KAMİL']:
                    # İki panelli layout: Sol - AXA (A-K), Sağ - Diğer Şirketler (M-X)
                    
                    # AXA ve diğer şirket verilerini ayır
                    df_axa = df_ay[df_ay['sirket'] == 'AXA'].copy()
                    df_other = df_ay[df_ay['sirket'] != 'AXA'].copy()
                    
                    # Sol Panel (AXA) sütun başlıkları - A-J (ŞİRKET yok)
                    axa_columns = ['SİGORTALI', 'TARİH', 'POLİÇE NO', 'TÜRÜ', 'ÖDEME ŞEKLİ', 
                                   'TUTAR', 'NET', 'KOMİSYON', 'ALINAN KOMİSYON', 'ORAN']
                    axa_col_widths = [20, 12, 16, 10, 12, 14, 14, 12, 16, 8]
                    
                    # Sağ Panel (Diğer) sütun başlıkları - M-W (L boş ayırıcı)
                    other_columns = ['SİGORTALI', 'TARİH', 'POLİÇE NO', 'ŞİRKET', 'TÜRÜ', 'ÖDEME ŞEKLİ', 
                                     'TUTAR', 'NET', 'KOMİSYON', 'ALINAN KOMİSYON', 'ORAN']
                    other_col_widths = [20, 12, 16, 12, 10, 12, 14, 14, 12, 16, 8]
                    
                    # Sütun genişliklerini ayarla - Sol Panel (A-K)
                    for i, width in enumerate(axa_col_widths):
                        ws.column_dimensions[get_column_letter(i + 1)].width = width
                    
                    # L sütunu ayırıcı (boş)
                    ws.column_dimensions[get_column_letter(12)].width = 3
                    
                    # Sütun genişliklerini ayarla - Sağ Panel (M-X)
                    for i, width in enumerate(other_col_widths):
                        ws.column_dimensions[get_column_letter(13 + i)].width = width
                    
                    # Sarı arka plan stili (başlık için)
                    tezer_header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    tezer_header_font = Font(bold=True, size=10)
                    
                    # Kişi adını al (TEZER veya KAMİL)
                    kisi_adi = self.current_kisi
                    
                    # Satır 1: Panel başlıkları
                    # Sol: AXA'DAN YAPILAN [KİŞİ] SİGORTA POLİÇELERİ
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
                    cell = ws.cell(row=1, column=1)
                    cell.value = f"AXA'DAN YAPILAN {kisi_adi} SİGORTA POLİÇELERİ"
                    cell.font = Font(bold=True, size=11, color="0000FF")
                    cell.fill = tezer_header_fill
                    cell.alignment = Alignment(horizontal='center')
                    
                    # Sağ: [KİŞİ] ÖZDAYAN'DAN YAPILAN POLİÇELER
                    ws.merge_cells(start_row=1, start_column=13, end_row=1, end_column=24)
                    cell = ws.cell(row=1, column=13)
                    cell.value = f"{kisi_adi} ÖZDAYAN'DAN YAPILAN POLİÇELER"
                    cell.font = Font(bold=True, size=11, color="0000FF")
                    cell.fill = tezer_header_fill
                    cell.alignment = Alignment(horizontal='center')
                    
                    # Satır 2: Ay ve yıl
                    ws.cell(row=2, column=1).value = ay
                    ws.cell(row=2, column=1).font = Font(bold=True)
                    ws.cell(row=2, column=2).value = "2025"
                    ws.cell(row=2, column=2).font = Font(bold=True)
                    
                    ws.cell(row=2, column=13).value = ay
                    ws.cell(row=2, column=13).font = Font(bold=True)
                    ws.cell(row=2, column=14).value = "2025"
                    ws.cell(row=2, column=14).font = Font(bold=True)
                    
                    # Satır 3: Sütun başlıkları - Sol Panel
                    for i, col_name in enumerate(axa_columns):
                        cell = ws.cell(row=3, column=i + 1)
                        cell.value = col_name
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border
                    
                    # Satır 3: Sütun başlıkları - Sağ Panel (sarı arka plan)
                    for i, col_name in enumerate(other_columns):
                        cell = ws.cell(row=3, column=13 + i)
                        cell.value = col_name
                        cell.fill = tezer_header_fill  # Sarı arka plan
                        cell.font = tezer_header_font  # Koyu yazı
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border
                    
                    # Veri satırları için başlangıç
                    row_num = 4
                    
                    # ÖNCEKİ AYDAN DEVREDEN BAKİYE - Sol Panel (Satır 4)
                    ws.cell(row=row_num, column=1).value = "DEVREDEN BAKİYE"
                    ws.cell(row=row_num, column=1).font = Font(bold=True, color="FF0000")
                    ws.cell(row=row_num, column=1).fill = bakiye_fill
                    # Değer hücreleri boş (elle doldurulacak)
                    for col in range(2, 12):
                        ws.cell(row=row_num, column=col).fill = bakiye_fill
                    
                    # DEVREDEN BAKİYE - Sağ Panel (Satır 4) - görüntüdeki gibi basit format
                    ws.cell(row=row_num, column=13).value = "DEVREDEN BAKİYE"
                    ws.cell(row=row_num, column=13).font = Font(bold=True)
                    # Değer için 23. sütun (ALINAN KOMİSYON sütunu) - elle doldurulacak boş bırakıldı
                    
                    row_num = 5
                    
                    # AXA verilerini yaz (Sol Panel)
                    axa_toplam_brut = 0
                    axa_toplam_net = 0
                    axa_toplam_komisyon = 0
                    axa_toplam_odenen = 0
                    # İptal toplamları
                    axa_iptal_brut = 0
                    axa_iptal_net = 0
                    axa_iptal_komisyon = 0
                    axa_iptal_odenen = 0
                    axa_row_num = row_num
                    
                    # Normal kayıtlar
                    df_axa_normal = df_axa[df_axa['iptal'] == 0]
                    df_axa_iptal = df_axa[df_axa['iptal'] == 1]
                    
                    for _, row in df_axa_normal.iterrows():
                        brut = float(row['brut_prim']) if row['brut_prim'] else 0
                        net = float(row['net_prim']) if row['net_prim'] else 0
                        tur = str(row.get('tur', '')).upper()
                        
                        # Komisyon oranı belirle (toplamlar için Python'da hesapla)
                        if 'TRAFİK' in tur:
                            komisyon_hesap = net * 0.10
                            komisyon_oran = 0.10
                        elif 'DASK' in tur:
                            komisyon_hesap = net * 0.0725
                            komisyon_oran = 0.0725
                        else:
                            if self.current_kisi == 'TEZER':
                                komisyon_hesap = net * 0.13
                                komisyon_oran = 0.13
                            else:  # KAMİL
                                komisyon_hesap = net * 0.15
                                komisyon_oran = 0.15
                        
                        alinan_hesap = komisyon_hesap * 0.50
                        
                        axa_toplam_brut += brut
                        axa_toplam_net += net
                        axa_toplam_komisyon += komisyon_hesap
                        axa_toplam_odenen += alinan_hesap
                        
                        # Sabit değerler
                        ws.cell(row=axa_row_num, column=1).value = row['sigortali']
                        ws.cell(row=axa_row_num, column=2).value = row['tarih']
                        ws.cell(row=axa_row_num, column=3).value = str(row['police_no'])
                        ws.cell(row=axa_row_num, column=4).value = row['tur']
                        ws.cell(row=axa_row_num, column=5).value = row['odeme_turu']
                        ws.cell(row=axa_row_num, column=6).value = brut  # TUTAR
                        ws.cell(row=axa_row_num, column=7).value = net   # NET
                        
                        # KOMİSYON formülü: Türe göre dinamik hesaplama
                        # IF(ISNUMBER(SEARCH("TRAFİK",D{row})),G{row}*0.1,IF(ISNUMBER(SEARCH("DASK",D{row})),G{row}*0.0725,G{row}*oran))
                        if self.current_kisi == 'TEZER':
                            ws.cell(row=axa_row_num, column=8).value = f'=IF(ISNUMBER(SEARCH("TRAFİK",D{axa_row_num})),G{axa_row_num}*0.1,IF(ISNUMBER(SEARCH("DASK",D{axa_row_num})),G{axa_row_num}*0.0725,G{axa_row_num}*0.13))'
                        else:  # KAMİL
                            ws.cell(row=axa_row_num, column=8).value = f'=IF(ISNUMBER(SEARCH("TRAFİK",D{axa_row_num})),G{axa_row_num}*0.1,IF(ISNUMBER(SEARCH("DASK",D{axa_row_num})),G{axa_row_num}*0.0725,G{axa_row_num}*0.15))'
                        
                        # ALINAN KOMİSYON formülü: =H{row}*0.5
                        ws.cell(row=axa_row_num, column=9).value = f"=H{axa_row_num}*0.5"
                        
                        ws.cell(row=axa_row_num, column=10).value = '50%'
                        
                        # Stil ve format
                        for i in range(10):
                            cell = ws.cell(row=axa_row_num, column=i + 1)
                            cell.border = border
                            if i == 4:  # ÖDEME ŞEKLİ sütunu
                                apply_odeme_turu_color(cell, ws.cell(row=axa_row_num, column=5).value)
                            if i >= 5 and i <= 8:  # TUTAR, NET, KOMİSYON, ALINAN KOMİSYON
                                cell.number_format = '#,##0.00'
                                cell.alignment = Alignment(horizontal='right')
                            elif i == 9:  # ORAN
                                cell.alignment = Alignment(horizontal='center')
                        
                        axa_row_num += 1
                    
                    # İptal kayıtları (kırmızı yazı)
                    for _, row in df_axa_iptal.iterrows():
                        brut = float(row['brut_prim']) if row['brut_prim'] else 0
                        net = float(row['net_prim']) if row['net_prim'] else 0
                        tur = str(row.get('tur', '')).upper()
                        
                        # Komisyon oranı belirle (toplamlar için Python'da hesapla)
                        if 'TRAFİK' in tur:
                            komisyon_hesap = net * 0.10
                        else:
                            if self.current_kisi == 'TEZER':
                                komisyon_hesap = net * 0.13
                            else:  # KAMİL
                                komisyon_hesap = net * 0.15
                        
                        alinan_hesap = komisyon_hesap * 0.50
                        
                        # İptal toplamlarını biriktir
                        axa_iptal_brut += brut
                        axa_iptal_net += net
                        axa_iptal_komisyon += komisyon_hesap
                        axa_iptal_odenen += alinan_hesap
                        
                        # Sabit değerler
                        ws.cell(row=axa_row_num, column=1).value = row['sigortali']
                        ws.cell(row=axa_row_num, column=2).value = row['tarih']
                        ws.cell(row=axa_row_num, column=3).value = str(row['police_no'])
                        ws.cell(row=axa_row_num, column=4).value = row['tur']
                        ws.cell(row=axa_row_num, column=5).value = row['odeme_turu']
                        ws.cell(row=axa_row_num, column=6).value = brut  # TUTAR
                        ws.cell(row=axa_row_num, column=7).value = net   # NET
                        
                        # KOMİSYON formülü: Türe göre dinamik hesaplama
                        if self.current_kisi == 'TEZER':
                            ws.cell(row=axa_row_num, column=8).value = f'=IF(ISNUMBER(SEARCH("TRAFİK",D{axa_row_num})),G{axa_row_num}*0.1,IF(ISNUMBER(SEARCH("DASK",D{axa_row_num})),G{axa_row_num}*0.0725,G{axa_row_num}*0.13))'
                        else:  # KAMİL
                            ws.cell(row=axa_row_num, column=8).value = f'=IF(ISNUMBER(SEARCH("TRAFİK",D{axa_row_num})),G{axa_row_num}*0.1,IF(ISNUMBER(SEARCH("DASK",D{axa_row_num})),G{axa_row_num}*0.0725,G{axa_row_num}*0.15))'
                        
                        # ALINAN KOMİSYON formülü: =H{row}*0.5
                        ws.cell(row=axa_row_num, column=9).value = f"=H{axa_row_num}*0.5"
                        
                        ws.cell(row=axa_row_num, column=10).value = '50%'
                        
                        # Stil ve format - iptal için kırmızı yazı
                        for i in range(10):
                            cell = ws.cell(row=axa_row_num, column=i + 1)
                            cell.font = iptal_font
                            cell.border = border
                            if i == 4:  # ÖDEME ŞEKLİ sütunu
                                apply_odeme_turu_color(cell, ws.cell(row=axa_row_num, column=5).value)
                            if i >= 5 and i <= 8:
                                cell.number_format = '#,##0.00'
                                cell.alignment = Alignment(horizontal='right')
                            elif i == 9:
                                cell.alignment = Alignment(horizontal='center')
                        
                        axa_row_num += 1
                    
                    # Diğer şirket verilerini yaz (Sağ Panel)
                    other_toplam_brut = 0
                    other_toplam_net = 0
                    other_toplam_komisyon = 0
                    other_toplam_odenen = 0
                    # İptal toplamları
                    other_iptal_brut = 0
                    other_iptal_net = 0
                    other_iptal_komisyon = 0
                    other_iptal_odenen = 0
                    other_row_num = row_num
                    
                    df_other_normal = df_other[df_other['iptal'] == 0]
                    df_other_iptal = df_other[df_other['iptal'] == 1]
                    
                    for _, row in df_other_normal.iterrows():
                        brut = float(row['brut_prim']) if row['brut_prim'] else 0
                        net = float(row['net_prim']) if row['net_prim'] else 0
                        sirket = row.get('sirket', '-')
                        tur = str(row.get('tur', '')).upper()
                        
                        # Komisyon oranı belirle (toplamlar için Python'da hesapla)
                        if 'TRAFİK' in tur:
                            komisyon_hesap = net * 0.10
                        elif 'DASK' in tur:
                            komisyon_hesap = net * 0.0725
                        else:
                            if self.current_kisi == 'TEZER':
                                komisyon_hesap = net * 0.13
                            else:  # KAMİL
                                komisyon_hesap = net * 0.15
                        
                        alinan_hesap = komisyon_hesap * 0.50
                        
                        other_toplam_brut += brut
                        other_toplam_net += net
                        other_toplam_komisyon += komisyon_hesap
                        other_toplam_odenen += alinan_hesap
                        
                        # Sabit değerler (sağ panel M-W, sütun 13-23)
                        ws.cell(row=other_row_num, column=13).value = row['sigortali']
                        ws.cell(row=other_row_num, column=14).value = row['tarih']
                        ws.cell(row=other_row_num, column=15).value = str(row['police_no'])
                        ws.cell(row=other_row_num, column=16).value = sirket
                        ws.cell(row=other_row_num, column=17).value = row['tur']
                        ws.cell(row=other_row_num, column=18).value = row['odeme_turu']
                        ws.cell(row=other_row_num, column=19).value = brut  # TUTAR
                        ws.cell(row=other_row_num, column=20).value = net   # NET
                        
                        # KOMİSYON formülü: Türe göre dinamik hesaplama (Q=tür, T=net)
                        if self.current_kisi == 'TEZER':
                            ws.cell(row=other_row_num, column=21).value = f'=IF(ISNUMBER(SEARCH("TRAFİK",Q{other_row_num})),T{other_row_num}*0.1,IF(ISNUMBER(SEARCH("DASK",Q{other_row_num})),T{other_row_num}*0.0725,T{other_row_num}*0.13))'
                        else:  # KAMİL
                            ws.cell(row=other_row_num, column=21).value = f'=IF(ISNUMBER(SEARCH("TRAFİK",Q{other_row_num})),T{other_row_num}*0.1,IF(ISNUMBER(SEARCH("DASK",Q{other_row_num})),T{other_row_num}*0.0725,T{other_row_num}*0.15))'
                        
                        # ALINAN KOMİSYON formülü: =U{row}*0.5
                        ws.cell(row=other_row_num, column=22).value = f"=U{other_row_num}*0.5"
                        
                        ws.cell(row=other_row_num, column=23).value = '50%'
                        
                        # Stil ve format
                        for i in range(11):
                            cell = ws.cell(row=other_row_num, column=13 + i)
                            cell.border = border
                            if i == 5:  # ÖDEME ŞEKLİ sütunu (sağ panel)
                                apply_odeme_turu_color(cell, ws.cell(row=other_row_num, column=18).value)
                            if i >= 6 and i <= 9:  # TUTAR, NET, KOMİSYON, ALINAN KOMİSYON
                                cell.number_format = '#,##0.00'
                                cell.alignment = Alignment(horizontal='right')
                            elif i == 10:  # ORAN
                                cell.alignment = Alignment(horizontal='center')
                        
                        other_row_num += 1
                    
                    # İptal kayıtları (kırmızı yazı)
                    for _, row in df_other_iptal.iterrows():
                        brut = float(row['brut_prim']) if row['brut_prim'] else 0
                        net = float(row['net_prim']) if row['net_prim'] else 0
                        sirket = row.get('sirket', '-')
                        tur = str(row.get('tur', '')).upper()
                        
                        # Komisyon oranı belirle (toplamlar için Python'da hesapla)
                        if 'TRAFİK' in tur:
                            komisyon_hesap = net * 0.10
                        elif 'DASK' in tur:
                            komisyon_hesap = net * 0.0725
                        else:
                            if self.current_kisi == 'TEZER':
                                komisyon_hesap = net * 0.13
                            else:  # KAMİL
                                komisyon_hesap = net * 0.15
                        
                        alinan_hesap = komisyon_hesap * 0.50
                        
                        # İptal toplamlarını biriktir
                        other_iptal_brut += brut
                        other_iptal_net += net
                        other_iptal_komisyon += komisyon_hesap
                        other_iptal_odenen += alinan_hesap
                        
                        # Sabit değerler (sağ panel M-W, sütun 13-23)
                        ws.cell(row=other_row_num, column=13).value = row['sigortali']
                        ws.cell(row=other_row_num, column=14).value = row['tarih']
                        ws.cell(row=other_row_num, column=15).value = str(row['police_no'])
                        ws.cell(row=other_row_num, column=16).value = sirket
                        ws.cell(row=other_row_num, column=17).value = row['tur']
                        ws.cell(row=other_row_num, column=18).value = row['odeme_turu']
                        ws.cell(row=other_row_num, column=19).value = brut  # TUTAR
                        ws.cell(row=other_row_num, column=20).value = net   # NET
                        
                        # KOMİSYON formülü: Türe göre dinamik hesaplama (Q=tür, T=net)
                        if self.current_kisi == 'TEZER':
                            ws.cell(row=other_row_num, column=21).value = f'=IF(ISNUMBER(SEARCH("TRAFİK",Q{other_row_num})),T{other_row_num}*0.1,IF(ISNUMBER(SEARCH("DASK",Q{other_row_num})),T{other_row_num}*0.0725,T{other_row_num}*0.13))'
                        else:  # KAMİL
                            ws.cell(row=other_row_num, column=21).value = f'=IF(ISNUMBER(SEARCH("TRAFİK",Q{other_row_num})),T{other_row_num}*0.1,IF(ISNUMBER(SEARCH("DASK",Q{other_row_num})),T{other_row_num}*0.0725,T{other_row_num}*0.15))'
                        
                        # ALINAN KOMİSYON formülü: =U{row}*0.5
                        ws.cell(row=other_row_num, column=22).value = f"=U{other_row_num}*0.5"
                        
                        ws.cell(row=other_row_num, column=23).value = '50%'
                        
                        # Stil ve format - iptal için kırmızı yazı
                        for i in range(11):
                            cell = ws.cell(row=other_row_num, column=13 + i)
                            cell.font = iptal_font
                            cell.border = border
                            if i == 5:  # ÖDEME ŞEKLİ sütunu (iptal olsa bile)
                                apply_odeme_turu_color(cell, ws.cell(row=other_row_num, column=18).value)
                            if i >= 6 and i <= 9:
                                cell.number_format = '#,##0.00'
                                cell.alignment = Alignment(horizontal='right')
                            elif i == 10:
                                cell.alignment = Alignment(horizontal='center')
                        
                        other_row_num += 1
                    
                    # En uzun paneli bul (toplam satırları için)
                    max_row = max(axa_row_num, other_row_num) + 2
                    
                    # ============ SOL PANEL TOPLAM SATIRLARI ============
                    # Sütunlar: SİGORTALI(1), TARİH(2), POLİÇE NO(3), TÜRÜ(4), ÖDEME ŞEKLİ(5), TUTAR(6), NET(7), KOMİSYON(8), ALINAN KOMİSYON(9), ORAN(10)
                    # TOPLAM satırı
                    ws.cell(row=max_row, column=4).value = "TOPLAM"
                    ws.cell(row=max_row, column=4).font = toplam_font
                    ws.cell(row=max_row, column=4).fill = toplam_fill
                    
                    # TUTAR, NET, KOMİSYON, ALINAN KOMİSYON değerleri
                    for i, val in enumerate([axa_toplam_brut, axa_toplam_net, axa_toplam_komisyon, axa_toplam_odenen]):
                        cell = ws.cell(row=max_row, column=6 + i)
                        cell.value = val
                        cell.number_format = '#,##0.00'
                        cell.font = toplam_font
                        cell.fill = toplam_fill
                        cell.alignment = Alignment(horizontal='right')
                        cell.border = border
                    
                    # İPTAL TOPLAMI satırı (Sol Panel)
                    iptal_row = max_row + 1
                    ws.cell(row=iptal_row, column=4).value = "İPTAL TOPLAMI"
                    ws.cell(row=iptal_row, column=4).font = Font(bold=True)
                    cell = ws.cell(row=iptal_row, column=9)
                    cell.value = axa_iptal_odenen
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                    cell.border = border
                    
                    # BAKİYE satırı (Sol Panel)
                    bakiye_row = max_row + 2
                    ws.cell(row=bakiye_row, column=4).value = "BAKİYE"
                    ws.cell(row=bakiye_row, column=4).font = bakiye_font
                    ws.cell(row=bakiye_row, column=4).fill = bakiye_fill
                    cell = ws.cell(row=bakiye_row, column=9)
                    cell.value = axa_toplam_odenen - axa_iptal_odenen
                    cell.number_format = '#,##0.00'
                    cell.font = bakiye_font
                    cell.fill = bakiye_fill
                    cell.alignment = Alignment(horizontal='right')
                    cell.border = border
                    
                    # ============ SAĞ PANEL TOPLAM SATIRLARI ============
                    # Sütunlar: SİGORTALI(13), TARİH(14), POLİÇE NO(15), ŞİRKET(16), TÜRÜ(17), ÖDEME ŞEKLİ(18), TUTAR(19), NET(20), KOMİSYON(21), ALINAN KOMİSYON(22), ORAN(23)
                    # TOPLAM satırı
                    ws.cell(row=max_row, column=17).value = "TOPLAM"
                    ws.cell(row=max_row, column=17).font = toplam_font
                    ws.cell(row=max_row, column=17).fill = toplam_fill
                    
                    # TUTAR, NET, KOMİSYON, ALINAN KOMİSYON değerleri
                    for i, val in enumerate([other_toplam_brut, other_toplam_net, other_toplam_komisyon, other_toplam_odenen]):
                        cell = ws.cell(row=max_row, column=19 + i)
                        cell.value = val
                        cell.number_format = '#,##0.00'
                        cell.font = toplam_font
                        cell.fill = toplam_fill
                        cell.alignment = Alignment(horizontal='right')
                        cell.border = border
                    
                    # İPTAL TOPLAMI satırı (Sağ Panel)
                    ws.cell(row=iptal_row, column=17).value = "İPTAL TOPLAMI"
                    ws.cell(row=iptal_row, column=17).font = Font(bold=True)
                    cell = ws.cell(row=iptal_row, column=22)
                    cell.value = other_iptal_odenen
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                    cell.border = border
                    
                    # BAKİYE satırı (Sağ Panel)
                    ws.cell(row=bakiye_row, column=17).value = "BAKİYE"
                    ws.cell(row=bakiye_row, column=17).font = bakiye_font
                    ws.cell(row=bakiye_row, column=17).fill = bakiye_fill
                    cell = ws.cell(row=bakiye_row, column=22)
                    cell.value = other_toplam_odenen - other_iptal_odenen
                    cell.number_format = '#,##0.00'
                    cell.font = bakiye_font
                    cell.fill = bakiye_fill
                    cell.alignment = Alignment(horizontal='right')
                    cell.border = border
                
                # ========== CMC VE DİĞERLERİ İÇİN MEVCUT FORMAT ==========
                else:
                    # Normal ve iptal kayıtları ayır
                    df_normal = df_ay[df_ay['iptal'] == 0]
                    df_iptal = df_ay[df_ay['iptal'] == 1]
                    
                    # Başlık (Satır 1)
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
                    cell = ws.cell(row=1, column=1)
                    cell.value = f"{self.current_kisi} GROUP"
                    cell.font = Font(bold=True, size=12)
                    cell.alignment = Alignment(horizontal='center')
                    
                    # Ay adı (Satır 2)
                    ws.cell(row=2, column=1).value = ay
                    ws.cell(row=2, column=1).font = Font(bold=True, size=11)
                    
                    # Sütun başlıkları (Satır 3)
                    for i, (col_name, width) in enumerate(zip(columns, col_widths)):
                        col_idx = i + 1
                        cell = ws.cell(row=3, column=col_idx)
                        cell.value = col_name
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border
                    
                    # Veri satırları
                    row_num = 4
                    toplam_tutar = 0
                    toplam_net = 0
                    toplam_komisyon = 0
                    toplam_alinan = 0
                    iptal_toplam = 0
                    
                    # Normal kayıtlar
                    for _, row in df_normal.iterrows():
                        tutar = float(row['brut_prim']) if row['brut_prim'] else 0
                        net = float(row['net_prim']) if row['net_prim'] else 0
                        tur = str(row.get('tur', '')).upper()
                        
                        # Komisyon oranı belirle (toplamlar için Python'da hesapla)
                        if 'TRAFİK' in tur:
                            komisyon_hesap = net * 0.10
                        elif 'DASK' in tur:
                            komisyon_hesap = net * 0.0725
                        else:
                            komisyon_hesap = net * 0.15  # CMC için %15
                        
                        alinan_hesap = komisyon_hesap * 0.50
                        
                        toplam_tutar += tutar
                        toplam_net += net
                        toplam_komisyon += komisyon_hesap
                        toplam_alinan += alinan_hesap
                        
                        # Sabit değerler
                        ws.cell(row=row_num, column=1).value = row['sigortali']
                        ws.cell(row=row_num, column=2).value = row['tarih']
                        ws.cell(row=row_num, column=3).value = str(row['police_no'])
                        ws.cell(row=row_num, column=4).value = row['tur']
                        ws.cell(row=row_num, column=5).value = row['odeme_turu']
                        ws.cell(row=row_num, column=6).value = tutar  # TUTAR
                        ws.cell(row=row_num, column=7).value = net    # NET
                        
                        # KOMİSYON formülü: Türe göre dinamik hesaplama (D=tür, G=net)
                        ws.cell(row=row_num, column=8).value = f'=IF(ISNUMBER(SEARCH("TRAFİK",D{row_num})),G{row_num}*0.1,IF(ISNUMBER(SEARCH("DASK",D{row_num})),G{row_num}*0.0725,G{row_num}*0.15))'
                        
                        # ALINAN KOMİSYON formülü: =H{row}*0.5
                        ws.cell(row=row_num, column=9).value = f"=H{row_num}*0.5"
                        
                        ws.cell(row=row_num, column=10).value = oran_text
                        
                        # Stil ve format
                        for i in range(10):
                            cell = ws.cell(row=row_num, column=i + 1)
                            cell.border = border
                            
                            if i == 4:  # ÖDEME ŞEKLİ sütunu
                                apply_odeme_turu_color(cell, ws.cell(row=row_num, column=5).value)
                            if i >= 5 and i <= 8:
                                cell.number_format = '#,##0.00'
                                cell.alignment = Alignment(horizontal='right')
                            elif i == 9:
                                cell.alignment = Alignment(horizontal='center')
                        
                        row_num += 1
                    
                    # İptal kayıtlar
                    for _, row in df_iptal.iterrows():
                        tutar = float(row['brut_prim']) if row['brut_prim'] else 0
                        net = float(row['net_prim']) if row['net_prim'] else 0
                        tur = str(row.get('tur', '')).upper()
                        
                        # Komisyon oranı belirle (toplamlar için Python'da hesapla)
                        if 'TRAFİK' in tur:
                            komisyon_hesap = net * 0.10
                        elif 'DASK' in tur:
                            komisyon_hesap = net * 0.0725
                        else:
                            komisyon_hesap = net * 0.15  # CMC için %15
                        
                        alinan_hesap = komisyon_hesap * 0.50
                        iptal_toplam += alinan_hesap
                        
                        # Sabit değerler
                        ws.cell(row=row_num, column=1).value = row['sigortali']
                        ws.cell(row=row_num, column=2).value = row['tarih']
                        ws.cell(row=row_num, column=3).value = str(row['police_no'])
                        ws.cell(row=row_num, column=4).value = row['tur']
                        ws.cell(row=row_num, column=5).value = row['odeme_turu']
                        ws.cell(row=row_num, column=6).value = tutar  # TUTAR
                        ws.cell(row=row_num, column=7).value = net    # NET
                        
                        # KOMİSYON formülü: Türe göre dinamik hesaplama (D=tür, G=net)
                        ws.cell(row=row_num, column=8).value = f'=IF(ISNUMBER(SEARCH("TRAFİK",D{row_num})),G{row_num}*0.1,IF(ISNUMBER(SEARCH("DASK",D{row_num})),G{row_num}*0.0725,G{row_num}*0.15))'
                        
                        # ALINAN KOMİSYON formülü: =H{row}*0.5
                        ws.cell(row=row_num, column=9).value = f"=H{row_num}*0.5"
                        
                        ws.cell(row=row_num, column=10).value = oran_text
                        
                        # Stil ve format - iptal için kırmızı yazı
                        for i in range(10):
                            cell = ws.cell(row=row_num, column=i + 1)
                            cell.font = iptal_font  # Kırmızı yazı
                            cell.border = border
                            
                            if i == 4:  # ÖDEME ŞEKLİ sütunu (iptal olsa bile)
                                apply_odeme_turu_color(cell, ws.cell(row=row_num, column=5).value)
                            if i >= 5 and i <= 8:
                                cell.number_format = '#,##0.00'
                                cell.alignment = Alignment(horizontal='right')
                            elif i == 9:
                                cell.alignment = Alignment(horizontal='center')
                        
                        row_num += 1
                    
                    # Boş satır
                    row_num += 1
                    
                    # TOPLAM satırı
                    ws.cell(row=row_num, column=5).value = "TOPLAM"
                    ws.cell(row=row_num, column=5).font = toplam_font
                    ws.cell(row=row_num, column=5).fill = toplam_fill
                    
                    for i, val in enumerate([toplam_tutar, toplam_net, toplam_komisyon, toplam_alinan]):
                        cell = ws.cell(row=row_num, column=6 + i)
                        cell.value = val
                        cell.number_format = '#,##0.00'
                        cell.font = toplam_font
                        cell.fill = toplam_fill
                        cell.alignment = Alignment(horizontal='right')
                        cell.border = border
                    
                    row_num += 1
                    
                    # İPTAL TOPLAMI satırı
                    ws.cell(row=row_num, column=5).value = "İPTAL TOPLAMI"
                    ws.cell(row=row_num, column=5).font = Font(bold=True)
                    
                    cell = ws.cell(row=row_num, column=9)
                    cell.value = iptal_toplam
                    cell.number_format = '#,##0.00'
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='right')
                    
                    row_num += 2
                    
                    # BAKİYE satırı
                    bakiye = toplam_alinan - iptal_toplam
                    
                    ws.cell(row=row_num, column=5).value = "BAKİYE"
                    ws.cell(row=row_num, column=5).font = bakiye_font
                    ws.cell(row=row_num, column=5).fill = bakiye_fill
                    
                    cell = ws.cell(row=row_num, column=6)
                    cell.value = bakiye
                    cell.number_format = '#,##0.00'
                    cell.font = bakiye_font
                    cell.fill = bakiye_fill
                    cell.alignment = Alignment(horizontal='right')
                    
                    cell = ws.cell(row=row_num, column=9)
                    cell.value = bakiye
                    cell.number_format = '#,##0.00'
                    cell.font = bakiye_font
                    cell.fill = bakiye_fill
                    cell.alignment = Alignment(horizontal='right')
            
            # Varsayılan boş sheet'i sil
            if sheet_count > 0:
                wb.remove(default_sheet)
            else:
                default_sheet.title = "Veri Yok"
            
            # Dosyayı kaydet
            wb.save(filename)
            
            self.log(f"✅ Excel kaydedildi: {os.path.basename(filename)}", "success")
            messagebox.showinfo("Başarılı", f"Excel dosyası kaydedildi:\n{filename}\n\n{sheet_count} ay için sayfa oluşturuldu.")
        
        except Exception as e:
            self.log(f"❌ Excel hatası: {str(e)}", "error")
            messagebox.showerror("Hata", str(e))
    
    def import_excel_data(self):
        """Excel dosyasından veri içe aktar - Export formatına uyumlu"""
        filename = filedialog.askopenfilename(
            title="Excel Dosyası Seç",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if not filename:
            return
        
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(filename, data_only=True)
            total_imported = 0
            total_skipped = 0
            
            # Kırmızı font kontrolü için yardımcı fonksiyon
            def is_red_font(cell):
                """Hücrenin font renginin kırmızı olup olmadığını kontrol et"""
                try:
                    if cell.font and cell.font.color:
                        color = cell.font.color
                        # RGB formatı (FF0000 = kırmızı)
                        if color.rgb and isinstance(color.rgb, str):
                            rgb = color.rgb.upper()
                            # FFFF0000 veya FF0000 formatı (kırmızı)
                            if rgb in ['FFFF0000', 'FF0000'] or rgb.endswith('FF0000'):
                                return True
                            # Kırmızı tonları
                            if len(rgb) >= 6:
                                r_hex = rgb[-6:-4] if len(rgb) >= 6 else '00'
                                g_hex = rgb[-4:-2] if len(rgb) >= 4 else '00'
                                b_hex = rgb[-2:] if len(rgb) >= 2 else '00'
                                try:
                                    r = int(r_hex, 16)
                                    g = int(g_hex, 16)
                                    b = int(b_hex, 16)
                                    # Kırmızı bileşen yüksek, diğerleri düşük
                                    if r > 200 and g < 100 and b < 100:
                                        return True
                                except:
                                    pass
                except:
                    pass
                return False
            
            # Her sheet için işle (aylar)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Sheet'teki tüm satırları oku - format bilgisiyle birlikte
                rows_with_format = list(ws.iter_rows(values_only=False))
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 4:  # En az başlık ve bir veri satırı olmalı
                    continue
                
                # İlk satırdan formatı anla
                first_row = rows[0]
                first_cell = str(first_row[0]) if first_row[0] else ""
                
                # TEZER/KAMİL formatı mı kontrol et (iki panelli)
                is_dual_panel = "AXA'DAN YAPILAN" in first_cell.upper() if first_cell else False
                
                if is_dual_panel:
                    # İki panelli format (TEZER/KAMİL)
                    # Satır 4'ten başla (0-indexed: 3)
                    # Sol panel: A-J (0-9), Sağ panel: M-W (12-22)
                    
                    for row_idx, row in enumerate(rows[4:], start=4):  # 0-indexed: 4. satırdan başla
                        # Format satırı al (iptal kontrolü için)
                        format_row = rows_with_format[row_idx] if row_idx < len(rows_with_format) else None
                        
                        # TOPLAM, İPTAL TOPLAMI, BAKİYE satırlarını atla
                        first_val = str(row[0]) if row[0] else ""
                        fourth_val = str(row[3]) if len(row) > 3 and row[3] else ""
                        
                        if any(x in first_val.upper() for x in ['TOPLAM', 'BAKİYE', 'İPTAL', 'DEVREDEN']):
                            continue
                        if any(x in fourth_val.upper() for x in ['TOPLAM', 'BAKİYE', 'İPTAL']):
                            continue
                        
                        # Sol panel (AXA) - Sütunlar: 0-9
                        # SİGORTALI(0), TARİH(1), POLİÇE NO(2), TÜRÜ(3), ÖDEME ŞEKLİ(4), TUTAR(5), NET(6), KOMİSYON(7), ALINAN KOMİSYON(8), ORAN(9)
                        if row[0] and str(row[0]).strip() and str(row[0]) != 'nan':
                            sigortali = str(row[0]).strip()
                            # Boş veya başlık satırlarını atla
                            if sigortali and len(sigortali) > 2 and sigortali.upper() not in ['SİGORTALI', 'SIGORTALI']:
                                try:
                                    tarih = str(row[1]) if row[1] else ''
                                    police_no = str(row[2]).replace('.0', '') if row[2] else ''
                                    tur = str(row[3]) if row[3] else 'TRAFİK'
                                    odeme_turu = str(row[4]) if row[4] else 'K.KART'
                                    
                                    # Sayısal değerler
                                    def parse_num(val):
                                        if val is None or str(val) == 'nan':
                                            return 0
                                        try:
                                            if isinstance(val, (int, float)):
                                                return float(val)
                                            return float(str(val).replace('.', '').replace(',', '.'))
                                        except:
                                            return 0
                                    
                                    brut = parse_num(row[5])
                                    net = parse_num(row[6])
                                    komisyon = parse_num(row[7])
                                    alinan = parse_num(row[8])
                                    
                                    if brut == 0 and net == 0:
                                        continue
                                    
                                    # Net prim 0 ise brüt kullan
                                    effective_net = net if net > 0 else brut
                                    
                                    # Komisyon oranı belirle
                                    tur_clean = tur if tur != 'nan' else 'TRAFİK'
                                    komisyon_orani = hesapla_komisyon(self.current_kisi, tur_clean, 1)['komisyon_orani']
                                    odeme_orani = 0.60 if self.current_kisi == 'YAŞAR' else 0.50
                                    
                                    # Komisyon/ödenen 0 ise (formül cache yok) otomatik hesapla
                                    if komisyon == 0 and effective_net > 0:
                                        komisyon = effective_net * komisyon_orani
                                    if alinan == 0 and komisyon > 0:
                                        alinan = komisyon * odeme_orani
                                    
                                    # Yıl çıkar
                                    yil = datetime.now().year
                                    if tarih and ('/' in tarih or '.' in tarih):
                                        try:
                                            parts = tarih.replace('.', '/').split('/')
                                            if len(parts) >= 3:
                                                yil = int(parts[2]) if len(parts[2]) == 4 else 2000 + int(parts[2])
                                        except:
                                            pass
                                    
                                    # İptal kontrolü - font rengi veya negatif değer
                                    is_iptal = 0
                                    if format_row and len(format_row) > 0:
                                        # İlk hücrenin font rengini kontrol et
                                        if is_red_font(format_row[0]):
                                            is_iptal = 1
                                    # Negatif değerler de iptal olabilir
                                    if brut < 0 or net < 0:
                                        is_iptal = 1
                                    
                                    record = {
                                        'kisi': self.current_kisi,
                                        'sirket': 'AXA',
                                        'ay': '-',
                                        'yil': yil,
                                        'sigortali': sigortali,
                                        'police_no': police_no,
                                        'tarih': tarih,
                                        'plaka': '-',
                                        'tur': tur_clean,
                                        'odeme_turu': odeme_turu if odeme_turu != 'nan' else 'K.KART',
                                        'brut_prim': brut,
                                        'net_prim': effective_net,
                                        'komisyon_orani': komisyon_orani,
                                        'toplam_komisyon': komisyon,
                                        'odeme_orani': odeme_orani,
                                        'odenen_komisyon': alinan,
                                        'iptal': is_iptal,
                                        'notlar': f'Excel import: {sheet_name}'
                                    }
                                    
                                    if self.db.insert_record(record):
                                        total_imported += 1
                                    else:
                                        total_skipped += 1
                                except Exception as e:
                                    print(f"Sol panel satır hatası: {e}")
                                    continue
                        
                        # Sağ panel (Diğer Şirketler) - Sütunlar: 12-22 (M-W)
                        # SİGORTALI(12), TARİH(13), POLİÇE NO(14), ŞİRKET(15), TÜRÜ(16), ÖDEME ŞEKLİ(17), TUTAR(18), NET(19), KOMİSYON(20), ALINAN KOMİSYON(21), ORAN(22)
                        if len(row) > 12 and row[12] and str(row[12]).strip() and str(row[12]) != 'nan':
                            sigortali = str(row[12]).strip()
                            if sigortali and len(sigortali) > 2 and sigortali.upper() not in ['SİGORTALI', 'SIGORTALI']:
                                try:
                                    tarih = str(row[13]) if len(row) > 13 and row[13] else ''
                                    police_no = str(row[14]).replace('.0', '') if len(row) > 14 and row[14] else ''
                                    sirket = str(row[15]) if len(row) > 15 and row[15] else '-'
                                    tur = str(row[16]) if len(row) > 16 and row[16] else 'TRAFİK'
                                    odeme_turu = str(row[17]) if len(row) > 17 and row[17] else 'K.KART'
                                    
                                    def parse_num(val):
                                        if val is None or str(val) == 'nan':
                                            return 0
                                        try:
                                            if isinstance(val, (int, float)):
                                                return float(val)
                                            return float(str(val).replace('.', '').replace(',', '.'))
                                        except:
                                            return 0
                                    
                                    brut = parse_num(row[18]) if len(row) > 18 else 0
                                    net = parse_num(row[19]) if len(row) > 19 else 0
                                    komisyon = parse_num(row[20]) if len(row) > 20 else 0
                                    alinan = parse_num(row[21]) if len(row) > 21 else 0
                                    
                                    if brut == 0 and net == 0:
                                        continue
                                    
                                    # Net prim 0 ise brüt kullan
                                    effective_net = net if net > 0 else brut
                                    
                                    # Komisyon oranı belirle
                                    tur_clean = tur if tur != 'nan' else 'TRAFİK'
                                    komisyon_orani = hesapla_komisyon(self.current_kisi, tur_clean, 1)['komisyon_orani']
                                    odeme_orani = 0.60 if self.current_kisi == 'YAŞAR' else 0.50
                                    
                                    # Komisyon/ödenen 0 ise (formül cache yok) otomatik hesapla
                                    if komisyon == 0 and effective_net > 0:
                                        komisyon = effective_net * komisyon_orani
                                    if alinan == 0 and komisyon > 0:
                                        alinan = komisyon * odeme_orani
                                    
                                    # Yıl çıkar
                                    yil = datetime.now().year
                                    if tarih and ('/' in tarih or '.' in tarih):
                                        try:
                                            parts = tarih.replace('.', '/').split('/')
                                            if len(parts) >= 3:
                                                yil = int(parts[2]) if len(parts[2]) == 4 else 2000 + int(parts[2])
                                        except:
                                            pass
                                    
                                    # İptal kontrolü - font rengi veya negatif değer
                                    is_iptal = 0
                                    if format_row and len(format_row) > 12:
                                        # Sağ panelin ilk hücresinin (12. sütun) font rengini kontrol et
                                        if is_red_font(format_row[12]):
                                            is_iptal = 1
                                    # Negatif değerler de iptal olabilir
                                    if brut < 0 or net < 0:
                                        is_iptal = 1
                                    
                                    record = {
                                        'kisi': self.current_kisi,
                                        'sirket': sirket if sirket != 'nan' else '-',
                                        'ay': '-',
                                        'yil': yil,
                                        'sigortali': sigortali,
                                        'police_no': police_no,
                                        'tarih': tarih,
                                        'plaka': '-',
                                        'tur': tur_clean,
                                        'odeme_turu': odeme_turu if odeme_turu != 'nan' else 'K.KART',
                                        'brut_prim': brut,
                                        'net_prim': effective_net,
                                        'komisyon_orani': komisyon_orani,
                                        'toplam_komisyon': komisyon,
                                        'odeme_orani': odeme_orani,
                                        'odenen_komisyon': alinan,
                                        'iptal': is_iptal,
                                        'notlar': f'Excel import: {sheet_name}'
                                    }
                                    
                                    if self.db.insert_record(record):
                                        total_imported += 1
                                    else:
                                        total_skipped += 1
                                except Exception as e:
                                    print(f"Sağ panel satır hatası: {e}")
                                    continue
                
                else:
                    # Tek panelli format (YAŞAR, CMC veya eski format)
                    # Başlık satırını bul
                    header_row = None
                    header_idx = 0
                    for idx, row in enumerate(rows):
                        if row[0] and 'SİGORTALI' in str(row[0]).upper():
                            header_row = row
                            header_idx = idx
                            break
                    
                    if not header_row:
                        # Başlık bulunamadı, 3. satırı dene
                        if len(rows) > 2:
                            header_row = rows[2]
                            header_idx = 2
                        else:
                            continue
                    
                    # Sütun indekslerini bul
                    col_map = {}
                    for idx, col in enumerate(header_row):
                        if col:
                            col_name = str(col).upper().strip()
                            col_map[col_name] = idx
                    
                    # Veri satırlarını işle
                    for row_idx, row in enumerate(rows[header_idx + 1:], start=header_idx + 1):
                        # Format satırı al (iptal kontrolü için)
                        format_row = rows_with_format[row_idx] if row_idx < len(rows_with_format) else None
                        
                        try:
                            # Boş satırları ve özet satırlarını atla
                            first_val = str(row[0]) if row[0] else ""
                            if not first_val or first_val == 'nan':
                                continue
                            if any(x in first_val.upper() for x in ['TOPLAM', 'BAKİYE', 'İPTAL', 'KASKO', 'TRAFİK', 'DİĞER', 'AÇIK', 'ALINAN', 'YAŞAR', 'CARİ', 'NİHAT']):
                                continue
                            
                            # Değerleri al
                            def get_val(names):
                                for name in names:
                                    if name in col_map:
                                        val = row[col_map[name]]
                                        if val and str(val) != 'nan':
                                            return val
                                return None
                            
                            sigortali = str(get_val(['SİGORTALI', 'SIGORTALI']) or '')
                            if not sigortali or len(sigortali) < 2:
                                continue
                            
                            tarih = str(get_val(['TARİH', 'TARIH']) or '')
                            police_no = str(get_val(['POL. NO.', 'POLİÇE NO', 'POLICE NO']) or '').replace('.0', '')
                            tur = str(get_val(['TÜRÜ', 'TÜR', 'TUR']) or 'TRAFİK')
                            odeme_turu = str(get_val(['ÖDEME TÜRÜ', 'ÖDEME ŞEKLİ', 'ODEME TURU']) or 'K.KART')
                            
                            def parse_num(val):
                                if val is None or str(val) == 'nan':
                                    return 0
                                try:
                                    if isinstance(val, (int, float)):
                                        return float(val)
                                    return float(str(val).replace('.', '').replace(',', '.'))
                                except:
                                    return 0
                            
                            brut = parse_num(get_val(['TUTAR', 'BRÜT PRİM', 'BRUT PRIM']))
                            net = parse_num(get_val(['NET', 'NET PRİM', 'NET PRIM']))
                            komisyon = parse_num(get_val(['KOMİSYON', 'KOMISYON', 'TOPLAM KOMİSYON']))
                            alinan = parse_num(get_val(['ALINAN KOMİSYON', 'ÖDENEN KOMİSYON', 'ODENEN KOMISYON']))
                            
                            if brut == 0 and net == 0:
                                continue
                            
                            # Net prim 0 ise brüt kullan
                            effective_net = net if net > 0 else brut
                            
                            # Komisyon oranı belirle
                            tur_clean = tur if tur != 'nan' else 'TRAFİK'
                            komisyon_orani = hesapla_komisyon(self.current_kisi, tur_clean, 1)['komisyon_orani']
                            odeme_orani = 0.60 if self.current_kisi == 'YAŞAR' else 0.50
                            
                            # Komisyon/ödenen 0 ise (formül cache yok) otomatik hesapla
                            if komisyon == 0 and effective_net > 0:
                                komisyon = effective_net * komisyon_orani
                            if alinan == 0 and komisyon > 0:
                                alinan = komisyon * odeme_orani
                            
                            # Yıl çıkar
                            yil = datetime.now().year
                            if tarih and ('/' in tarih or '.' in tarih):
                                try:
                                    parts = tarih.replace('.', '/').split('/')
                                    if len(parts) >= 3:
                                        yil = int(parts[2]) if len(parts[2]) == 4 else 2000 + int(parts[2])
                                except:
                                    pass
                            
                            # İptal kontrolü - font rengi veya negatif değer
                            is_iptal = 0
                            if format_row and len(format_row) > 0:
                                # İlk hücrenin font rengini kontrol et
                                if is_red_font(format_row[0]):
                                    is_iptal = 1
                            # Negatif değerler de iptal olabilir
                            if brut < 0 or net < 0:
                                is_iptal = 1
                            
                            record = {
                                'kisi': self.current_kisi,
                                'sirket': 'AXA',
                                'ay': '-',
                                'yil': yil,
                                'sigortali': sigortali,
                                'police_no': police_no,
                                'tarih': tarih,
                                'plaka': str(get_val(['PLAKA']) or '-'),
                                'tur': tur_clean,
                                'odeme_turu': odeme_turu if odeme_turu != 'nan' else 'K.KART',
                                'brut_prim': brut,
                                'net_prim': effective_net,
                                'komisyon_orani': komisyon_orani,
                                'toplam_komisyon': komisyon,
                                'odeme_orani': odeme_orani,
                                'odenen_komisyon': alinan,
                                'iptal': is_iptal,
                                'notlar': f'Excel import: {sheet_name}'
                            }
                            
                            if self.db.insert_record(record):
                                total_imported += 1
                            else:
                                total_skipped += 1
                        
                        except Exception as e:
                            print(f"Satır hatası: {e}")
                            continue
            
            msg = f"✅ İçe aktarıldı: {total_imported}\n"
            if total_skipped > 0:
                msg += f"⚠️ Atlandı (duplicate): {total_skipped}"
            
            messagebox.showinfo("İçe Aktarma Tamamlandı", msg)
            self.log(f"✅ Excel'den {total_imported} kayıt içe aktarıldı", "success")
            self.load_data()
            self.update_summary()
        
        except Exception as e:
            self.log(f"❌ Excel okuma hatası: {str(e)}", "error")
            messagebox.showerror("Hata", str(e))
    
    def log(self, message, level="info"):
        """Log ekle"""
        timestamp = datetime.now().strftime('%H:%M:%S')
        log_entry = f"[{timestamp}] {message}\n"
        self.log_text.insert(tk.END, log_entry, level)
        self.log_text.see(tk.END)
        self.status_bar.config(text=message)
    
    def clear_logs(self):
        """Logları temizle"""
        self.log_text.delete('1.0', tk.END)
    
    def show_add_policy_dialog(self, table_type='axa'):
        """Yeni poliçe eklemek için popup dialog göster"""
        # Dialog penceresi oluştur
        dialog = tk.Toplevel(self.root)
        dialog.title("➕ Yeni Poliçe Ekle")
        
        # Boyutu büyüt - 'other' için daha büyük
        if table_type == 'other':
            dialog_width = 450
            dialog_height = 620
        else:
            dialog_width = 450
            dialog_height = 580
        
        dialog.geometry(f"{dialog_width}x{dialog_height}")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Pencereyi ortala
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() - dialog_width) // 2
        y = (dialog.winfo_screenheight() - dialog_height) // 2
        dialog.geometry(f"+{x}+{y}")
        
        # Ana container - scrollable
        canvas = tk.Canvas(dialog, bg='#f8fafc', highlightthickness=0)
        scrollbar = ttk.Scrollbar(dialog, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#f8fafc')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Mouse wheel scroll
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", on_mousewheel)
        
        canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y")
        
        # Ana container içinde form
        main_frame = tk.Frame(scrollable_frame, bg='#f8fafc', padx=10, pady=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Başlık
        title_label = tk.Label(
            main_frame,
            text="📋 Poliçe Bilgileri",
            font=('Segoe UI', 14, 'bold'),
            bg='#f8fafc',
            fg='#1e3a5f'
        )
        title_label.pack(pady=(0, 15))
        
        # Form frame
        form_frame = tk.Frame(main_frame, bg='#f8fafc')
        form_frame.pack(fill=tk.BOTH, expand=True)
        
        # Değişkenler için dictionary
        entries = {}
        
        # Şirket (her iki tablo için de göster)
        tk.Label(form_frame, text="Şirket:", font=('Segoe UI', 10), bg='#f8fafc', anchor='w').pack(fill=tk.X, pady=(5, 2))
        sirket_container = tk.Frame(form_frame, bg='#f8fafc')
        sirket_container.pack(fill=tk.X, pady=(0, 5))
        sirketler = ['AXA', 'ALLIANZ', 'ANADOLU', 'GÜNEŞ', 'MAPFRE', 'QUICK', 'HDI', 'DOĞA', 'SOMPO', 'RAY', 'ETHICA', 'HEPİYİ', 'DİĞER']
        entries['sirket'] = ttk.Combobox(sirket_container, values=sirketler, font=('Segoe UI', 10), state='readonly')
        if table_type == 'axa':
            entries['sirket'].set('AXA')
        else:
            entries['sirket'].set('')
        entries['sirket'].pack(fill=tk.X)
        
        # DİĞER seçildiğinde elle giriş aç
        def on_sirket_diger(e=None):
            if entries['sirket'].get() == 'DİĞER':
                entries['sirket'].destroy()
                entries['sirket'] = tk.Entry(sirket_container, font=('Segoe UI', 10))
                entries['sirket'].pack(fill=tk.X)
                entries['sirket'].focus_set()
        entries['sirket'].bind('<<ComboboxSelected>>', on_sirket_diger)
        
        # Sigortalı
        tk.Label(form_frame, text="Sigortalı:", font=('Segoe UI', 10), bg='#f8fafc', anchor='w').pack(fill=tk.X, pady=(5, 2))
        entries['sigortali'] = tk.Entry(form_frame, font=('Segoe UI', 10))
        entries['sigortali'].pack(fill=tk.X, pady=(0, 5))
        
        # Tarih - seçili aya göre varsayılan tarih
        tk.Label(form_frame, text="Tarih (GG/AA/YYYY):", font=('Segoe UI', 10), bg='#f8fafc', anchor='w').pack(fill=tk.X, pady=(5, 2))
        entries['tarih'] = tk.Entry(form_frame, font=('Segoe UI', 10))
        # Seçili aya göre varsayılan tarih oluştur
        ay_index = AYLAR.index(self.current_ay) + 1  # 1-12
        default_tarih = f"01/{ay_index:02d}/{datetime.now().year}"
        entries['tarih'].insert(0, default_tarih)
        entries['tarih'].pack(fill=tk.X, pady=(0, 5))
        
        # Poliçe No
        tk.Label(form_frame, text="Poliçe No:", font=('Segoe UI', 10), bg='#f8fafc', anchor='w').pack(fill=tk.X, pady=(5, 2))
        entries['police_no'] = tk.Entry(form_frame, font=('Segoe UI', 10))
        entries['police_no'].pack(fill=tk.X, pady=(0, 5))
        
        # Tür
        tk.Label(form_frame, text="Tür:", font=('Segoe UI', 10), bg='#f8fafc', anchor='w').pack(fill=tk.X, pady=(5, 2))
        entries['tur'] = ttk.Combobox(form_frame, values=TURLER, font=('Segoe UI', 10), state='readonly')
        entries['tur'].set(TURLER[0])
        entries['tur'].pack(fill=tk.X, pady=(0, 5))
        
        # Ödeme Türü
        tk.Label(form_frame, text="Ödeme Türü:", font=('Segoe UI', 10), bg='#f8fafc', anchor='w').pack(fill=tk.X, pady=(5, 2))
        entries['odeme'] = ttk.Combobox(form_frame, values=ODEME_TURLERI, font=('Segoe UI', 10), state='readonly')
        entries['odeme'].set('K.KART')
        entries['odeme'].pack(fill=tk.X, pady=(0, 5))
        
        # Brüt Prim
        tk.Label(form_frame, text="Brüt Prim (TL):", font=('Segoe UI', 10), bg='#f8fafc', anchor='w').pack(fill=tk.X, pady=(5, 2))
        entries['brut_prim'] = tk.Entry(form_frame, font=('Segoe UI', 10))
        entries['brut_prim'].pack(fill=tk.X, pady=(0, 5))
        
        # Net Prim
        tk.Label(form_frame, text="Net Prim (TL):", font=('Segoe UI', 10), bg='#f8fafc', anchor='w').pack(fill=tk.X, pady=(5, 2))
        entries['net_prim'] = tk.Entry(form_frame, font=('Segoe UI', 10))
        entries['net_prim'].pack(fill=tk.X, pady=(0, 5))
        
        # İptal checkbox
        iptal_var = tk.BooleanVar(value=False)
        iptal_cb = tk.Checkbutton(
            form_frame,
            text="🔴 İptal Poliçesi",
            variable=iptal_var,
            font=('Segoe UI', 10),
            bg='#f8fafc',
            fg='#dc2626',
            activebackground='#f8fafc'
        )
        iptal_cb.pack(anchor='w', pady=(10, 5))
        
        # Butonlar frame
        btn_frame = tk.Frame(main_frame, bg='#f8fafc')
        btn_frame.pack(fill=tk.X, pady=(15, 0))
        
        def save_policy():
            """Poliçeyi kaydet"""
            # Değerleri al
            sigortali = entries['sigortali'].get().strip()
            tarih = entries['tarih'].get().strip()
            police_no = entries['police_no'].get().strip()
            tur = entries['tur'].get()
            odeme = entries['odeme'].get()
            brut_prim_str = entries['brut_prim'].get().strip()
            net_prim_str = entries['net_prim'].get().strip()
            iptal = 1 if iptal_var.get() else 0
            
            # Şirket - entries'den oku, tabloya göre yönlendir
            sirket = entries['sirket'].get().strip().upper()
            if not sirket:
                sirket = 'AXA' if table_type == 'axa' else 'ALLIANZ'
            
            # Brüt prim kontrolü
            try:
                brut_prim = float(brut_prim_str.replace('.', '').replace(',', '.'))
            except:
                brut_prim = 0
            
            # Net prim kontrolü
            try:
                net_prim = float(net_prim_str.replace('.', '').replace(',', '.'))
            except:
                net_prim = brut_prim  # Net girilmezse brüt kullan
            
            # Poliçe no kontrolü
            if not police_no:
                police_no = f"MNL-{datetime.now().strftime('%Y%m%d%H%M%S')}"
            
            # Yıl çıkar
            yil = datetime.now().year
            if tarih and '/' in tarih:
                try:
                    yil = int(tarih.split('/')[2])
                except:
                    pass
            
            # Komisyon hesapla - NET PRİM üzerinden (mevcut sistemdeki gibi)
            if tur and net_prim > 0:
                komisyon = hesapla_komisyon(self.current_kisi, tur, net_prim)
            else:
                komisyon = {
                    'komisyon_orani': 0,
                    'komisyon': 0,
                    'odeme_orani': 0,
                    'odenen': 0
                }
            
            # Kayıt oluştur - İptal poliçelerinde değerler pozitif kalır, sadece iptal=1
            record = {
                'kisi': self.current_kisi,
                'ay': self.current_ay,
                'yil': yil,
                'sigortali': sigortali or '-',
                'police_no': police_no,
                'tarih': tarih or datetime.now().strftime('%d/%m/%Y'),
                'plaka': '-',
                'tur': tur or '-',
                'odeme_turu': odeme,
                'brut_prim': brut_prim,
                'net_prim': net_prim,
                'komisyon_orani': komisyon['komisyon_orani'],
                'toplam_komisyon': komisyon['komisyon'],
                'odeme_orani': komisyon['odeme_orani'],
                'odenen_komisyon': komisyon['odenen'],
                'iptal': iptal,
                'notlar': '',
                'sirket': sirket
            }
            
            if self.db.insert_record(record):
                iptal_text = " 🔴 İPTAL" if iptal else ""
                self.log(f"✅ Yeni poliçe eklendi: {sigortali} - {tur}{iptal_text}", "success")
                # Mouse wheel binding'i kaldır
                canvas.unbind_all("<MouseWheel>")
                dialog.destroy()
                self.load_data()
                self.update_summary()
            else:
                messagebox.showerror("Hata", "Kayıt eklenemedi! (Poliçe no zaten mevcut olabilir)")
        
        def cancel():
            """İptal et"""
            # Mouse wheel binding'i kaldır
            canvas.unbind_all("<MouseWheel>")
            dialog.destroy()
            self.log("❌ Poliçe ekleme iptal edildi", "warning")
        
        # Kaydet butonu
        save_btn = tk.Button(
            btn_frame,
            text="💾 Kaydet",
            command=save_policy,
            font=('Segoe UI', 11, 'bold'),
            bg='#16a34a',
            fg='white',
            relief=tk.FLAT,
            cursor='hand2',
            padx=20,
            pady=8
        )
        save_btn.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(0, 5))
        
        # İptal butonu
        cancel_btn = tk.Button(
            btn_frame,
            text="❌ İptal",
            command=cancel,
            font=('Segoe UI', 11, 'bold'),
            bg='#dc2626',
            fg='white',
            relief=tk.FLAT,
            cursor='hand2',
            padx=20,
            pady=8
        )
        cancel_btn.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(5, 0))
        
        # Enter ile kaydet
        dialog.bind('<Return>', lambda e: save_policy())
        dialog.bind('<Escape>', lambda e: cancel())
        
        # İlk alana focus
        if table_type == 'other':
            entries['sirket'].focus_set()
        else:
            entries['sigortali'].focus_set()
    
    def show_about(self):
        """Hakkında"""
        about_text = """
🏢 AXA Muhasebe ve Komisyon Takip Sistemi
Versiyon: 1.0

✨ Özellikler:
• Otomatik PDF okuma ve veri çıkarma
• Kişi bazlı komisyon takibi
• Otomatik komisyon hesaplama
• Aylık Excel raporlama
• Modern kullanıcı arayüzü

📋 Komisyon Oranları:
• Trafik: %10 (tüm kişiler)
• Diğer (Yaşar/Kamil): %15
• Diğer (Tezer): %13

💰 Ödeme Oranları:
• Yaşar: %60
• Kamil/Tezer: %50
        """
        messagebox.showinfo("Hakkında", about_text)
    
    def show_cari_window(self):
        """Cari verileri görüntülemek için ayrı pencere"""
        # Yeni pencere oluştur
        cari_win = tk.Toplevel(self.root)
        cari_win.title("💳 Cari Takip")
        cari_win.geometry("700x500")
        cari_win.configure(bg='#f0f0f0')
        
        # PDF yollarını saklamak için dictionary (item_id -> pdf_path)
        pdf_paths = {}
        
        # Başlık
        header = tk.Frame(cari_win, bg=COLORS['primary'], height=40)
        header.pack(fill='x')
        header.pack_propagate(False)
        tk.Label(header, text="💳 CARİ TAKİP", font=('Arial', 14, 'bold'), 
                 bg=COLORS['primary'], fg='white').pack(pady=8)
        
        # Ay seçimi
        filter_frame = tk.Frame(cari_win, bg='#f0f0f0')
        filter_frame.pack(fill='x', padx=10, pady=10)
        
        tk.Label(filter_frame, text="Ay:", font=('Arial', 10), bg='#f0f0f0').pack(side='left', padx=5)
        ay_combo = ttk.Combobox(filter_frame, values=AYLAR, state='readonly', width=12)
        ay_combo.pack(side='left', padx=5)
        ay_combo.set(self.current_ay)
        
        # Tablo
        table_frame = tk.Frame(cari_win)
        table_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        columns = ('İsim', 'Tutar', 'Tarih')
        tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=15)
        
        tree.heading('İsim', text='İSİM')
        tree.heading('Tutar', text='TUTAR')
        tree.heading('Tarih', text='TARİH')
        
        tree.column('İsim', width=250, anchor='w')
        tree.column('Tutar', width=150, anchor='e')
        tree.column('Tarih', width=120, anchor='center')
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(table_frame, orient='vertical', command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Toplam label
        toplam_frame = tk.Frame(cari_win, bg='#f0f0f0')
        toplam_frame.pack(fill='x', padx=10, pady=5)
        
        toplam_label = tk.Label(toplam_frame, text="TOPLAM: 0,00 TL", 
                               font=('Arial', 12, 'bold'), bg='#f0f0f0', fg=COLORS['success'])
        toplam_label.pack(side='right', padx=20)
        
        # Butonlar
        btn_frame = tk.Frame(cari_win, bg='#f0f0f0')
        btn_frame.pack(fill='x', padx=10, pady=10)
        
        def load_cari_data():
            """Cari verilerini yükle"""
            nonlocal pdf_paths
            pdf_paths = {}
            
            # Tabloyu temizle
            for item in tree.get_children():
                tree.delete(item)
            
            # Seçili ay numarasını al
            selected_ay = ay_combo.get()
            ay_numarasi = AYLAR.index(selected_ay) + 1 if selected_ay in AYLAR else 1
            
            # Cari klasöründen verileri çek
            script_dir = get_application_path()
            cari_folder = os.path.join(script_dir, 'cari')
            cari_data = get_cari_data_for_month(cari_folder, ay_numarasi)
            
            toplam = 0
            for i, cari in enumerate(cari_data):
                item_id = f"cari_{i}"
                tree.insert('', 'end', iid=item_id, values=(
                    cari['isim'],
                    format_turkish_currency(cari['tutar']),
                    cari['tarih']
                ))
                # PDF yolunu sakla
                if 'pdf_path' in cari:
                    pdf_paths[item_id] = cari['pdf_path']
                toplam += cari['tutar']
            
            toplam_label.config(text=f"TOPLAM: {format_turkish_currency(toplam)} TL")
        
        def delete_selected_cari():
            """Seçili cari kayıtları sil (PDF dosyalarını silerek)"""
            selection = tree.selection()
            if not selection:
                messagebox.showwarning("Uyarı", "Lütfen silmek istediğiniz kayıtları seçin!")
                return
            
            if messagebox.askyesno("Onay", f"{len(selection)} kayıt silinecek. Emin misiniz?"):
                deleted = 0
                for item_id in selection:
                    if item_id in pdf_paths:
                        pdf_path = pdf_paths[item_id]
                        try:
                            if os.path.exists(pdf_path):
                                os.remove(pdf_path)
                                deleted += 1
                        except Exception as e:
                            print(f"PDF silme hatası: {e}")
                
                messagebox.showinfo("Başarılı", f"{deleted} kayıt silindi!")
                load_cari_data()
        
        def open_selected_pdf():
            """Seçili kayıdın PDF'ini aç"""
            selection = tree.selection()
            if not selection:
                messagebox.showwarning("Uyarı", "Lütfen bir kayıt seçin!")
                return
            
            item_id = selection[0]
            if item_id in pdf_paths:
                pdf_path = pdf_paths[item_id]
                if os.path.exists(pdf_path):
                    os.startfile(pdf_path)
                else:
                    messagebox.showerror("Hata", "PDF dosyası bulunamadı!")
            else:
                messagebox.showerror("Hata", "PDF dosya yolu bulunamadı!")
        
        def show_cari_context_menu(event):
            """Sağ tık menüsü göster"""
            # Tıklanan öğeyi seç
            item = tree.identify_row(event.y)
            if item:
                tree.selection_set(item)
                cari_context_menu.tk_popup(event.x_root, event.y_root)
        
        # Sağ tık menüsü oluştur
        cari_context_menu = tk.Menu(cari_win, tearoff=0)
        cari_context_menu.add_command(label="📂 PDF'i Aç", command=open_selected_pdf)
        cari_context_menu.add_separator()
        cari_context_menu.add_command(label="🗑️ Sil", command=delete_selected_cari)
        
        # Sağ tık olayını bağla
        tree.bind("<Button-3>", show_cari_context_menu)
        
        # Çift tık ile PDF'i aç
        tree.bind("<Double-1>", lambda e: open_selected_pdf())
        
        def add_cari_pdf():
            """Cari PDF ekle"""
            files = filedialog.askopenfilenames(
                title="Cari PDF Seçin",
                filetypes=[("PDF files", "*.pdf")]
            )
            if files:
                script_dir = get_application_path()
                cari_folder = os.path.join(script_dir, 'cari')
                
                # Klasör yoksa oluştur
                if not os.path.exists(cari_folder):
                    os.makedirs(cari_folder)
                
                import shutil
                for f in files:
                    dest = os.path.join(cari_folder, os.path.basename(f))
                    shutil.copy2(f, dest)
                
                messagebox.showinfo("Başarılı", f"{len(files)} PDF dosyası eklendi!")
                load_cari_data()
        
        tk.Button(btn_frame, text="🔄 Yenile", command=load_cari_data,
                 font=('Arial', 10), bg=COLORS['primary'], fg='white', width=12).pack(side='left', padx=5)
        
        tk.Button(btn_frame, text="➕ PDF Ekle", command=add_cari_pdf,
                 font=('Arial', 10), bg=COLORS['success'], fg='white', width=12).pack(side='left', padx=5)
        
        tk.Button(btn_frame, text="🗑️ Seçili Sil", command=delete_selected_cari,
                 font=('Arial', 10), bg=COLORS['warning'], fg='white', width=12).pack(side='left', padx=5)
        
        tk.Button(btn_frame, text="❌ Kapat", command=cari_win.destroy,
                 font=('Arial', 10), bg=COLORS['danger'], fg='white', width=12).pack(side='right', padx=5)
        
        # Ay değiştiğinde verileri yenile
        ay_combo.bind('<<ComboboxSelected>>', lambda e: load_cari_data())
        
        # İlk yükleme
        load_cari_data()


# ==============================================================================
# ANA PROGRAM
# ==============================================================================

if __name__ == "__main__":
    root = tk.Tk()
    app = MuhasebeGUI(root)
    root.mainloop()
